import sys
try:
    import openpyxl
    print("openpyxl is available")
    
    file_path = r"c:/Users/sdnse/OneDrive/Dokumen/Yum/PROJEK SUDIN JU2/ASKA SUDIN JU 2/sudin_aska/.Data_Sekolah_SudinJU2/DAFTAR SEKOLAH NEGERI & SWASTA DI JAKARTA UTARA II.xlsx"
    wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
    ws = wb.active
    
    print("Sheet names:", wb.sheetnames)
    print("First few rows:")
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        print(row)
        if i >= 5:
            break
except ImportError:
    print("openpyxl not found")
except Exception as e:
    print(f"Error: {e}")
