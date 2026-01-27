import openpyxl
import re
from pathlib import Path
from collections import defaultdict

# Paths
EXCEL_PATH = r"c:/Users/sdnse/OneDrive/Dokumen/Yum/PROJEK SUDIN JU2/ASKA SUDIN JU 2/sudin_aska/.Data_Sekolah_SudinJU2/DAFTAR SEKOLAH NEGERI & SWASTA DI JAKARTA UTARA II.xlsx"
MD_PATH = r"c:/Users/sdnse/OneDrive/Dokumen/Yum/PROJEK SUDIN JU2/ASKA SUDIN JU 2/sudin_aska/kecerdasan/Data_Sekolah_Sudin_JU2.md"

def parse_markdown(path):
    if not Path(path).exists():
        return {}, ""
    
    content = Path(path).read_text(encoding="utf-8")
    lines = content.splitlines()
    
    # Extract header (everything before the first Kecamatan)
    header_lines = []
    data_start = 0
    for i, line in enumerate(lines):
        if line.startswith("## Kecamatan"):
            data_start = i
            break
        header_lines.append(line)
    
    header = "\n".join(header_lines)
    
    # Parse data
    structure = defaultdict(lambda: defaultdict(list))
    current_kec = None
    current_kel = None
    
    # Current school buffer
    current_school = {}
    
    i = data_start
    while i < len(lines):
        line = lines[i]
        
        if line.startswith("## Kecamatan"):
            current_kec = line.replace("## Kecamatan", "").strip()
            current_kel = None
        elif line.startswith("### Kelurahan"):
            current_kel = line.replace("### Kelurahan", "").strip()
        elif line.startswith("#### "):
            # Specific optimization: if we were parsing a school, save it
            if current_school:
                # Save previous school
                if current_kec and current_kel:
                    structure[current_kec][current_kel].append(current_school)
                current_school = {}
            
            school_name = line.replace("#### ", "").strip()
            current_school = {"Nama Sekolah": school_name}
        elif line.strip().startswith("- **"):
            # Parse attributes
            match = re.match(r"- \*\*(.*?)\*\*: (.*)", line.strip())
            if match and current_school is not None:
                key, value = match.groups()
                current_school[key] = value.strip()
        
        i += 1
        
    # Add last school
    if current_school and current_kec and current_kel:
        structure[current_kec][current_kel].append(current_school)
        
    return structure, header

def parse_excel(path):
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    
    header = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    # Assume headers map to columns. Let's find indices.
    # We expect: NPSN, Nama Sekolah, Jenjang, Status, Alamat, Kecamatan, Kelurahan
    # Mapping based on typical header names or just index if consistent
    
    # Helper to find index
    def get_idx(name_part):
        for idx, h in enumerate(header):
            if h and name_part.lower() in str(h).lower():
                return idx
        return -1
    
    idx_npsn = get_idx("NPSN")
    idx_nama = get_idx("Nama Sekolah")
    if idx_nama == -1: idx_nama = get_idx("Nama") 
    idx_jenjang = get_idx("Jenjang")
    idx_status = get_idx("Status")
    idx_alamat = get_idx("Alamat")
    idx_kec = get_idx("Kecamatan")
    idx_kel = get_idx("Kelurahan")
    
    if idx_npsn == -1 or idx_nama == -1:
        print(f"Critical columns missing in Excel. Header: {header}")
        return []

    schools = []
    
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row[idx_nama]: continue
        
        school = {
            "Nama Sekolah": str(row[idx_nama]).strip(),
            "NPSN": str(row[idx_npsn]).strip() if idx_npsn != -1 and row[idx_npsn] else "-",
            "Jenjang": str(row[idx_jenjang]).strip() if idx_jenjang != -1 and row[idx_jenjang] else "-",
            "Status": str(row[idx_status]).strip() if idx_status != -1 and row[idx_status] else "-",
            "Alamat": str(row[idx_alamat]).strip() if idx_alamat != -1 and row[idx_alamat] else "-",
            "Kecamatan": str(row[idx_kec]).strip() if idx_kec != -1 and row[idx_kec] else "Unknown",
            "Kelurahan": str(row[idx_kel]).strip() if idx_kel != -1 and row[idx_kel] else "Unknown",
        }
        schools.append(school)
        
    return schools

def merge_data(md_structure, excel_data):
    # Flatten md_structure to check NPSNs
    existing_npsns = set()
    for kec in md_structure:
        for kel in md_structure[kec]:
            for school in md_structure[kec][kel]:
                if "NPSN" in school and school["NPSN"] != "-":
                    existing_npsns.add(school["NPSN"])
    
    added_count = 0
    for school in excel_data:
        npsn = school["NPSN"]
        # Basic dedup by NPSN
        if npsn in existing_npsns:
            continue
            
        # Add to structure
        kec = school["Kecamatan"]
        kel = school["Kelurahan"]
        
        # Normalize keys if possible (e.g. Title Case)
        # But we assume excel data is good or we just add new keys
        
        # Remove keys not for display
        display_school = school.copy()
        del display_school["Kecamatan"]
        del display_school["Kelurahan"]
        
        md_structure[kec][kel].append(display_school)
        existing_npsns.add(npsn)
        added_count += 1
        
    print(f"Added {added_count} new schools.")
    return md_structure

def generate_markdown(header, structure):
    output = [header.strip()]
    if not output[0]:
        output = ["# Data Sekolah SUDINDIK JU 2"]
        
    output.append("")
    output.append("Berikut adalah daftar sekolah di wilayah Suku Dinas Pendidikan Jakarta Utara 2, dikelompokkan berdasarkan Kecamatan dan Kelurahan.")
    output.append("")
    
    # Sort keys for consistent output
    sorted_kec = sorted(structure.keys())
    
    for kec in sorted_kec:
        output.append(f"## Kecamatan {kec}")
        sorted_kel = sorted(structure[kec].keys())
        for kel in sorted_kel:
            output.append(f"### Kelurahan {kel}")
            
            # Sort schools by name
            schools = sorted(structure[kec][kel], key=lambda x: x.get("Nama Sekolah", ""))
            
            for school in schools:
                output.append(f"#### {school.get('Nama Sekolah', 'Unnamed')}")
                # Order attributes: NPSN, Jenjang, Status, Alamat
                attrs = ["NPSN", "Jenjang", "Status", "Alamat"]
                for attr in attrs:
                    val = school.get(attr, "-")
                    output.append(f"- **{attr.strip()}**: {val}")
                output.append("")
    
    return "\n".join(output)

def main():
    print("Parsing existing markdown...")
    md_structure, md_header = parse_markdown(MD_PATH)
    
    print("Parsing excel...")
    excel_data = parse_excel(EXCEL_PATH)
    
    print("Merging data...")
    merged_structure = merge_data(md_structure, excel_data)
    
    print("Generating new markdown...")
    new_content = generate_markdown(md_header, merged_structure)
    
    # Backup
    backup_path = Path(MD_PATH).with_suffix(".md.bak")
    Path(MD_PATH).rename(backup_path)
    print(f"Backed up to {backup_path}")
    
    Path(MD_PATH).write_text(new_content, encoding="utf-8")
    print(f"Written to {MD_PATH}")

if __name__ == "__main__":
    main()
