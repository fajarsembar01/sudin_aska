#!/usr/bin/env python3
"""Convert a BOP realization workbook into the repository claim JSON format."""

from __future__ import annotations

import argparse
import json
import re
from pathlib import Path

from openpyxl import load_workbook


def normalize_account_code(code: str) -> str:
    """Normalize legacy six-segment account codes to the current segment widths."""
    segments = str(code).strip().split(".")
    if len(segments) == 6:
        return ".".join(segments[:4] + [segments[4].zfill(3), segments[5].zfill(5)])
    return str(code).strip()


def convert(source: Path) -> dict:
    worksheet = load_workbook(source, read_only=True, data_only=True).active
    year_match = re.search(r"(20\d{2})", str(worksheet.cell(3, 1).value or ""))
    tw_match = re.search(r"(\d+)", str(worksheet.cell(2, 1).value or ""))
    if not year_match or not tw_match:
        raise ValueError("Tahun atau triwulan tidak ditemukan pada workbook.")
    year = int(year_match.group(1))
    tw = int(tw_match.group(1))
    organization = str(worksheet.cell(4, 1).value or "").strip()
    headers = [cell.value for cell in worksheet[6]]
    descriptions = [cell.value for cell in worksheet[7]]
    account_columns = [
        index
        for index, value in enumerate(headers)
        if isinstance(value, str) and value[:1].isdigit()
    ]
    accounts = [
        {
            "code": normalize_account_code(headers[index]),
            "source_code": str(headers[index]).strip(),
            "name": str(descriptions[index]).strip(),
        }
        for index in account_columns
    ]

    schools = []
    education_level = None
    for row in worksheet.iter_rows(min_row=8, values_only=True):
        if isinstance(row[0], str) and row[0].startswith("Penyediaan Biaya Operasional Pendidikan Jenjang "):
            education_level = row[0].rsplit(" ", 1)[-1]
            continue
        if not isinstance(row[0], (int, float)) or not isinstance(row[1], str) or " \\ " not in row[1]:
            continue

        npsn, school_name = (part.strip() for part in row[1].split(" \\ ", 1))
        transactions = []
        for index in account_columns:
            amount = row[index]
            if not isinstance(amount, (int, float)) or amount <= 0:
                continue
            transactions.append(
                {
                    "account_code": normalize_account_code(headers[index]),
                    "source_account_code": str(headers[index]).strip(),
                    "description": str(descriptions[index]).strip(),
                    "amount": int(amount),
                }
            )
        schools.append(
            {
                "npsn": npsn,
                "name": school_name,
                "education_level": education_level,
                "transactions": transactions,
                "total_amount": sum(item["amount"] for item in transactions),
            }
        )

    return {
        "schema_version": 1,
        "dataset_id": f"bop-{year}-tw{tw:02d}-jakut-wilayah-2",
        "fund_source": "BOP",
        "year": year,
        "tw": tw,
        "organization": organization,
        "accounts": accounts,
        "schools": schools,
        "school_count": len(schools),
        "transaction_count": sum(len(school["transactions"]) for school in schools),
        "total_amount": sum(school["total_amount"] for school in schools),
    }


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("source", type=Path)
    parser.add_argument("output", type=Path)
    args = parser.parse_args()
    result = convert(args.source)
    args.output.parent.mkdir(parents=True, exist_ok=True)
    args.output.write_text(json.dumps(result, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")


if __name__ == "__main__":
    main()
