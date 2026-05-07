"""Routes for Daftar Tamu (guestbook) module."""

from __future__ import annotations

import base64
import csv
import hashlib
import json
import math
import os
import time
from io import BytesIO, StringIO
import requests
from pathlib import Path
from datetime import date, datetime
from typing import Optional

from flask import (
    Blueprint,
    Response,
    current_app,
    flash,
    jsonify,
    redirect,
    render_template,
    request,
    session,
    send_file,
    stream_with_context,
    url_for,
)
from psycopg2.extras import Json
from urllib.parse import quote_plus
from PIL import Image, ImageDraw, ImageFont
import qrcode

from dashboard.auth import current_user, role_required
from dashboard.db_access import get_cursor
from dashboard.queries import record_admin_action
from dashboard.portal.permissions import can_access_aska, get_permission_summary, is_superadmin
from dashboard.portal.queries import (
    fetch_admin_pending_summary,
    fetch_portal_undo_window_seconds,
    get_staff_assigned_schools,
    get_user_kecamatan_details,
    get_user_kecamatan_ids,
    list_portal_kontak,
    list_schools_by_kecamatan,
    PORTAL_UNDO_WINDOW_DEFAULT_SECONDS,
    PORTAL_UNDO_WINDOW_MIN_SECONDS,
    PORTAL_UNDO_WINDOW_MAX_SECONDS,
)
from utils import current_jakarta_time, to_jakarta

from .media import stamp_guestbook_photo
from .queries import (
    DEFAULT_USER_SORT,
    DEFAULT_SORT,
    SORT_OPTIONS,
    USER_SORT_OPTIONS,
    ensure_daftar_tamu_seed_data,
    fetch_dashboard_summary,
    fetch_guest_history,
    fetch_map_data,
    fetch_recent_visits,
    fetch_school_rankings,
    fetch_school_visit_bucket_rows,
    fetch_school_visit_day_guests,
    fetch_school_visit_days,
    fetch_school_visit_histogram,
    fetch_school_visit_history,
    fetch_user_guestbook_history,
    fetch_user_rankings,
    fetch_user_visit_history,
    fetch_unvisited_schools,
    fetch_school_pending_counts,
    get_transaction_detail,
    list_admin_public_school_summary,
    list_admin_public_transactions,
    list_admin_transactions,
    list_transaction_previous_single_guest_photos,
    list_guest_candidates,
    list_general_guest_candidates,
    list_general_guests_admin,
    list_user_transactions,
    list_user_visited_school_ids,
    list_purpose_keyword_rows,
    list_purpose_keywords_by_usage,
    list_contact_priority_rows,
    list_school_public_transactions,
    list_school_transactions,
    update_contact_priority,
    set_purpose_keyword_active,
    fetch_guestbook_ux_metric_rows,
    USER_APP_NOTIFICATION_CATEGORIES,
    fetch_user_notification_summary,
    upsert_guestbook_ux_metrics,
    list_user_notifications,
    mark_user_notifications_read,
    create_guestbook_status_notifications,
    upsert_transaction_staff_note,
    upsert_purpose_keyword,
    update_public_transaction_status,
    update_transaction_status,
)

DAFTAR_TAMU_URL_PREFIX = "/daftar-tamu"
_HISTORY_TAB_OPTIONS = {"beranda", "detail"}
_HISTORY_STATUS_OPTIONS = {"pending", "approved", "rejected"}
_HISTORY_SORT_OPTIONS = {"date_desc", "date_asc"}
_STAFF_NOTE_LEVEL_LABEL_MAP = {
    "tidak_perlu": "Tidak Perlu Penanganan",
    "pantau": "Perlu Dipantau",
    "tindak_lanjut": "Butuh Penanganan",
    "mendesak": "Sangat Butuh Penanganan",
}
_STAFF_NOTE_LEVEL_TONE_MAP = {
    "tidak_perlu": "secondary",
    "pantau": "info",
    "tindak_lanjut": "warning",
    "mendesak": "danger",
}
_VISIT_DISTRIBUTION_BUCKETS = [
    {"key": "d0", "label": "0x", "min_visits": 0, "max_visits": 0},
    {"key": "d1", "label": "1x", "min_visits": 1, "max_visits": 1},
    {"key": "d2", "label": "2x", "min_visits": 2, "max_visits": 2},
    {"key": "d3", "label": "3x", "min_visits": 3, "max_visits": 3},
    {"key": "d4", "label": "4x", "min_visits": 4, "max_visits": 4},
    {"key": "d5", "label": "5x", "min_visits": 5, "max_visits": 5},
    {"key": "d6", "label": "6x", "min_visits": 6, "max_visits": 6},
    {"key": "d7", "label": "7x", "min_visits": 7, "max_visits": 7},
    {"key": "d8_plus", "label": "8+", "min_visits": 8, "max_visits": None},
]
_VISIT_FREQUENCY_BUCKETS = [
    {"key": "f0", "label": "0x", "min_visits": 0, "max_visits": 0},
    {"key": "f1_4", "label": "1-4x", "min_visits": 1, "max_visits": 4},
    {"key": "f5_9", "label": "5-9x", "min_visits": 5, "max_visits": 9},
    {"key": "f10_plus", "label": "10+", "min_visits": 10, "max_visits": None},
]


daftar_tamu_bp = Blueprint(
    "daftar_tamu",
    __name__,
    url_prefix=DAFTAR_TAMU_URL_PREFIX,
    template_folder="templates",
    static_folder="static",
)

_PREVIEW_ADMIN_SESSION_KEY = "preview_admin_user"
_PREVIEW_TARGET_SESSION_KEY = "preview_target_user"
_PREVIEW_READ_ONLY_METHODS = {"POST", "PUT", "PATCH", "DELETE"}


def _is_preview_read_only_session() -> bool:
    preview_admin = session.get(_PREVIEW_ADMIN_SESSION_KEY)
    preview_target = session.get(_PREVIEW_TARGET_SESSION_KEY)
    return (
        isinstance(preview_admin, dict)
        and preview_admin.get("role") == "admin"
        and isinstance(preview_target, dict)
        and bool(preview_target.get("id"))
    )


def _preview_read_only_block_response(*, fallback_url: str) -> Response:
    message = "Mode preview aktif. Aksi edit dinonaktifkan."
    accept_header = (request.headers.get("Accept") or "").lower()
    content_type = (request.content_type or "").lower()
    wants_json = (
        request.is_json
        or request.headers.get("X-Requested-With") == "XMLHttpRequest"
        or "application/json" in accept_header
        or "application/json" in content_type
    )
    if wants_json:
        return jsonify({"success": False, "message": message, "preview_read_only": True}), 403

    flash(message, "warning")
    target_url = (request.referrer or "").strip() or fallback_url
    return redirect(target_url, code=303)


@daftar_tamu_bp.before_request
def _enforce_preview_read_only_mode() -> Response | None:
    if request.method not in _PREVIEW_READ_ONLY_METHODS:
        return None
    if not _is_preview_read_only_session():
        return None
    return _preview_read_only_block_response(fallback_url=url_for("portal.preview_accounts"))


def _parse_iso_date(value: Optional[str]) -> Optional[date]:
    if not value:
        return None
    try:
        return datetime.strptime(value, "%Y-%m-%d").date()
    except ValueError:
        return None


def _to_int(value: Optional[str], default: int) -> int:
    try:
        return int(value or default)
    except (TypeError, ValueError):
        return default


def _today_jakarta() -> date:
    return current_jakarta_time().date()


def _needs_profile_photo_completion(user: dict | None) -> bool:
    """Return True when staff/coordinator must complete profile photo first."""
    if not isinstance(user, dict):
        return False
    role_value = (user.get("role") or "").strip().lower()
    if role_value not in {"staff", "coordinator"}:
        return False
    if _is_preview_read_only_session():
        return False
    return not bool((user.get("profile_photo_path") or "").strip())


def _require_profile_photo_redirect(user: dict | None) -> Response | None:
    """Redirect to profile page when photo completion is mandatory."""
    if not _needs_profile_photo_completion(user):
        return None
    flash("Foto profil wajib diisi untuk melanjutkan akses OSS.", "warning")
    return redirect(url_for("portal.user_profile_settings"))


def _parse_guest_scope(value: Optional[str], default: str = "sudin") -> str:
    scope = (value or "").strip().lower()
    if scope == "semua":
        scope = "all"
    if scope not in {"sudin", "umum", "all"}:
        scope = default
    return scope


def _parse_school_status(value: Optional[str], default: Optional[str] = None) -> str:
    status = (value or "").strip().lower()
    if not status:
        return default or "all"
    if status in {"all", "semua"}:
        return "all"
    if status in {"negeri", "state"}:
        return "negeri"
    if status in {"swasta", "private"}:
        return "swasta"
    return default or "all"


def _normalize_staff_note_level(value: Optional[str], default: str = "") -> str:
    level = (value or "").strip().lower()
    if level in {"mendesak", "urgent", "critical", "sangat_mendesak", "sangat mendesak"}:
        level = "mendesak"
    elif level in {"tindak_lanjut", "tindak lanjut", "normal", "follow_up", "perlu_tindakan"}:
        level = "tindak_lanjut"
    elif level in {"pantau", "monitor", "other", "lainnya", "lainnya/pantau"}:
        level = "pantau"
    elif level in {"tidak_perlu", "tidak perlu", "info", "informasi", "arsip", "no_action"}:
        level = "tidak_perlu"
    if level not in {"tidak_perlu", "pantau", "tindak_lanjut", "mendesak"}:
        level = default
    return level


def _parse_guest_ids(raw: Optional[str]) -> list[int]:
    if not raw:
        return []
    try:
        payload = json.loads(raw)
    except json.JSONDecodeError:
        return []
    if not isinstance(payload, list):
        return []
    ids: list[int] = []
    for item in payload:
        try:
            ids.append(int(item))
        except (TypeError, ValueError):
            continue
    # Deduplicate while preserving order
    seen = set()
    unique_ids = []
    for val in ids:
        if val in seen:
            continue
        seen.add(val)
        unique_ids.append(val)
    return unique_ids


def _notify_guestbook_status_change(
    *,
    transaction_id: int,
    status: str,
    actor: Optional[dict],
    is_public: bool = False,
) -> None:
    # Do not push Telegram notifications for public-web guestbook flow
    # or any status change performed by school verifiers.
    if is_public or (actor or {}).get("role") == "sekolah":
        return

    normalized_status = (status or "").strip().lower()
    if normalized_status not in {"approved", "rejected"}:
        return

    if normalized_status == "approved":
        status_label = "✅ Disetujui"
    else:
        status_label = "❌ Ditolak"

    actor_name = (actor or {}).get("full_name") or (actor or {}).get("email")
    school_name = None
    photo_links: list[dict] = []
    guest_names: list[str] = []

    detail = get_transaction_detail(transaction_id)
    if detail:
        school_name = detail.get("school_name")
        photo_links = _build_guestbook_photo_links(transaction_id=transaction_id, detail=detail)
        guest_names = _extract_guest_names_from_detail(detail)

    from dashboard.telegram_notifications import notify_guestbook_status_update

    notify_guestbook_status_update(
        transaction_id=transaction_id,
        school_name=school_name,
        status_label=status_label,
        actor_name=actor_name,
        actor_username=None,
        purpose=(detail or {}).get("purpose"),
        notes=(detail or {}).get("notes"),
        guest_names=guest_names,
        photo_links=photo_links,
    )


def _notify_user_app_status_change(
    *,
    transaction_id: int,
    status: str,
    actor: Optional[dict],
    reviewer_notes: Optional[str] = None,
) -> None:
    safe_status = (status or "").strip().lower()
    if safe_status not in {"pending", "approved", "rejected"}:
        return

    actor_name = (actor or {}).get("full_name") or (actor or {}).get("email")
    history_link = url_for("daftar_tamu.user_guestbook_history", tab="detail")
    school_history_link = url_for("daftar_tamu.sekolah_riwayat")

    create_guestbook_status_notifications(
        transaction_id=transaction_id,
        status=safe_status,
        actor_name=actor_name,
        reviewer_notes=reviewer_notes,
        link=history_link,
        school_link=school_history_link,
    )


def _parse_guest_payload(raw: Optional[str]) -> tuple[list[int], list[int]]:
    """Parse guest payload list into (sudin_ids, umum_ids)."""
    if not raw:
        return [], []
    try:
        payload = json.loads(raw)
    except json.JSONDecodeError:
        return [], []
    if not isinstance(payload, list):
        return [], []
    sudin_ids: list[int] = []
    umum_ids: list[int] = []
    for item in payload:
        if not isinstance(item, dict):
            continue
        guest_type = (item.get("type") or "").strip().lower()
        guest_id = item.get("id")
        try:
            guest_id = int(guest_id)
        except (TypeError, ValueError):
            continue
        if guest_type == "umum":
            umum_ids.append(guest_id)
        else:
            sudin_ids.append(guest_id)
    # Deduplicate while preserving order
    def _dedupe(values: list[int]) -> list[int]:
        seen = set()
        output = []
        for val in values:
            if val in seen:
                continue
            seen.add(val)
            output.append(val)
        return output

    return _dedupe(sudin_ids), _dedupe(umum_ids)


def _is_truthy(value: Optional[str]) -> bool:
    normalized = (value or "").strip().lower()
    return normalized in {"1", "true", "yes", "on"}


def _find_sudin_same_day_approved_duplicates(
    *,
    school_id: int,
    sudin_ids: list[int],
    visit_at: datetime,
) -> list[dict]:
    """Find SUDIN guests already approved at the same school on the same Jakarta date."""
    if not school_id or not sudin_ids:
        return []

    with get_cursor() as cur:
        cur.execute(
            """
            SELECT
                g.user_id AS guest_id,
                COALESCE(
                    NULLIF(TRIM(u.full_name), ''),
                    NULLIF(TRIM(u.email), ''),
                    CONCAT('ID ', g.user_id::TEXT)
                ) AS guest_name,
                COUNT(DISTINCT t.id)::INT AS approved_count
            FROM daftar_tamu_transactions t
            JOIN daftar_tamu_transaction_guests g ON g.transaction_id = t.id
            LEFT JOIN dashboard_users u ON u.id = g.user_id
            WHERE t.school_id = %s
              AND t.status = 'approved'
              AND (g.guest_type = 'sudin' OR g.guest_type IS NULL)
              AND g.user_id = ANY(%s::INT[])
              AND DATE(t.visit_at AT TIME ZONE 'Asia/Jakarta') = DATE(%s AT TIME ZONE 'Asia/Jakarta')
            GROUP BY g.user_id, guest_name
            HAVING COUNT(DISTINCT t.id) >= 1
            ORDER BY approved_count DESC, guest_name ASC
            """,
            (school_id, sudin_ids, visit_at),
        )
        rows = cur.fetchall() or []
    return [dict(row) for row in rows]


def _build_sudin_duplicate_warning_message(*, school_name: Optional[str], duplicate_rows: list[dict]) -> str:
    guest_names = [
        str(row.get("guest_name") or "").strip()
        for row in duplicate_rows
        if str(row.get("guest_name") or "").strip()
    ]
    unique_names: list[str] = []
    seen: set[str] = set()
    for name in guest_names:
        key = name.casefold()
        if key in seen:
            continue
        seen.add(key)
        unique_names.append(name)

    guest_text = ", ".join(unique_names) if unique_names else "Tamu terpilih"
    school_text = (school_name or "sekolah ini").strip() or "sekolah ini"
    return (
        f"{guest_text} sudah terverifikasi di {school_text} pada hari ini. "
        "Yakin ingin menambahkan kunjungan lagi?"
    )


def _web_aska_base_url() -> str:
    base = (
        os.getenv("WEB_ASKA_BASE_URL")
        or os.getenv("WEB_ASKA_PUBLIC_URL")
        or os.getenv("ASKA_PUBLIC_BASE_URL")
        or "https://aska.sudindikju2.com"
    )
    base = (base or "").strip()
    if base and not base.startswith(("http://", "https://")):
        base = f"https://{base}"
    return base.rstrip("/")


def _get_guestbook_qr_payload(school_id: int) -> Optional[dict]:
    with get_cursor() as cur:
        cur.execute(
            """
            SELECT metadata->'guestbook_qr' AS guestbook_qr
            FROM portal_schools
            WHERE id = %s
            """,
            (school_id,),
        )
        row = cur.fetchone()
    payload = dict(row).get("guestbook_qr") if row else None
    if isinstance(payload, str):
        try:
            payload = json.loads(payload)
        except json.JSONDecodeError:
            payload = None
    return payload if isinstance(payload, dict) else None


def _store_guestbook_qr_payload(school_id: int, payload: dict) -> None:
    with get_cursor(commit=True) as cur:
        cur.execute(
            """
            UPDATE portal_schools
            SET metadata = jsonb_set(
                    CASE
                        WHEN jsonb_typeof(COALESCE(metadata, '{}'::jsonb)) = 'object'
                            THEN COALESCE(metadata, '{}'::jsonb)
                        ELSE '{}'::jsonb
                    END,
                    '{guestbook_qr}',
                    %s::jsonb,
                    true
                ),
                updated_at = NOW()
            WHERE id = %s
            """,
            (Json(payload), school_id),
        )


def _build_guestbook_qr(target_url: str, size: int = 1024, logo_path: Optional[Path] = None) -> Image.Image:
    qr = qrcode.QRCode(
        version=None,
        error_correction=qrcode.constants.ERROR_CORRECT_H,
        box_size=10,
        border=4,
    )
    qr.add_data(target_url)
    qr.make(fit=True)
    qr_img = qr.make_image(fill_color="black", back_color="white").convert("RGBA")
    qr_img = qr_img.resize((size, size), Image.LANCZOS)


    base_dir = Path(__file__).resolve().parents[2]
    logo_candidates = []
    if logo_path:
        logo_candidates.append(logo_path)
    logo_candidates.extend([
        base_dir / "web_aska" / "static" / "favicon.ico",
        base_dir / "web_aska" / "static" / "logo.png",
    ])
    logo = None
    for logo_path in logo_candidates:
        if not logo_path.exists():
            continue
        try:
            logo = Image.open(logo_path)
            if getattr(logo, "is_animated", False):
                logo.seek(0)
            logo = logo.convert("RGBA")
            break
        except Exception:
            logo = None
            continue

    if logo is not None:
        # Keep logo small to preserve QR readability (<= ~18% of width).
        logo_size = int(size * 0.16)
        logo.thumbnail((logo_size, logo_size), Image.LANCZOS)
        pos = ((size - logo.width) // 2, (size - logo.height) // 2)
        padding = int(logo_size * 0.08)
        bg_box = Image.new(
            "RGBA",
            (logo.width + padding * 2, logo.height + padding * 2),
            (255, 255, 255, 235),
        )
        bg_pos = (pos[0] - padding, pos[1] - padding)
        qr_img.paste(bg_box, bg_pos, bg_box)
        qr_img.paste(logo, pos, logo)
    return qr_img


def _load_school_logo(school: dict) -> Optional[Image.Image]:
    logo_url = (school or {}).get("logo_url")
    if not logo_url:
        return None
    try:
        if str(logo_url).startswith(("http://", "https://")):
            resp = requests.get(logo_url, timeout=5)
            resp.raise_for_status()
            return Image.open(BytesIO(resp.content))
            
        # Clean path: remove leading slash to prevent absolute path issues
        clean_path = str(logo_url).lstrip("/")
        
        # Define base root directory
        root_dir = Path(__file__).resolve().parent.parent.parent
        
        # Logic to try multiple potential paths
        candidates = []
        
        # 1. Try as is (relative to root)
        candidates.append(root_dir / clean_path)
        
        # 2. Handle known mismatch: DB 'portal/uploads' -> FS 'uploads/portal'
        if clean_path.startswith("portal/uploads/"):
            swapped_path = clean_path.replace("portal/uploads/", "uploads/portal/", 1)
            candidates.append(root_dir / swapped_path)
            
        # 3. Try forcing 'uploads/portal/' + filename
        filename = Path(clean_path).name
        candidates.append(root_dir / "uploads" / "portal" / "logos" / filename)
        
        # 4. Try just 'uploads/' + clean_path (if portal prefix is extra)
        if clean_path.startswith("portal/"):
             candidates.append(root_dir / "uploads" / clean_path[7:]) # remove 'portal/'
             
        for candidate in candidates:
            if candidate.exists():
                return Image.open(candidate)
                
        return None
    except Exception:
        return None



def _measure_multiline(draw: ImageDraw.ImageDraw, text: str, font: ImageFont.ImageFont, spacing: int = 4) -> tuple[int, int]:
    if hasattr(draw, "multiline_textbbox"):
        bbox = draw.multiline_textbbox((0, 0), text, font=font, spacing=spacing)
        return bbox[2] - bbox[0], bbox[3] - bbox[1]
    if hasattr(draw, "multiline_textsize"):
        return draw.multiline_textsize(text, font=font, spacing=spacing)
    # Fallback
    lines = text.splitlines() or [text]
    widths = [draw.textlength(line, font=font) if hasattr(draw, "textlength") else len(line) * 6 for line in lines]
    return int(max(widths) if widths else 0), int(len(lines) * 12)


def _fetch_school_for_user(user_id: int) -> Optional[dict]:
    with get_cursor() as cur:
        cur.execute(
            """
            SELECT s.id,
                   s.npsn,
                   s.name,
                   s.jenjang,
                   s.alamat,
                   s.metadata,
                   s.logo_url,
                   l.name AS kelurahan_name,
                   k.name AS kecamatan_name
            FROM dashboard_users u
            JOIN portal_schools s ON u.school_id = s.id
            LEFT JOIN portal_kelurahan l ON s.kelurahan_id = l.id
            LEFT JOIN portal_kecamatan k ON l.kecamatan_id = k.id
            WHERE u.id = %s
            """,
            (user_id,),
        )
        row = cur.fetchone()
        return dict(row) if row else None


def _fetch_school_profile(school_id: int) -> Optional[dict]:
    with get_cursor() as cur:
        cur.execute(
            """
            SELECT s.id,
                   s.npsn,
                   s.name,
                   s.jenjang,
                   l.kecamatan_id AS kecamatan_id,
                   k.name AS kecamatan_name
            FROM portal_schools s
            LEFT JOIN portal_kelurahan l ON s.kelurahan_id = l.id
            LEFT JOIN portal_kecamatan k ON l.kecamatan_id = k.id
            WHERE s.id = %s
            """,
            (school_id,),
        )
        row = cur.fetchone()
        return dict(row) if row else None


def _fetch_dashboard_user(user_id: int) -> Optional[dict]:
    with get_cursor() as cur:
        cur.execute(
            """
            SELECT
                u.id,
                u.full_name,
                u.email,
                u.role,
                u.jabatan,
                u.requested_kecamatan,
                k.name AS requested_kecamatan_name
            FROM dashboard_users u
            LEFT JOIN portal_kecamatan k ON u.requested_kecamatan = k.id
            WHERE u.id = %s
            LIMIT 1
            """,
            (user_id,),
        )
        row = cur.fetchone()
    return dict(row) if row else None


def _resolve_dashboard_kecamatan_ids(user: Optional[dict]) -> Optional[list[int]]:
    if not isinstance(user, dict):
        return None
    role_value = (user.get("role") or "").strip().lower()
    if role_value != "coordinator":
        return None

    user_id = int(user.get("id") or 0)
    resolved_ids: list[int] = []
    seen_ids: set[int] = set()

    def _append_id(raw_value: object) -> None:
        try:
            kecamatan_id = int(raw_value or 0)
        except (TypeError, ValueError):
            return
        if kecamatan_id <= 0 or kecamatan_id in seen_ids:
            return
        seen_ids.add(kecamatan_id)
        resolved_ids.append(kecamatan_id)

    _append_id(user.get("requested_kecamatan"))
    if user_id > 0:
        db_profile = _fetch_dashboard_user(user_id) or {}
        _append_id(db_profile.get("requested_kecamatan"))
        try:
            for kec_id in get_user_kecamatan_ids(user_id):
                _append_id(kec_id)
        except Exception:
            pass

    return resolved_ids or None


def _can_access_school_for_dashboard(
    *,
    user: Optional[dict],
    school: Optional[dict],
    kecamatan_ids: Optional[list[int]] = None,
) -> bool:
    if not isinstance(user, dict) or not isinstance(school, dict):
        return False
    role_value = (user.get("role") or "").strip().lower()
    if role_value != "coordinator":
        return True

    allowed_ids = kecamatan_ids or _resolve_dashboard_kecamatan_ids(user)
    if not allowed_ids:
        return True
    try:
        school_kecamatan_id = int(school.get("kecamatan_id") or 0)
    except (TypeError, ValueError):
        school_kecamatan_id = 0
    if school_kecamatan_id <= 0:
        return False
    return school_kecamatan_id in set(allowed_ids)


def _list_unvisited_schools_for_user(
    *,
    user_id: int,
    date_from: Optional[date],
    date_to: Optional[date],
    guest_scope: str,
    assigned_schools: Optional[list[dict]] = None,
    kecamatan_ids: Optional[list[int]] = None,
) -> list[dict]:
    assigned_schools = assigned_schools or []
    candidates: list[dict] = []

    if assigned_schools:
        for school in assigned_schools:
            candidates.append(
                {
                    "school_id": school.get("school_id"),
                    "school_name": school.get("school_name"),
                    "npsn": school.get("npsn"),
                    "kecamatan": school.get("kecamatan_name"),
                }
            )
    else:
        if kecamatan_ids is None:
            kecamatan_ids = get_user_kecamatan_ids(user_id)
        if not kecamatan_ids:
            return []
        schools = list_schools_by_kecamatan(kecamatan_ids, active_only=True)
        for school in schools:
            candidates.append(
                {
                    "school_id": school.get("id"),
                    "school_name": school.get("name"),
                    "npsn": school.get("npsn"),
                    "kecamatan": school.get("kecamatan_name"),
                }
            )

    visited_ids = set(
        list_user_visited_school_ids(
            user_id=user_id,
            date_from=date_from,
            date_to=date_to,
            guest_scope=guest_scope,
        )
    )
    return [school for school in candidates if school.get("school_id") not in visited_ids]


def _sanitize_phone(phone: str) -> str:
    digits_only = "".join(ch for ch in phone if ch.isdigit())
    if digits_only.startswith("0"):
        digits_only = "62" + digits_only[1:]
    return digits_only


def _photo_filename_from_path(photo_path: Optional[str]) -> str:
    if not photo_path:
        return ""
    normalized = str(photo_path).replace("\\", "/").strip()
    if not normalized:
        return ""
    if "uploads/portal/" in normalized:
        normalized = normalized.split("uploads/portal/", 1)[1]
    else:
        normalized = normalized.lstrip("/")
    normalized = normalized.replace("\\", "/").lstrip("/")
    if ".." in Path(normalized).parts:
        return ""
    return normalized


def _resolve_portal_upload_path(filename: str) -> Optional[Path]:
    safe_name = _photo_filename_from_path(filename)
    if not safe_name:
        return None
    upload_root = Path(__file__).resolve().parents[2] / "uploads" / "portal"
    candidate = (upload_root / safe_name).resolve()
    try:
        candidate.relative_to(upload_root.resolve())
    except ValueError:
        return None
    if not candidate.exists() or not candidate.is_file():
        return None
    return candidate


def _build_photo_url(photo_path: Optional[str], *, external: bool = False) -> Optional[str]:
    filename = _photo_filename_from_path(photo_path)
    if not filename:
        return None
    return url_for("portal.uploaded_file", filename=filename, _external=external)


def _build_photo_thumb_url(
    photo_path: Optional[str],
    *,
    width: int = 360,
    quality: int = 72,
    external: bool = False,
) -> Optional[str]:
    filename = _photo_filename_from_path(photo_path)
    if not filename:
        return None
    safe_width = max(120, min(int(width), 1200))
    safe_quality = max(40, min(int(quality), 90))
    return url_for(
        "daftar_tamu.guestbook_photo_thumb",
        filename=filename,
        w=safe_width,
        q=safe_quality,
        _external=external,
    )


def _format_guest_reference_button_label(
    *,
    kind: str,
    index: int,
    total_items: int,
    guest_name: Optional[str],
) -> str:
    base = "Foto Profil" if kind == "profile" else "Foto Sebelumnya"
    if total_items <= 1:
        return base

    base_with_index = f"{base} {index}"
    name = (guest_name or "").strip()
    if not name:
        return base_with_index

    name_parts = [part for part in name.split() if part]
    short_name = " ".join(name_parts[:2]) if name_parts else name
    label = f"{base_with_index}: {short_name}"
    if len(label) <= 64:
        return label
    max_name_len = max(1, 64 - len(base_with_index) - 5)
    return f"{base_with_index}: {short_name[:max_name_len]}..."


def _build_guestbook_photo_links(
    *,
    transaction_id: int,
    detail: Optional[dict] = None,
) -> list[dict]:
    links: list[dict] = []
    seen_keys: set[tuple[str, str]] = set()

    current_photo_url = None
    if detail:
        current_photo_url = _build_photo_url(detail.get("photo_path"), external=True)
    else:
        current_detail = get_transaction_detail(transaction_id)
        current_photo_url = _build_photo_url((current_detail or {}).get("photo_path"), external=True)

    if current_photo_url:
        key = ("Foto Transaksi", current_photo_url)
        seen_keys.add(key)
        links.append({"text": key[0], "url": key[1]})

    preferred_refs: list[dict] = []
    for row in list_transaction_previous_single_guest_photos(transaction_id):
        profile_photo_url = _build_photo_url(row.get("profile_photo_path"), external=True)
        previous_photo_url = _build_photo_url(row.get("previous_photo_path"), external=True)
        selected_url: Optional[str] = None
        selected_kind: Optional[str] = None
        if profile_photo_url:
            selected_url = profile_photo_url
            selected_kind = "profile"
        elif previous_photo_url:
            selected_url = previous_photo_url
            selected_kind = "previous"

        if not selected_url or not selected_kind:
            continue
        if current_photo_url and selected_url == current_photo_url:
            continue

        preferred_refs.append(
            {
                "kind": selected_kind,
                "url": selected_url,
                "guest_name": row.get("guest_name"),
            }
        )

    total_refs = len(preferred_refs)
    for idx, row in enumerate(preferred_refs, start=1):
        text = _format_guest_reference_button_label(
            kind=row.get("kind") or "previous",
            index=idx,
            total_items=total_refs,
            guest_name=row.get("guest_name"),
        )
        key = (text, row["url"])
        if key in seen_keys:
            continue
        seen_keys.add(key)
        links.append({"text": text, "url": row["url"]})

    return links


def _extract_guest_names_from_detail(detail: Optional[dict]) -> list[str]:
    names: list[str] = []
    seen: set[str] = set()
    for row in (detail or {}).get("guests") or []:
        full_name = str((row or {}).get("full_name") or "").strip()
        if not full_name:
            continue
        dedupe_key = full_name.casefold()
        if dedupe_key in seen:
            continue
        seen.add(dedupe_key)
        names.append(full_name)
    return names


def _format_date_dmy(value: Optional[datetime | date]) -> str:
    if not value:
        return ""
    if isinstance(value, datetime):
        value = value.date()
    return value.strftime("%d/%m/%Y")


def _build_csv_response(headers: list[str], rows: list[list[object]], filename: str) -> Response:
    buffer = StringIO()
    writer = csv.writer(buffer, delimiter=";", quoting=csv.QUOTE_MINIMAL, lineterminator="\n")
    writer.writerow(headers)
    for row in rows:
        writer.writerow(row)
    response = Response(buffer.getvalue(), mimetype="text/csv; charset=utf-8")
    response.headers["Content-Disposition"] = f"attachment; filename={filename}"
    return response


def _build_xlsx_response(
    headers: list[str],
    rows: list[list[object]],
    filename: str,
    fill_ranges: Optional[list[tuple[int, int, int, str]]] = None,
) -> Response:
    try:
        from openpyxl import Workbook
        from openpyxl.styles import PatternFill
    except ImportError as exc:
        return Response("Library openpyxl belum terinstall.", status=500)

    output = BytesIO()
    wb = Workbook()
    ws = wb.active
    ws.title = "Data"
    ws.append(headers)
    for row in rows:
        ws.append(list(row))
    if fill_ranges:
        for row_index, start_col, end_col, fill_color in fill_ranges:
            fill = PatternFill(fill_type="solid", fgColor=fill_color)
            for col_index in range(start_col, end_col + 1):
                ws.cell(row=row_index, column=col_index).fill = fill
    wb.save(output)
    output.seek(0)
    return send_file(
        output,
        as_attachment=True,
        download_name=filename,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


@daftar_tamu_bp.route("/media/photo-thumb/<path:filename>")
@role_required("admin", "sekolah", "staff", "coordinator")
def guestbook_photo_thumb(filename: str) -> Response:
    """Serve resized WEBP thumbnail with cache headers for mobile timeline."""
    safe_width = max(120, min(_to_int(request.args.get("w"), 360), 1200))
    safe_quality = max(40, min(_to_int(request.args.get("q"), 72), 90))
    image_path = _resolve_portal_upload_path(filename)
    if not image_path:
        return Response("Foto tidak ditemukan.", status=404)

    stat = image_path.stat()
    etag_token = hashlib.sha1(
        f"{image_path.as_posix()}:{stat.st_mtime_ns}:{stat.st_size}:{safe_width}:{safe_quality}".encode("utf-8")
    ).hexdigest()
    etag_value = f'W/"{etag_token}"'
    if request.headers.get("If-None-Match") == etag_value:
        response = Response(status=304)
        response.headers["ETag"] = etag_value
        response.headers["Cache-Control"] = "public, max-age=2592000, stale-while-revalidate=604800"
        return response

    try:
        with Image.open(image_path) as original:
            source = original.convert("RGBA") if original.mode not in {"RGB", "RGBA"} else original.copy()
            if source.width > safe_width:
                target_height = max(1, int((safe_width / source.width) * source.height))
                source = source.resize((safe_width, target_height), Image.LANCZOS)
            output = BytesIO()
            source.save(output, format="WEBP", quality=safe_quality, method=6)
            output.seek(0)
    except Exception:
        current_app.logger.exception("Gagal membuat thumbnail foto daftar tamu.")
        return Response("Gagal memproses foto.", status=500)

    response = send_file(output, mimetype="image/webp")
    response.headers["ETag"] = etag_value
    response.headers["Cache-Control"] = "public, max-age=2592000, stale-while-revalidate=604800"
    response.headers["Content-Disposition"] = "inline"
    return response


def _build_area_contacts(school: Optional[dict], message: Optional[str] = None) -> list[dict]:
    if not school:
        return []
    area_name = (school.get("kecamatan_name") or "").strip()
    if not message:
        message = (
            f"Halo, kami dari {school.get('name')} (NPSN {school.get('npsn')}) "
            "baru mengisi buku tamu dan mohon bantuan percepatan verifikasi."
        )
    contacts: list[dict] = []
    for row in list_portal_kontak():
        area = (row.get("wilayah") or "").strip()
        if not area:
            continue
        is_user_area = bool(area_name) and area.lower() in area_name.lower()
        for idx, (name_key, phone_key, active_key) in enumerate(
            (
                ("nama", "kontak", "kontak_1_active"),
                ("nama_2", "kontak_2", "kontak_2_active"),
            ),
            start=1,
        ):
            name = (row.get(name_key) or "").strip()
            phone = (row.get(phone_key) or "").strip()
            if not name and not phone:
                continue
            is_active = row.get(active_key)
            if is_active is None:
                is_active = True
            phone_for_link = _sanitize_phone(phone)
            if not phone_for_link:
                continue
            contacts.append(
                {
                    "area": area,
                    "name": name,
                    "phone": phone,
                    "wa_link": f"https://api.whatsapp.com/send?phone={phone_for_link}&text={quote_plus(message)}",
                    "normalized_area": area.lower(),
                    "is_user_area": is_user_area,
                    "is_active": bool(is_active),
                    "contact_index": idx,
                }
            )
    # Prioritize contacts in the same area
    has_match = any(c["is_user_area"] for c in contacts)
    if has_match:
        contacts = [c for c in contacts if c["is_user_area"]]
    contacts.sort(key=lambda c: (not c["is_user_area"], c["area"], c["contact_index"]))
    return contacts


@daftar_tamu_bp.context_processor
def inject_daftar_tamu_context() -> dict:
    """Inject shared portal-like context into daftar tamu templates."""
    user = current_user()
    if not user:
        return {}

    admin_pending = {
        "pending_users": 0,
        "pending_assignment_requests": 0,
        "pending_team_member_requests": 0,
        "pending_reopen_requests": 0,
        "pending_guestbook": 0,
        "total": 0,
    }
    user_app_notifications = {
        "unread_count": 0,
        "total_count": 0,
    }
    undo_window_seconds = PORTAL_UNDO_WINDOW_DEFAULT_SECONDS
    if user.get("role") == "admin":
        try:
            admin_pending = fetch_admin_pending_summary()
        except Exception:
            pass
    elif user.get("role") in {"staff", "coordinator", "sekolah"}:
        try:
            user_app_notifications = fetch_user_notification_summary(
                user_id=int(user.get("id")),
                categories=list(USER_APP_NOTIFICATION_CATEGORIES),
            )
        except Exception:
            user_app_notifications = {"unread_count": 0, "total_count": 0}
    try:
        undo_window_seconds = fetch_portal_undo_window_seconds()
    except Exception:
        undo_window_seconds = PORTAL_UNDO_WINDOW_DEFAULT_SECONDS

    context = {
        "permissions": get_permission_summary(user),
        "is_superadmin": is_superadmin(user),
        "can_access_aska": can_access_aska(user),
        "admin_pending": admin_pending,
        "user_app_notifications": user_app_notifications,
        "undo_window_seconds": undo_window_seconds,
        "undo_window_min_seconds": PORTAL_UNDO_WINDOW_MIN_SECONDS,
        "undo_window_max_seconds": PORTAL_UNDO_WINDOW_MAX_SECONDS,
    }
    if user.get("role") == "sekolah":
        school = _fetch_school_for_user(user.get("id"))
        context["user_school"] = school
        context["area_contacts"] = _build_area_contacts(school)
        context["school_pending"] = {"pending_sudin": 0, "pending_public": 0}
        if school:
            try:
                context["school_pending"] = fetch_school_pending_counts(school_id=school.get("id"))
            except Exception:
                context["school_pending"] = {"pending_sudin": 0, "pending_public": 0}
    return context


# ===============================
# Admin Dashboard
# ===============================

@daftar_tamu_bp.route("/admin/dashboard")
@role_required("admin", "coordinator")
def admin_dashboard() -> Response:
    """Render admin monitoring dashboard for school guest visits."""
    ensure_daftar_tamu_seed_data()
    user = current_user() or {}
    role_value = (user.get("role") or "").strip().lower()
    is_coordinator_dashboard = role_value == "coordinator"
    dashboard_kecamatan_ids = _resolve_dashboard_kecamatan_ids(user) if is_coordinator_dashboard else None

    date_from = _parse_iso_date(request.args.get("date_from"))
    date_to = _parse_iso_date(request.args.get("date_to"))
    if date_from and date_to and date_from > date_to:
        date_from, date_to = date_to, date_from

    search_query = (request.args.get("q") or "").strip()
    sort = (request.args.get("sort") or DEFAULT_SORT).strip().lower()
    if sort not in SORT_OPTIONS:
        sort = DEFAULT_SORT

    per_page = _to_int(request.args.get("per_page"), 10)
    per_page = max(5, min(per_page, 100))

    page = _to_int(request.args.get("page"), 1)
    page = max(1, page)

    guest_scope = _parse_guest_scope(request.args.get("guest_scope"))
    school_status = _parse_school_status(request.args.get("school_status"), default="all")
    show_user_rankings = (guest_scope != "umum") and not is_coordinator_dashboard
    user_rank_guest_scope = "sudin" if guest_scope == "all" else guest_scope
    user_rank_scope_label = "SUDIN" if user_rank_guest_scope == "sudin" else "Umum"
    summary = fetch_dashboard_summary(
        date_from=date_from,
        date_to=date_to,
        guest_scope=guest_scope,
        school_status=school_status,
        kecamatan_ids=dashboard_kecamatan_ids,
    )
    visit_histogram = fetch_school_visit_histogram(
        date_from=date_from,
        date_to=date_to,
        guest_scope=guest_scope,
        school_status=school_status,
        kecamatan_ids=dashboard_kecamatan_ids,
    )

    def _count_histogram_bucket(min_visits: int, max_visits: Optional[int]) -> int:
        total = 0
        for visit_count, school_count in visit_histogram.items():
            if visit_count < min_visits:
                continue
            if max_visits is not None and visit_count > max_visits:
                continue
            total += int(school_count or 0)
        return total

    visit_dist_labels = [bucket["label"] for bucket in _VISIT_DISTRIBUTION_BUCKETS]
    visit_dist_values = [
        _count_histogram_bucket(
            int(bucket.get("min_visits") or 0),
            int(bucket["max_visits"]) if bucket.get("max_visits") is not None else None,
        )
        for bucket in _VISIT_DISTRIBUTION_BUCKETS
    ]
    visit_frequency_groups = [
        {
            "key": bucket["key"],
            "label": bucket["label"],
            "school_count": _count_histogram_bucket(
                int(bucket.get("min_visits") or 0),
                int(bucket["max_visits"]) if bucket.get("max_visits") is not None else None,
            ),
        }
        for bucket in _VISIT_FREQUENCY_BUCKETS
    ]

    top_visit_rankings, _ = fetch_school_rankings(
        page=1,
        per_page=10,
        sort_key="visits_desc",
        search_query="",
        date_from=date_from,
        date_to=date_to,
        guest_scope=guest_scope,
        school_status=school_status,
        kecamatan_ids=dashboard_kecamatan_ids,
    )
    bottom_visit_rankings, _ = fetch_school_rankings(
        page=1,
        per_page=10,
        sort_key="visits_asc",
        search_query="",
        date_from=date_from,
        date_to=date_to,
        guest_scope=guest_scope,
        school_status=school_status,
        kecamatan_ids=dashboard_kecamatan_ids,
    )

    rankings, total_rows = fetch_school_rankings(
        page=page,
        per_page=per_page,
        sort_key=sort,
        search_query=search_query,
        date_from=date_from,
        date_to=date_to,
        guest_scope=guest_scope,
        school_status=school_status,
        kecamatan_ids=dashboard_kecamatan_ids,
    )
    user_search_query = (request.args.get("user_q") or "").strip()
    user_sort = (request.args.get("user_sort") or "").strip().lower() or "visits_desc"
    user_per_page = _to_int(request.args.get("user_per_page"), 10)
    user_per_page = max(5, min(user_per_page, 100))
    user_page = _to_int(request.args.get("user_page"), 1)
    user_page = max(1, user_page)

    user_rankings = []
    user_total_rows = 0
    user_total_pages = 1
    if show_user_rankings:
        user_rankings, user_total_rows = fetch_user_rankings(
            page=user_page,
            per_page=user_per_page,
            sort_key=user_sort,
            search_query=user_search_query,
            date_from=date_from,
            date_to=date_to,
            guest_scope=user_rank_guest_scope,
            school_status=school_status,
        )
        user_total_pages = max(1, math.ceil(user_total_rows / user_per_page)) if user_total_rows else 1
        if user_page > user_total_pages:
            user_page = user_total_pages
            user_rankings, user_total_rows = fetch_user_rankings(
                page=user_page,
                per_page=user_per_page,
                sort_key=user_sort,
                search_query=user_search_query,
                date_from=date_from,
                date_to=date_to,
                guest_scope=user_rank_guest_scope,
                school_status=school_status,
            )

    total_pages = max(1, math.ceil(total_rows / per_page)) if total_rows else 1
    if page > total_pages:
        page = total_pages
        rankings, total_rows = fetch_school_rankings(
            page=page,
            per_page=per_page,
            sort_key=sort,
            search_query=search_query,
            date_from=date_from,
            date_to=date_to,
            guest_scope=guest_scope,
            school_status=school_status,
            kecamatan_ids=dashboard_kecamatan_ids,
        )

    unvisited_schools = fetch_unvisited_schools(
        limit=10,
        date_from=date_from,
        date_to=date_to,
        guest_scope=guest_scope,
        school_status=school_status,
        kecamatan_ids=dashboard_kecamatan_ids,
    )
    recent_visits = fetch_recent_visits(
        limit=8,
        date_from=date_from,
        date_to=date_to,
        guest_scope=guest_scope,
        school_status=school_status,
        kecamatan_ids=dashboard_kecamatan_ids,
    )

    date_from_str = date_from.isoformat() if date_from else ""
    date_to_str = date_to.isoformat() if date_to else ""

    return render_template(
        "daftar_tamu/admin_dashboard.html",
        summary=summary,
        visit_dist_labels=visit_dist_labels,
        visit_dist_values=visit_dist_values,
        visit_frequency_groups=visit_frequency_groups,
        top_visit_rankings=top_visit_rankings,
        bottom_visit_rankings=bottom_visit_rankings,
        rankings=rankings,
        user_rankings=user_rankings,
        user_total_rows=user_total_rows,
        user_total_pages=user_total_pages,
        user_page=user_page,
        user_per_page=user_per_page,
        user_search_query=user_search_query,
        user_sort=user_sort,
        show_user_rankings=show_user_rankings,
        user_rank_guest_scope=user_rank_guest_scope,
        user_rank_scope_label=user_rank_scope_label,
        unvisited_schools=unvisited_schools,
        recent_visits=recent_visits,
        page=page,
        per_page=per_page,
        total_rows=total_rows,
        total_pages=total_pages,
        sort=sort,
        search_query=search_query,
        date_from_str=date_from_str,
        date_to_str=date_to_str,
        today_str=_today_jakarta().isoformat(),
        guest_scope=guest_scope,
        school_status=school_status,
        dashboard_scope_role="coordinator" if is_coordinator_dashboard else "admin",
    )


@daftar_tamu_bp.route("/admin/user/<int:user_id>/riwayat")
@role_required("admin")
def admin_user_history(user_id: int) -> Response:
    user_profile = _fetch_dashboard_user(user_id)
    if not user_profile:
        return render_template(
            "daftar_tamu/admin_user_history.html",
            user_profile=None,
            rows=[],
            total_rows=0,
            total_pages=1,
            page=1,
            per_page=10,
            date_from_str="",
            date_to_str="",
            guest_scope="all",
            school_status="all",
            today_str=_today_jakarta().isoformat(),
            assigned_schools=[],
            assigned_kecamatan=[],
            unvisited_schools=[],
            error_message="User tidak ditemukan.",
        ), 404

    date_from = _parse_iso_date(request.args.get("date_from"))
    date_to = _parse_iso_date(request.args.get("date_to"))
    if date_from and date_to and date_from > date_to:
        date_from, date_to = date_to, date_from
    guest_scope = _parse_guest_scope(request.args.get("guest_scope"))
    school_status = _parse_school_status(request.args.get("school_status"), default="all")

    per_page = _to_int(request.args.get("per_page"), 10)
    per_page = max(5, min(per_page, 100))

    page = _to_int(request.args.get("page"), 1)
    page = max(1, page)

    rows, total_rows = list_user_transactions(
        user_id=user_id,
        date_from=date_from,
        date_to=date_to,
        guest_scope=guest_scope,
        school_status=school_status,
        page=page,
        per_page=per_page,
    )

    total_pages = max(1, math.ceil(total_rows / per_page)) if total_rows else 1
    if page > total_pages:
        page = total_pages
        rows, total_rows = list_user_transactions(
            user_id=user_id,
            date_from=date_from,
            date_to=date_to,
            guest_scope=guest_scope,
            school_status=school_status,
            page=page,
            per_page=per_page,
        )

    hide_assignments = True
    assigned_schools = []
    assigned_kecamatan = []
    unvisited_schools = []

    return render_template(
        "daftar_tamu/admin_user_history.html",
        user_profile=user_profile,
        rows=rows,
        total_rows=total_rows,
        total_pages=total_pages,
        page=page,
        per_page=per_page,
        date_from_str=date_from.isoformat() if date_from else "",
        date_to_str=date_to.isoformat() if date_to else "",
        guest_scope=guest_scope,
        school_status=school_status,
        today_str=_today_jakarta().isoformat(),
        assigned_schools=assigned_schools,
        assigned_kecamatan=assigned_kecamatan,
        unvisited_schools=unvisited_schools,
        hide_assignments=hide_assignments,
        error_message=None,
    )


@daftar_tamu_bp.route("/admin/user/<int:user_id>/visits")
@role_required("admin")
def admin_user_visits(user_id: int) -> Response:
    """Return visit history rows for modal on admin dashboard."""
    user_profile = _fetch_dashboard_user(user_id)
    if not user_profile:
        return jsonify({"success": False, "message": "User tidak ditemukan."}), 404

    page = _to_int(request.args.get("page"), 1)
    page = max(1, page)
    per_page = _to_int(request.args.get("per_page"), 10)
    per_page = max(5, min(per_page, 100))
    sort = (request.args.get("sort") or "").strip().lower() or "date_desc"
    search_query = (request.args.get("q") or "").strip()
    guest_scope = _parse_guest_scope(request.args.get("guest_scope"))
    school_status = _parse_school_status(request.args.get("school_status"))
    date_from = _parse_iso_date(request.args.get("date_from"))
    date_to = _parse_iso_date(request.args.get("date_to"))
    if date_from and date_to and date_from > date_to:
        date_from, date_to = date_to, date_from

    rows, total_rows = fetch_user_visit_history(
        user_id=user_id,
        page=page,
        per_page=per_page,
        sort_key=sort,
        search_query=search_query,
        date_from=date_from,
        date_to=date_to,
        guest_scope=guest_scope,
        school_status=school_status,
    )

    total_pages = max(1, math.ceil(total_rows / per_page)) if total_rows else 1
    if page > total_pages:
        page = total_pages
        rows, total_rows = fetch_user_visit_history(
            user_id=user_id,
            page=page,
            per_page=per_page,
            sort_key=sort,
            search_query=search_query,
            date_from=date_from,
            date_to=date_to,
            guest_scope=guest_scope,
            school_status=school_status,
        )

    for row in rows:
        visit_at = row.get("visit_at")
        row["visit_at"] = visit_at.isoformat() if visit_at else None
        row["photo_url"] = _build_photo_url(row.get("photo_path"))

    return jsonify(
        {
            "success": True,
            "rows": rows,
            "total_rows": total_rows,
            "total_pages": total_pages,
            "page": page,
            "per_page": per_page,
            "sort": sort,
        }
    )


@daftar_tamu_bp.route("/admin/sekolah/<int:school_id>/visits")
@role_required("admin", "coordinator")
def admin_school_visits(school_id: int) -> Response:
    """Return visit history rows for school modal on admin dashboard."""
    user = current_user() or {}
    dashboard_kecamatan_ids = _resolve_dashboard_kecamatan_ids(user)
    school = _fetch_school_profile(school_id)
    if not school:
        return jsonify({"success": False, "message": "Sekolah tidak ditemukan."}), 404
    if not _can_access_school_for_dashboard(
        user=user,
        school=school,
        kecamatan_ids=dashboard_kecamatan_ids,
    ):
        return jsonify({"success": False, "message": "Sekolah di luar lokasi unit kerja."}), 403

    date_from = _parse_iso_date(request.args.get("date_from"))
    date_to = _parse_iso_date(request.args.get("date_to"))
    if date_from and date_to and date_from > date_to:
        date_from, date_to = date_to, date_from

    page = _to_int(request.args.get("page"), 1)
    page = max(1, page)
    per_page = _to_int(request.args.get("per_page"), 10)
    per_page = max(5, min(per_page, 100))
    sort = (request.args.get("sort") or "").strip().lower() or "date_desc"
    if sort not in {"date_desc", "date_asc"}:
        sort = "date_desc"
    search_query = (request.args.get("q") or "").strip()
    guest_scope = _parse_guest_scope(request.args.get("guest_scope"))

    rows, total_rows = fetch_school_visit_history(
        school_id=school_id,
        page=page,
        per_page=per_page,
        sort_key=sort,
        search_query=search_query,
        date_from=date_from,
        date_to=date_to,
        guest_scope=guest_scope,
    )

    total_pages = max(1, math.ceil(total_rows / per_page)) if total_rows else 1
    if page > total_pages:
        page = total_pages
        rows, total_rows = fetch_school_visit_history(
            school_id=school_id,
            page=page,
            per_page=per_page,
            sort_key=sort,
            search_query=search_query,
            date_from=date_from,
            date_to=date_to,
            guest_scope=guest_scope,
        )

    for row in rows:
        visit_at = row.get("visit_at")
        row["visit_at"] = visit_at.isoformat() if visit_at else None
        row["photo_url"] = _build_photo_url(row.get("photo_path"))

    return jsonify(
        {
            "success": True,
            "rows": rows,
            "total_rows": total_rows,
            "total_pages": total_pages,
            "page": page,
            "per_page": per_page,
            "sort": sort,
        }
    )


@daftar_tamu_bp.route("/admin/sekolah/<int:school_id>/visit-days")
@role_required("admin")
def admin_school_visit_days(school_id: int) -> Response:
    """Return distinct visit dates for the school day drill-down modal."""
    school = _fetch_school_profile(school_id)
    if not school:
        return jsonify({"success": False, "message": "Sekolah tidak ditemukan."}), 404

    date_from = _parse_iso_date(request.args.get("date_from"))
    date_to = _parse_iso_date(request.args.get("date_to"))
    if date_from and date_to and date_from > date_to:
        date_from, date_to = date_to, date_from

    page = _to_int(request.args.get("page"), 1)
    page = max(1, page)
    per_page = _to_int(request.args.get("per_page"), 10)
    per_page = max(5, min(per_page, 100))
    search_query = (request.args.get("q") or "").strip()
    guest_scope = _parse_guest_scope(request.args.get("guest_scope"))

    rows, total_rows = fetch_school_visit_days(
        school_id=school_id,
        page=page,
        per_page=per_page,
        search_query=search_query,
        date_from=date_from,
        date_to=date_to,
        guest_scope=guest_scope,
    )

    total_pages = max(1, math.ceil(total_rows / per_page)) if total_rows else 1
    if page > total_pages:
        page = total_pages
        rows, total_rows = fetch_school_visit_days(
            school_id=school_id,
            page=page,
            per_page=per_page,
            search_query=search_query,
            date_from=date_from,
            date_to=date_to,
            guest_scope=guest_scope,
        )

    for row in rows:
        visit_date = row.get("visit_date")
        row["visit_date"] = visit_date.isoformat() if visit_date else None

    return jsonify(
        {
            "success": True,
            "rows": rows,
            "total_rows": total_rows,
            "total_pages": total_pages,
            "page": page,
            "per_page": per_page,
        }
    )


@daftar_tamu_bp.route("/admin/sekolah/<int:school_id>/visit-days/<visit_date>/guests")
@role_required("admin")
def admin_school_visit_day_guests(school_id: int, visit_date: str) -> Response:
    """Return guest names for a selected school visit date."""
    school = _fetch_school_profile(school_id)
    if not school:
        return jsonify({"success": False, "message": "Sekolah tidak ditemukan."}), 404

    parsed_visit_date = _parse_iso_date(visit_date)
    if not parsed_visit_date:
        return jsonify({"success": False, "message": "Tanggal kunjungan tidak valid."}), 400

    page = _to_int(request.args.get("page"), 1)
    page = max(1, page)
    per_page = _to_int(request.args.get("per_page"), 10)
    per_page = max(5, min(per_page, 100))
    search_query = (request.args.get("q") or "").strip()
    guest_scope = _parse_guest_scope(request.args.get("guest_scope"))

    rows, total_rows = fetch_school_visit_day_guests(
        school_id=school_id,
        visit_date=parsed_visit_date,
        page=page,
        per_page=per_page,
        search_query=search_query,
        guest_scope=guest_scope,
    )

    total_pages = max(1, math.ceil(total_rows / per_page)) if total_rows else 1
    if page > total_pages:
        page = total_pages
        rows, total_rows = fetch_school_visit_day_guests(
            school_id=school_id,
            visit_date=parsed_visit_date,
            page=page,
            per_page=per_page,
            search_query=search_query,
            guest_scope=guest_scope,
        )

    for row in rows:
        visit_at = row.get("visit_at")
        row["visit_at"] = visit_at.isoformat() if visit_at else None

    return jsonify(
        {
            "success": True,
            "rows": rows,
            "total_rows": total_rows,
            "total_pages": total_pages,
            "page": page,
            "per_page": per_page,
            "visit_date": parsed_visit_date.isoformat(),
        }
    )


@daftar_tamu_bp.route("/admin/user/<int:user_id>/visits/export")
@role_required("admin")
def admin_user_visits_export(user_id: int) -> Response:
    """Export user visit history (Excel-friendly CSV)."""
    user_profile = _fetch_dashboard_user(user_id)
    if not user_profile:
        return Response("User tidak ditemukan.", status=404)

    date_from = _parse_iso_date(request.args.get("date_from"))
    date_to = _parse_iso_date(request.args.get("date_to"))
    if date_from and date_to and date_from > date_to:
        date_from, date_to = date_to, date_from

    search_query = (request.args.get("q") or "").strip()
    guest_scope = _parse_guest_scope(request.args.get("guest_scope"))
    school_status = _parse_school_status(request.args.get("school_status"))
    sort = (request.args.get("sort") or "").strip().lower() or "date_desc"
    if sort not in {"date_desc", "date_asc"}:
        sort = "date_desc"

    per_page = 100
    rows, total_rows = fetch_user_visit_history(
        user_id=user_id,
        page=1,
        per_page=per_page,
        sort_key=sort,
        search_query=search_query,
        date_from=date_from,
        date_to=date_to,
        guest_scope=guest_scope,
        school_status=school_status,
    )
    total_pages = max(1, math.ceil(total_rows / per_page)) if total_rows else 1
    if total_pages > 1:
        for page in range(2, total_pages + 1):
            page_rows, _ = fetch_user_visit_history(
                user_id=user_id,
                page=page,
                per_page=per_page,
                sort_key=sort,
                search_query=search_query,
                date_from=date_from,
                date_to=date_to,
                guest_scope=guest_scope,
                school_status=school_status,
            )
            rows.extend(page_rows)

    headers = [
        "Nama Sekolah",
        "Tanggal Kunjungan",
        "Tujuan",
        "Foto",
    ]
    data_rows: list[list[object]] = []
    for row in rows:
        photo_flag = "Ada" if row.get("photo_path") else "-"
        data_rows.append(
            [
                row.get("school_name") or "",
                _format_date_dmy(row.get("visit_at")),
                row.get("purpose") or "",
                photo_flag,
            ]
        )

    file_format = (request.args.get("format") or "excel").strip().lower()
    if file_format in {"excel", "xlsx"}:
        filename = f"riwayat_kunjungan_user_{user_id}_{_today_jakarta().isoformat()}.xlsx"
        return _build_xlsx_response(headers, data_rows, filename)

    filename = f"riwayat_kunjungan_user_{user_id}_{_today_jakarta().isoformat()}.csv"
    return _build_csv_response(headers, data_rows, filename)


@daftar_tamu_bp.route("/admin/sekolah/<int:school_id>/visits/export")
@role_required("admin", "coordinator")
def admin_school_visits_export(school_id: int) -> Response:
    """Export school visit history (Excel-friendly CSV)."""
    user = current_user() or {}
    dashboard_kecamatan_ids = _resolve_dashboard_kecamatan_ids(user)
    school = _fetch_school_profile(school_id)
    if not school:
        return Response("Sekolah tidak ditemukan.", status=404)
    if not _can_access_school_for_dashboard(
        user=user,
        school=school,
        kecamatan_ids=dashboard_kecamatan_ids,
    ):
        return Response("Sekolah di luar lokasi unit kerja.", status=403)

    date_from = _parse_iso_date(request.args.get("date_from"))
    date_to = _parse_iso_date(request.args.get("date_to"))
    if date_from and date_to and date_from > date_to:
        date_from, date_to = date_to, date_from

    search_query = (request.args.get("q") or "").strip()
    guest_scope = _parse_guest_scope(request.args.get("guest_scope"))
    sort = (request.args.get("sort") or "").strip().lower() or "date_desc"
    if sort not in {"date_desc", "date_asc"}:
        sort = "date_desc"

    per_page = 100
    rows, total_rows = fetch_school_visit_history(
        school_id=school_id,
        page=1,
        per_page=per_page,
        sort_key=sort,
        search_query=search_query,
        date_from=date_from,
        date_to=date_to,
        guest_scope=guest_scope,
    )
    total_pages = max(1, math.ceil(total_rows / per_page)) if total_rows else 1
    if total_pages > 1:
        for page in range(2, total_pages + 1):
            page_rows, _ = fetch_school_visit_history(
                school_id=school_id,
                page=page,
                per_page=per_page,
                sort_key=sort,
                search_query=search_query,
                date_from=date_from,
                date_to=date_to,
                guest_scope=guest_scope,
            )
            rows.extend(page_rows)

    headers = [
        "Nama Tamu Terakhir",
        "Tanggal Kunjungan",
        "Tujuan",
        "Foto",
    ]
    data_rows: list[list[object]] = []
    for row in rows:
        photo_flag = "Ada" if row.get("photo_path") else "-"
        data_rows.append(
            [
                row.get("guest_display") or row.get("guest_names") or "",
                _format_date_dmy(row.get("visit_at")),
                row.get("purpose") or "",
                photo_flag,
            ]
        )

    file_format = (request.args.get("format") or "excel").strip().lower()
    if file_format in {"excel", "xlsx"}:
        filename = f"riwayat_kunjungan_sekolah_{school_id}_{_today_jakarta().isoformat()}.xlsx"
        return _build_xlsx_response(headers, data_rows, filename)

    filename = f"riwayat_kunjungan_sekolah_{school_id}_{_today_jakarta().isoformat()}.csv"
    return _build_csv_response(headers, data_rows, filename)


@daftar_tamu_bp.route("/admin/map-data")
@role_required("admin")
def admin_map_data() -> Response:
    """Return map dots for daftar tamu dashboard."""
    ensure_daftar_tamu_seed_data()
    date_from = _parse_iso_date(request.args.get("date_from"))
    date_to = _parse_iso_date(request.args.get("date_to"))
    if date_from and date_to and date_from > date_to:
        date_from, date_to = date_to, date_from
    guest_scope = _parse_guest_scope(request.args.get("guest_scope"))
    school_status = _parse_school_status(request.args.get("school_status"))
    return jsonify(
        fetch_map_data(
            date_from=date_from,
            date_to=date_to,
            guest_scope=guest_scope,
            school_status=school_status,
        )
    )


@daftar_tamu_bp.route("/admin/rankings/more")
@role_required("admin", "coordinator")
def admin_rankings_more() -> Response:
    """Load more school rankings for dashboard card."""
    ensure_daftar_tamu_seed_data()
    user = current_user() or {}
    dashboard_kecamatan_ids = _resolve_dashboard_kecamatan_ids(user)

    date_from = _parse_iso_date(request.args.get("date_from"))
    date_to = _parse_iso_date(request.args.get("date_to"))
    if date_from and date_to and date_from > date_to:
        date_from, date_to = date_to, date_from

    guest_scope = _parse_guest_scope(request.args.get("guest_scope"))
    school_status = _parse_school_status(request.args.get("school_status"), default="all")
    ranking_type = (request.args.get("type") or "best").strip().lower()
    if ranking_type not in {"best", "worst"}:
        ranking_type = "best"
    sort_key = "visits_desc" if ranking_type == "best" else "visits_asc"

    limit = max(1, min(_to_int(request.args.get("limit"), 10), 50))
    offset = max(0, _to_int(request.args.get("offset"), 0))
    page = (offset // limit) + 1
    skip = offset % limit

    rows, _ = fetch_school_rankings(
        page=page,
        per_page=limit + skip,
        sort_key=sort_key,
        search_query="",
        date_from=date_from,
        date_to=date_to,
        guest_scope=guest_scope,
        school_status=school_status,
        kecamatan_ids=dashboard_kecamatan_ids,
    )
    if skip:
        rows = rows[skip:]
    rows = rows[:limit]

    payload = [
        {
            "rank": row.get("rank"),
            "name": row.get("school_name") or "",
            "jenjang": row.get("jenjang") or "",
            "npsn": row.get("npsn") or "",
            "visit_count": int(row.get("visit_count") or 0),
            "people_count": int(row.get("people_count") or 0),
            "visit_day_count": int(row.get("visit_day_count") or 0),
        }
        for row in rows
    ]
    return jsonify(payload)


@daftar_tamu_bp.route("/admin/stats/visit-buckets")
@role_required("admin", "coordinator")
def admin_visit_bucket_detail() -> Response:
    """Detailed school list for selected visit-count bucket."""
    ensure_daftar_tamu_seed_data()
    user = current_user() or {}
    dashboard_kecamatan_ids = _resolve_dashboard_kecamatan_ids(user)

    date_from = _parse_iso_date(request.args.get("date_from"))
    date_to = _parse_iso_date(request.args.get("date_to"))
    if date_from and date_to and date_from > date_to:
        date_from, date_to = date_to, date_from

    guest_scope = _parse_guest_scope(request.args.get("guest_scope"))
    school_status = _parse_school_status(request.args.get("school_status"), default="all")

    source = (request.args.get("source") or "distribution").strip().lower()
    if source not in {"distribution", "frequency"}:
        source = "distribution"
    bucket_options = _VISIT_DISTRIBUTION_BUCKETS if source == "distribution" else _VISIT_FREQUENCY_BUCKETS

    selected_bucket_key = (request.args.get("bucket") or "").strip().lower()
    selected_bucket = next((row for row in bucket_options if row.get("key") == selected_bucket_key), None)
    if not selected_bucket:
        selected_bucket = bucket_options[0]

    sort = (request.args.get("sort") or "visits_desc").strip().lower()
    if sort not in {"visits_desc", "visits_asc", "days_desc", "days_asc", "name_asc", "name_desc", "last_visit_desc", "last_visit_asc"}:
        sort = DEFAULT_SORT

    per_page = _to_int(request.args.get("per_page"), 25)
    per_page = max(10, min(per_page, 100))

    page = _to_int(request.args.get("page"), 1)
    page = max(1, page)

    min_visits = int(selected_bucket.get("min_visits") or 0)
    raw_max_visits = selected_bucket.get("max_visits")
    max_visits = int(raw_max_visits) if raw_max_visits is not None else None

    rows, total_rows = fetch_school_visit_bucket_rows(
        min_visits=min_visits,
        max_visits=max_visits,
        page=page,
        per_page=per_page,
        sort_key=sort,
        date_from=date_from,
        date_to=date_to,
        guest_scope=guest_scope,
        school_status=school_status,
        kecamatan_ids=dashboard_kecamatan_ids,
    )
    total_pages = max(1, math.ceil(total_rows / per_page)) if total_rows else 1
    if page > total_pages:
        page = total_pages
        rows, total_rows = fetch_school_visit_bucket_rows(
            min_visits=min_visits,
            max_visits=max_visits,
            page=page,
            per_page=per_page,
            sort_key=sort,
            date_from=date_from,
            date_to=date_to,
            guest_scope=guest_scope,
            school_status=school_status,
            kecamatan_ids=dashboard_kecamatan_ids,
        )

    date_from_str = date_from.isoformat() if date_from else ""
    date_to_str = date_to.isoformat() if date_to else ""

    return render_template(
        "daftar_tamu/admin_visit_bucket_detail.html",
        source=source,
        bucket_options=bucket_options,
        selected_bucket=selected_bucket,
        rows=rows,
        total_rows=total_rows,
        page=page,
        per_page=per_page,
        total_pages=total_pages,
        sort=sort,
        date_from_str=date_from_str,
        date_to_str=date_to_str,
        guest_scope=guest_scope,
        school_status=school_status,
        today_str=_today_jakarta().isoformat(),
    )


@daftar_tamu_bp.route("/admin/export")
@role_required("admin", "coordinator")
def export_rankings() -> Response:
    """Export school rankings in CSV/Excel."""
    ensure_daftar_tamu_seed_data()
    user = current_user() or {}
    dashboard_kecamatan_ids = _resolve_dashboard_kecamatan_ids(user)

    date_from = _parse_iso_date(request.args.get("date_from"))
    date_to = _parse_iso_date(request.args.get("date_to"))
    if date_from and date_to and date_from > date_to:
        date_from, date_to = date_to, date_from

    search_query = (request.args.get("q") or "").strip()
    sort = (request.args.get("sort") or DEFAULT_SORT).strip().lower()
    if sort not in SORT_OPTIONS:
        sort = DEFAULT_SORT

    guest_scope = _parse_guest_scope(request.args.get("guest_scope"))
    school_status = _parse_school_status(request.args.get("school_status"), default="all")

    rows, _ = fetch_school_rankings(
        page=1,
        per_page=10000,
        sort_key=sort,
        search_query=search_query,
        date_from=date_from,
        date_to=date_to,
        guest_scope=guest_scope,
        school_status=school_status,
        kecamatan_ids=dashboard_kecamatan_ids,
    )

    csv_headers = [
        "Peringkat",
        "NPSN",
        "Nama Sekolah",
        "Jenjang",
        "Kecamatan",
        "Kelurahan",
        "Orang",
        "Hari",
        "Kunjungan Terakhir",
        "Tamu Terakhir",
    ]
    csv_rows: list[list[object]] = []
    for row in rows:
        csv_rows.append(
            [
                row.get("rank"),
                row.get("npsn"),
                row.get("school_name"),
                row.get("jenjang"),
                row.get("kecamatan"),
                row.get("kelurahan"),
                row.get("people_count"),
                row.get("visit_day_count"),
                _format_date_dmy(row.get("last_visit_date")),
                row.get("last_guest_display") or "",
            ]
        )

    file_format = (request.args.get("format") or "csv").strip().lower()
    if file_format in {"excel", "xlsx"}:
        excel_headers = [
            "Peringkat",
            "NPSN",
            "Nama Sekolah",
            "Jenjang",
            "Kecamatan",
            "Kelurahan",
            "Kunjungan Ke",
            "Tanggal Kunjungan",
            "Nama Pengunjung",
            "Tujuan",
            "Link Foto",
            "Catatan Sekolah (opsional)",
            "Catatan Staf/Koordinator (opsional)",
        ]
        detail_rows: list[list[object]] = []
        visit_page_size = 100

        for row in rows:
            school_id = int(row.get("school_id") or 0)
            if school_id <= 0:
                continue

            school_rank = row.get("rank")
            school_npsn = row.get("npsn") or ""
            school_name = row.get("school_name") or ""
            school_jenjang = row.get("jenjang") or ""
            school_kecamatan = row.get("kecamatan") or ""
            school_kelurahan = row.get("kelurahan") or ""
            visit_count = int(row.get("visit_count") or 0)

            if visit_count <= 0:
                detail_rows.append(
                    [
                        school_rank,
                        school_npsn,
                        school_name,
                        school_jenjang,
                        school_kecamatan,
                        school_kelurahan,
                        "",
                        "",
                        "",
                        "",
                        "",
                        "",
                        "",
                    ]
                )
                continue

            visit_rows, visit_total_rows = fetch_school_visit_history(
                school_id=school_id,
                page=1,
                per_page=visit_page_size,
                sort_key="date_asc",
                search_query="",
                date_from=date_from,
                date_to=date_to,
                guest_scope=guest_scope,
            )
            visit_total_pages = max(1, math.ceil(visit_total_rows / visit_page_size)) if visit_total_rows else 1
            if visit_total_pages > 1:
                for visit_page in range(2, visit_total_pages + 1):
                    page_rows, _ = fetch_school_visit_history(
                        school_id=school_id,
                        page=visit_page,
                        per_page=visit_page_size,
                        sort_key="date_asc",
                        search_query="",
                        date_from=date_from,
                        date_to=date_to,
                        guest_scope=guest_scope,
                    )
                    visit_rows.extend(page_rows)

            for visit_index, visit in enumerate(visit_rows, start=1):
                photo_url = _build_photo_url(visit.get("photo_path"), external=True) or ""
                detail_rows.append(
                    [
                        school_rank,
                        school_npsn,
                        school_name,
                        school_jenjang,
                        school_kecamatan,
                        school_kelurahan,
                        visit_index,
                        _format_date_dmy(visit.get("visit_at")),
                        visit.get("guest_names") or visit.get("guest_display") or "",
                        visit.get("purpose") or "",
                        photo_url,
                        visit.get("notes") or "",
                        visit.get("staff_note_text") or "",
                    ]
                )

        excel_rows: list[list[object]] = []
        for row in detail_rows:
            excel_row = list(row)
            photo_url = str(excel_row[10] or "").strip()
            if photo_url:
                safe_url = photo_url.replace('"', '""')
                excel_row[10] = f'=HYPERLINK("{safe_url}","Foto")'
            else:
                excel_row[10] = ""
            excel_rows.append(excel_row)

        filename = f"ranking_daftar_tamu_{_today_jakarta().isoformat()}.xlsx"
        return _build_xlsx_response(excel_headers, excel_rows, filename)

    filename = f"ranking_daftar_tamu_{_today_jakarta().isoformat()}.csv"
    return _build_csv_response(csv_headers, csv_rows, filename)


@daftar_tamu_bp.route("/admin/export/users")
@role_required("admin")
def export_user_rankings() -> Response:
    """Export user rankings in CSV/Excel."""
    ensure_daftar_tamu_seed_data()

    date_from = _parse_iso_date(request.args.get("date_from"))
    date_to = _parse_iso_date(request.args.get("date_to"))
    if date_from and date_to and date_from > date_to:
        date_from, date_to = date_to, date_from

    user_search_query = (request.args.get("user_q") or request.args.get("q") or "").strip()
    user_sort = (request.args.get("user_sort") or DEFAULT_USER_SORT).strip().lower()
    if user_sort not in USER_SORT_OPTIONS:
        user_sort = DEFAULT_USER_SORT

    guest_scope = _parse_guest_scope(request.args.get("guest_scope"))
    user_rank_guest_scope = "sudin" if guest_scope == "all" else guest_scope
    school_status = _parse_school_status(request.args.get("school_status"))

    per_page = 100
    rows: list[dict] = []
    total_rows = 0
    if user_rank_guest_scope != "umum":
        rows, total_rows = fetch_user_rankings(
            page=1,
            per_page=per_page,
            sort_key=user_sort,
            search_query=user_search_query,
            date_from=date_from,
            date_to=date_to,
            guest_scope=user_rank_guest_scope,
            school_status=school_status,
        )
        total_pages = max(1, math.ceil(total_rows / per_page)) if total_rows else 1
        if total_pages > 1:
            for page in range(2, total_pages + 1):
                page_rows, _ = fetch_user_rankings(
                    page=page,
                    per_page=per_page,
                    sort_key=user_sort,
                    search_query=user_search_query,
                    date_from=date_from,
                    date_to=date_to,
                    guest_scope=user_rank_guest_scope,
                    school_status=school_status,
                )
                rows.extend(page_rows)

    headers = [
        "Peringkat",
        "Nama User",
        "Email",
        "Role",
        "Kunjungan Ke",
        "Tanggal Kunjungan",
        "Nama Sekolah",
        "NPSN",
        "Kecamatan",
        "Tujuan",
        "Link Foto",
        "Catatan Sekolah (opsional)",
        "Catatan Staf/Koordinator (opsional)",
    ]
    data_rows: list[list[object]] = []
    visit_page_size = 100
    for row in rows:
        user_id = int(row.get("user_id") or 0)
        if user_id <= 0:
            continue

        visit_rows, visit_total_rows = fetch_user_visit_history(
            user_id=user_id,
            page=1,
            per_page=visit_page_size,
            sort_key="date_asc",
            search_query="",
            date_from=date_from,
            date_to=date_to,
            guest_scope=user_rank_guest_scope,
            school_status=school_status,
        )
        visit_total_pages = max(1, math.ceil(visit_total_rows / visit_page_size)) if visit_total_rows else 1
        if visit_total_pages > 1:
            for visit_page in range(2, visit_total_pages + 1):
                page_rows, _ = fetch_user_visit_history(
                    user_id=user_id,
                    page=visit_page,
                    per_page=visit_page_size,
                    sort_key="date_asc",
                    search_query="",
                    date_from=date_from,
                    date_to=date_to,
                    guest_scope=user_rank_guest_scope,
                    school_status=school_status,
                )
                visit_rows.extend(page_rows)

        for visit_index, visit in enumerate(visit_rows, start=1):
            photo_url = _build_photo_url(visit.get("photo_path"), external=True) or ""
            data_rows.append(
                [
                    row.get("rank"),
                    row.get("full_name"),
                    row.get("email"),
                    row.get("role"),
                    visit_index,
                    _format_date_dmy(visit.get("visit_at")),
                    visit.get("school_name") or "",
                    visit.get("school_npsn") or "",
                    visit.get("school_kecamatan") or "",
                    visit.get("purpose") or "",
                    photo_url,
                    visit.get("notes") or "",
                    visit.get("staff_note_text") or "",
                ]
            )

    file_format = (request.args.get("format") or "excel").strip().lower()
    if file_format in {"excel", "xlsx"}:
        excel_rows: list[list[object]] = []
        for row in data_rows:
            excel_row = list(row)
            photo_url = str(excel_row[10] or "").strip()
            if photo_url:
                safe_url = photo_url.replace('"', '""')
                excel_row[10] = f'=HYPERLINK("{safe_url}","Foto")'
            else:
                excel_row[10] = ""
            excel_rows.append(excel_row)
        filename = f"ranking_user_kunjungan_{_today_jakarta().isoformat()}.xlsx"
        return _build_xlsx_response(headers, excel_rows, filename)

    filename = f"ranking_user_kunjungan_{_today_jakarta().isoformat()}.csv"
    return _build_csv_response(headers, data_rows, filename)


# ===============================
# Admin Validation
# ===============================

@daftar_tamu_bp.route("/admin/validasi")
@role_required("admin")
def admin_validation() -> Response:
    status = (request.args.get("status") or "pending").strip().lower()
    if status not in ("pending", "approved", "rejected", "history"):
        status = "pending"
    staff_note_level = _normalize_staff_note_level(request.args.get("staff_note_level"), default="")

    date_from = _parse_iso_date(request.args.get("date_from"))
    date_to = _parse_iso_date(request.args.get("date_to"))
    if date_from and date_to and date_from > date_to:
        date_from, date_to = date_to, date_from

    search_query = (request.args.get("q") or "").strip()

    per_page = _to_int(request.args.get("per_page"), 10)
    per_page = max(5, min(per_page, 100))

    page = _to_int(request.args.get("page"), 1)
    page = max(1, page)
    detail_param = _to_int(request.args.get("detail"), 0)
    open_transaction_id = detail_param if detail_param > 0 else None

    rows, total_rows = list_admin_transactions(
        status=status,
        staff_note_level=staff_note_level,
        search_query=search_query,
        date_from=date_from,
        date_to=date_to,
        page=page,
        per_page=per_page,
    )

    total_pages = max(1, math.ceil(total_rows / per_page)) if total_rows else 1
    if page > total_pages:
        page = total_pages
        rows, total_rows = list_admin_transactions(
            status=status,
            staff_note_level=staff_note_level,
            search_query=search_query,
            date_from=date_from,
            date_to=date_to,
            page=page,
            per_page=per_page,
        )

    date_from_str = date_from.isoformat() if date_from else ""
    date_to_str = date_to.isoformat() if date_to else ""

    return render_template(
        "daftar_tamu/admin_validation.html",
        rows=rows,
        status=status,
        staff_note_level=staff_note_level,
        search_query=search_query,
        page=page,
        per_page=per_page,
        total_rows=total_rows,
        total_pages=total_pages,
        date_from_str=date_from_str,
        date_to_str=date_to_str,
        today_str=_today_jakarta().isoformat(),
        open_transaction_id=open_transaction_id,
    )


@daftar_tamu_bp.route("/admin/transactions/<int:transaction_id>")
@role_required("admin")
def admin_transaction_detail(transaction_id: int) -> Response:
    detail = get_transaction_detail(transaction_id)
    if not detail:
        return jsonify({"success": False, "message": "Transaksi tidak ditemukan"}), 404

    return jsonify({"success": True, "transaction": detail})


def _get_school_umum_transaction_detail(
    *, transaction_id: int, school_id: int
) -> tuple[Optional[dict], Optional[str]]:
    detail = get_transaction_detail(transaction_id)
    if not detail or detail.get("school_id") != school_id:
        return None, "not_found"
    guests = detail.get("guests") or []
    has_umum = any((guest or {}).get("guest_type") == "umum" for guest in guests)
    if not has_umum:
        return None, "forbidden"
    return detail, None


@daftar_tamu_bp.route("/sekolah/transactions/<int:transaction_id>")
@role_required("sekolah")
def sekolah_transaction_detail(transaction_id: int) -> Response:
    user = current_user()
    school = _fetch_school_for_user(user.get("id"))
    if not school:
        return jsonify({"success": False, "message": "Akun sekolah belum terhubung."}), 400

    detail, err = _get_school_umum_transaction_detail(
        transaction_id=transaction_id,
        school_id=school.get("id"),
    )
    if err == "not_found":
        return jsonify({"success": False, "message": "Transaksi tidak ditemukan."}), 404
    if err == "forbidden":
        return jsonify({"success": False, "message": "Transaksi hanya untuk tamu Sudin."}), 403

    return jsonify({"success": True, "transaction": detail})


def _sekolah_update_transaction_status(
    *, transaction_id: int, status: str
) -> Response:
    user = current_user()
    school = _fetch_school_for_user(user.get("id"))
    if not school:
        return jsonify({"success": False, "message": "Akun sekolah belum terhubung."}), 400

    _, err = _get_school_umum_transaction_detail(
        transaction_id=transaction_id,
        school_id=school.get("id"),
    )
    if err == "not_found":
        return jsonify({"success": False, "message": "Transaksi tidak ditemukan."}), 404
    if err == "forbidden":
        return jsonify({"success": False, "message": "Transaksi hanya untuk tamu Sudin."}), 403

    note = (request.form.get("reviewer_note") or "").strip()
    if status == "rejected" and not note:
        return jsonify({"success": False, "message": "Catatan penolakan wajib diisi."}), 400

    try:
        ok = update_transaction_status(
            transaction_id=transaction_id,
            status=status,
            reviewer_id=user["id"],
            reviewer_notes=note or None,
        )
    except ValueError:
        ok = False
    if not ok:
        return jsonify({"success": False, "message": "Gagal memperbarui transaksi."}), 400

    try:
        _notify_user_app_status_change(
            transaction_id=transaction_id,
            status=status,
            actor=user,
            reviewer_notes=note or None,
        )
    except Exception:
        current_app.logger.exception("Gagal menyimpan notifikasi status buku tamu aplikasi.")

    try:
        _notify_guestbook_status_change(
            transaction_id=transaction_id,
            status=status,
            actor=user,
            is_public=False,
        )
    except Exception:
        current_app.logger.exception("Gagal mengirim notifikasi Telegram status buku tamu.")

    return jsonify({"success": True})


@daftar_tamu_bp.route("/sekolah/transactions/<int:transaction_id>/approve", methods=["POST"])
@role_required("sekolah")
def sekolah_transaction_approve(transaction_id: int) -> Response:
    return _sekolah_update_transaction_status(transaction_id=transaction_id, status="approved")


@daftar_tamu_bp.route("/sekolah/transactions/<int:transaction_id>/reject", methods=["POST"])
@role_required("sekolah")
def sekolah_transaction_reject(transaction_id: int) -> Response:
    return _sekolah_update_transaction_status(transaction_id=transaction_id, status="rejected")


@daftar_tamu_bp.route("/sekolah/transactions/<int:transaction_id>/pending", methods=["POST"])
@role_required("sekolah")
def sekolah_transaction_pending(transaction_id: int) -> Response:
    return _sekolah_update_transaction_status(transaction_id=transaction_id, status="pending")


def _guestbook_status_label(status: Optional[str]) -> str:
    normalized = (status or "").strip().lower()
    if normalized == "approved":
        return "✅ Disetujui"
    if normalized == "rejected":
        return "❌ Ditolak"
    return "⏳ Menunggu Verifikasi"


@daftar_tamu_bp.route("/public/detail/<int:transaction_id>")
def public_transaction_detail(transaction_id: int) -> Response:
    detail = get_transaction_detail(transaction_id)
    if not detail:
        return Response("Detail transaksi tidak ditemukan.", status=404)

    visit_at = to_jakarta(detail.get("visit_at"))
    reviewed_at = to_jakarta(detail.get("reviewed_at"))
    photo_url = _build_photo_url(detail.get("photo_path"), external=False)

    reviewer_name = (detail.get("reviewer_name") or "").strip() or "-"
    reviewer_telegram_username = (detail.get("reviewer_telegram_username") or "").strip().lstrip("@")
    reviewer_telegram_user_id = str(detail.get("reviewer_telegram_user_id") or "").strip()
    reviewer_telegram_parts: list[str] = []
    if reviewer_telegram_username:
        reviewer_telegram_parts.append(f"@{reviewer_telegram_username}")
    if reviewer_telegram_user_id:
        reviewer_telegram_parts.append(reviewer_telegram_user_id)

    gps_label = "-"
    raw_lat = detail.get("latitude")
    raw_lon = detail.get("longitude")
    if raw_lat is not None and raw_lon is not None:
        try:
            gps_label = f"{float(raw_lat):.5f}, {float(raw_lon):.5f}"
        except (TypeError, ValueError):
            gps_label = "-"

    guests = []
    for row in detail.get("guests") or []:
        guest_item = dict(row)
        guest_item["profile_photo_url"] = _build_photo_url(guest_item.get("profile_photo_path"), external=False)
        guests.append(guest_item)
    return render_template(
        "daftar_tamu/public_transaction_detail.html",
        transaction_id=transaction_id,
        school_name=detail.get("school_name") or "-",
        npsn=detail.get("npsn") or "-",
        jenjang=detail.get("jenjang") or "-",
        kecamatan=detail.get("kecamatan") or "-",
        kelurahan=detail.get("kelurahan") or "-",
        visit_at_label=visit_at.strftime("%d %b %Y, %H:%M") if visit_at else "-",
        purpose=detail.get("purpose") or "-",
        notes=detail.get("notes") or "-",
        status_label=_guestbook_status_label(detail.get("status")),
        created_by_name=detail.get("created_by_name") or "-",
        reviewer_name=reviewer_name,
        reviewer_telegram_label=" • ".join(reviewer_telegram_parts) if reviewer_telegram_parts else "-",
        reviewed_at_label=reviewed_at.strftime("%d %b %Y, %H:%M") if reviewed_at else "-",
        reviewer_notes=detail.get("reviewer_notes") or "-",
        gps_label=gps_label,
        photo_url=photo_url,
        guests=guests,
    )


@daftar_tamu_bp.route("/admin/transactions/<int:transaction_id>/approve", methods=["POST"])
@role_required("admin")
def admin_transaction_approve(transaction_id: int) -> Response:
    user = current_user()
    note = (request.form.get("reviewer_note") or "").strip()
    try:
        ok = update_transaction_status(
            transaction_id=transaction_id,
            status="approved",
            reviewer_id=user["id"],
            reviewer_notes=note or None,
        )
    except ValueError:
        ok = False
    if not ok:
        return jsonify({"success": False, "message": "Gagal memperbarui transaksi."}), 400
    try:
        _notify_user_app_status_change(
            transaction_id=transaction_id,
            status="approved",
            actor=user,
            reviewer_notes=note or None,
        )
    except Exception:
        current_app.logger.exception("Gagal menyimpan notifikasi status buku tamu aplikasi.")
    try:
        _notify_guestbook_status_change(
            transaction_id=transaction_id,
            status="approved",
            actor=user,
            is_public=False,
        )
    except Exception:
        current_app.logger.exception("Gagal mengirim notifikasi Telegram status buku tamu.")
    try:
        record_admin_action(
            user_id=user.get("id"),
            feature_key="daftar_tamu",
            action="VERIFY_APPROVE",
            target_type="GUESTBOOK_TRANSACTION",
            target_id=transaction_id,
            target_name=f"Transaksi #{transaction_id}",
            metadata={"status": "approved", "reviewer_note": note or None},
        )
    except Exception:
        current_app.logger.exception("Gagal mencatat action admin buku tamu.")
    return jsonify({"success": True})


@daftar_tamu_bp.route("/admin/transactions/<int:transaction_id>/reject", methods=["POST"])
@role_required("admin")
def admin_transaction_reject(transaction_id: int) -> Response:
    user = current_user()
    note = (request.form.get("reviewer_note") or "").strip()
    if not note:
        return jsonify({"success": False, "message": "Catatan penolakan wajib diisi."}), 400
    try:
        ok = update_transaction_status(
            transaction_id=transaction_id,
            status="rejected",
            reviewer_id=user["id"],
            reviewer_notes=note,
        )
    except ValueError:
        ok = False
    if not ok:
        return jsonify({"success": False, "message": "Gagal memperbarui transaksi."}), 400
    try:
        _notify_user_app_status_change(
            transaction_id=transaction_id,
            status="rejected",
            actor=user,
            reviewer_notes=note,
        )
    except Exception:
        current_app.logger.exception("Gagal menyimpan notifikasi status buku tamu aplikasi.")
    try:
        _notify_guestbook_status_change(
            transaction_id=transaction_id,
            status="rejected",
            actor=user,
            is_public=False,
        )
    except Exception:
        current_app.logger.exception("Gagal mengirim notifikasi Telegram status buku tamu.")
    try:
        record_admin_action(
            user_id=user.get("id"),
            feature_key="daftar_tamu",
            action="VERIFY_REJECT",
            target_type="GUESTBOOK_TRANSACTION",
            target_id=transaction_id,
            target_name=f"Transaksi #{transaction_id}",
            metadata={"status": "rejected", "reviewer_note": note},
        )
    except Exception:
        current_app.logger.exception("Gagal mencatat action admin buku tamu.")
    return jsonify({"success": True})


@daftar_tamu_bp.route("/admin/transactions/<int:transaction_id>/pending", methods=["POST"])
@role_required("admin")
def admin_transaction_pending(transaction_id: int) -> Response:
    user = current_user()
    note = (request.form.get("reviewer_note") or "").strip()
    try:
        ok = update_transaction_status(
            transaction_id=transaction_id,
            status="pending",
            reviewer_id=user["id"],
            reviewer_notes=note or None,
        )
    except ValueError:
        ok = False
    if not ok:
        return jsonify({"success": False, "message": "Gagal memperbarui transaksi."}), 400
    try:
        _notify_user_app_status_change(
            transaction_id=transaction_id,
            status="pending",
            actor=user,
            reviewer_notes=note or None,
        )
    except Exception:
        current_app.logger.exception("Gagal menyimpan notifikasi status buku tamu aplikasi.")
    try:
        record_admin_action(
            user_id=user.get("id"),
            feature_key="daftar_tamu",
            action="VERIFY_PENDING",
            target_type="GUESTBOOK_TRANSACTION",
            target_id=transaction_id,
            target_name=f"Transaksi #{transaction_id}",
            metadata={"status": "pending", "reviewer_note": note or None},
        )
    except Exception:
        current_app.logger.exception("Gagal mencatat action admin buku tamu.")
    return jsonify({"success": True})


@daftar_tamu_bp.route("/admin/transactions/bulk-approve", methods=["POST"])
@role_required("admin")
def admin_bulk_approve_transactions() -> Response:
    user = current_user()
    raw_ids = request.form.getlist("transaction_ids")
    ids: list[int] = []
    for raw in raw_ids:
        try:
            ids.append(int(raw))
        except (TypeError, ValueError):
            continue
    if not ids:
        flash("Pilih transaksi yang ingin disetujui.", "warning")
        return redirect(_build_admin_validation_redirect(request.form))

    note = (request.form.get("reviewer_note") or "").strip()
    success_count = 0
    for tx_id in ids:
        try:
            ok = update_transaction_status(
                transaction_id=tx_id,
                status="approved",
                reviewer_id=user["id"],
                reviewer_notes=note or None,
            )
        except Exception:
            ok = False
        if ok:
            success_count += 1
            try:
                record_admin_action(
                    user_id=user.get("id"),
                    feature_key="daftar_tamu",
                    action="VERIFY_APPROVE",
                    target_type="GUESTBOOK_TRANSACTION",
                    target_id=tx_id,
                    target_name=f"Transaksi #{tx_id}",
                    metadata={"status": "approved", "reviewer_note": note or None, "mode": "bulk"},
                )
            except Exception:
                current_app.logger.exception("Gagal mencatat bulk approve buku tamu.")
            try:
                _notify_user_app_status_change(
                    transaction_id=tx_id,
                    status="approved",
                    actor=user,
                    reviewer_notes=note or None,
                )
            except Exception:
                current_app.logger.exception("Gagal menyimpan notifikasi status buku tamu aplikasi.")
            try:
                _notify_guestbook_status_change(
                    transaction_id=tx_id,
                    status="approved",
                    actor=user,
                    is_public=False,
                )
            except Exception:
                current_app.logger.exception("Gagal mengirim notifikasi Telegram status buku tamu.")

    if success_count:
        flash(f"{success_count} transaksi berhasil disetujui.", "success")
    else:
        flash("Tidak ada transaksi yang berhasil disetujui.", "warning")

    return redirect(_build_admin_validation_redirect(request.form))


@daftar_tamu_bp.route("/admin/transactions/bulk-reject", methods=["POST"])
@role_required("admin")
def admin_bulk_reject_transactions() -> Response:
    user = current_user()
    raw_ids = request.form.getlist("transaction_ids")
    ids: list[int] = []
    for raw in raw_ids:
        try:
            ids.append(int(raw))
        except (TypeError, ValueError):
            continue
    if not ids:
        flash("Pilih transaksi yang ingin ditolak.", "warning")
        return redirect(_build_admin_validation_redirect(request.form))

    note = (request.form.get("reviewer_note") or "").strip()
    if not note:
        flash("Catatan penolakan wajib diisi.", "warning")
        return redirect(_build_admin_validation_redirect(request.form))

    success_count = 0
    for tx_id in ids:
        try:
            ok = update_transaction_status(
                transaction_id=tx_id,
                status="rejected",
                reviewer_id=user["id"],
                reviewer_notes=note,
            )
        except Exception:
            ok = False
        if ok:
            success_count += 1
            try:
                record_admin_action(
                    user_id=user.get("id"),
                    feature_key="daftar_tamu",
                    action="VERIFY_REJECT",
                    target_type="GUESTBOOK_TRANSACTION",
                    target_id=tx_id,
                    target_name=f"Transaksi #{tx_id}",
                    metadata={"status": "rejected", "reviewer_note": note, "mode": "bulk"},
                )
            except Exception:
                current_app.logger.exception("Gagal mencatat bulk reject buku tamu.")
            try:
                _notify_user_app_status_change(
                    transaction_id=tx_id,
                    status="rejected",
                    actor=user,
                    reviewer_notes=note,
                )
            except Exception:
                current_app.logger.exception("Gagal menyimpan notifikasi status buku tamu aplikasi.")
            try:
                _notify_guestbook_status_change(
                    transaction_id=tx_id,
                    status="rejected",
                    actor=user,
                    is_public=False,
                )
            except Exception:
                current_app.logger.exception("Gagal mengirim notifikasi Telegram status buku tamu.")

    if success_count:
        flash(f"{success_count} transaksi berhasil ditolak.", "success")
    else:
        flash("Tidak ada transaksi yang berhasil ditolak.", "warning")

    return redirect(_build_admin_validation_redirect(request.form))


@daftar_tamu_bp.route("/admin/transactions/bulk-pending", methods=["POST"])
@role_required("admin")
def admin_bulk_pending_transactions() -> Response:
    user = current_user()
    raw_ids = request.form.getlist("transaction_ids")
    ids: list[int] = []
    for raw in raw_ids:
        try:
            ids.append(int(raw))
        except (TypeError, ValueError):
            continue
    if not ids:
        flash("Pilih transaksi yang ingin dikembalikan ke pending.", "warning")
        return redirect(_build_admin_validation_redirect(request.form))

    note = (request.form.get("reviewer_note") or "").strip()
    success_count = 0
    for tx_id in ids:
        try:
            ok = update_transaction_status(
                transaction_id=tx_id,
                status="pending",
                reviewer_id=user["id"],
                reviewer_notes=note or None,
            )
        except Exception:
            ok = False
        if ok:
            success_count += 1
            try:
                _notify_user_app_status_change(
                    transaction_id=tx_id,
                    status="pending",
                    actor=user,
                    reviewer_notes=note or None,
                )
            except Exception:
                current_app.logger.exception("Gagal menyimpan notifikasi status buku tamu aplikasi.")

    if success_count:
        flash(f"{success_count} transaksi berhasil dikembalikan ke pending.", "success")
    else:
        flash("Tidak ada transaksi yang berhasil dikembalikan ke pending.", "warning")

    return redirect(_build_admin_validation_redirect(request.form))


def _build_admin_validation_redirect(form) -> str:
    status = (form.get("status") or "").strip() or "pending"
    staff_note_level = _normalize_staff_note_level(form.get("staff_note_level"), default="")
    search = (form.get("q") or "").strip()
    date_from = (form.get("date_from") or "").strip()
    date_to = (form.get("date_to") or "").strip()
    per_page_raw = (form.get("per_page") or "").strip()
    page_raw = (form.get("page") or "").strip()
    try:
        per_page = int(per_page_raw) if per_page_raw else None
    except (TypeError, ValueError):
        per_page = None
    try:
        page = int(page_raw) if page_raw else None
    except (TypeError, ValueError):
        page = None
    return url_for(
        "daftar_tamu.admin_validation",
        status=status,
        staff_note_level=staff_note_level,
        q=search,
        date_from=date_from,
        date_to=date_to,
        per_page=per_page,
        page=page,
    )


@daftar_tamu_bp.route("/admin/guests/<int:user_id>/history")
@role_required("admin")
def admin_guest_history(user_id: int) -> Response:
    limit = _to_int(request.args.get("limit"), 20)
    offset = _to_int(request.args.get("offset"), 0)
    rows = fetch_guest_history(user_id=user_id, limit=limit, offset=offset)
    return jsonify({"success": True, "history": rows})


@daftar_tamu_bp.route("/admin/umum")
@role_required("admin")
def admin_general_guests() -> Response:
    search_query = (request.args.get("q") or "").strip()
    verified_raw = (request.args.get("verified") or "").strip().lower()
    deleted_raw = (request.args.get("deleted") or "").strip().lower()
    verified = None
    if verified_raw == "true":
        verified = True
    elif verified_raw == "false":
        verified = False
    deleted = None
    if deleted_raw == "true":
        deleted = True
    elif deleted_raw == "false":
        deleted = False
    page = _to_int(request.args.get("page"), 1)
    per_page = _to_int(request.args.get("per_page"), 20)
    per_page = max(10, min(per_page, 100))
    rows, total_rows = list_general_guests_admin(
        search_query=search_query,
        verified=verified,
        deleted=deleted,
        page=page,
        per_page=per_page,
    )
    total_pages = max(1, math.ceil(total_rows / per_page)) if total_rows else 1
    if page > total_pages:
        page = total_pages
        rows, total_rows = list_general_guests_admin(
            search_query=search_query,
            verified=verified,
            deleted=deleted,
            page=page,
            per_page=per_page,
        )

    return render_template(
        "daftar_tamu/admin_general_guests.html",
        rows=rows,
        total_rows=total_rows,
        total_pages=total_pages,
        page=page,
        per_page=per_page,
        search_query=search_query,
        verified_filter=verified_raw,
        deleted_filter=deleted_raw,
    )


@daftar_tamu_bp.route("/admin/umum/search")
@role_required("admin")
def admin_general_guest_search() -> Response:
    query = (request.args.get("q") or "").strip()
    limit = _to_int(request.args.get("limit"), 15)
    results = list_general_guest_candidates(query, limit=limit)
    return jsonify({"success": True, "results": results})


@daftar_tamu_bp.route("/admin/umum-rekap")
@role_required("admin")
def admin_public_summary() -> Response:
    status = (request.args.get("status") or "").strip().lower()
    search_query = (request.args.get("q") or "").strip()
    date_from = _parse_iso_date(request.args.get("date_from"))
    date_to = _parse_iso_date(request.args.get("date_to"))
    if date_from and date_to and date_from > date_to:
        date_from, date_to = date_to, date_from

    per_page = _to_int(request.args.get("per_page"), 10)
    per_page = max(5, min(per_page, 200))
    page = _to_int(request.args.get("page"), 1)
    page = max(1, page)

    rows, total_rows = list_admin_public_school_summary(
        status=status,
        search_query=search_query,
        date_from=date_from,
        date_to=date_to,
        page=page,
        per_page=per_page,
    )
    total_pages = max(1, math.ceil(total_rows / per_page)) if total_rows else 1
    if page > total_pages:
        page = total_pages
        rows, total_rows = list_admin_public_school_summary(
            status=status,
            search_query=search_query,
            date_from=date_from,
            date_to=date_to,
            page=page,
            per_page=per_page,
        )

    return render_template(
        "daftar_tamu/admin_public_summary.html",
        rows=rows,
        status=status,
        search_query=search_query,
        page=page,
        per_page=per_page,
        total_rows=total_rows,
        total_pages=total_pages,
        date_from_str=date_from.isoformat() if date_from else "",
        date_to_str=date_to.isoformat() if date_to else "",
    )


@daftar_tamu_bp.route("/admin/umum-transactions")
@role_required("admin")
def admin_public_transactions() -> Response:
    status = (request.args.get("status") or "").strip().lower()
    search_query = (request.args.get("q") or "").strip()
    date_from = _parse_iso_date(request.args.get("date_from"))
    date_to = _parse_iso_date(request.args.get("date_to"))
    per_page = _to_int(request.args.get("per_page"), 10)
    per_page = max(5, min(per_page, 100))
    page = _to_int(request.args.get("page"), 1)
    page = max(1, page)

    rows, total_rows = list_admin_public_transactions(
        status=status,
        search_query=search_query,
        date_from=date_from,
        date_to=date_to,
        page=page,
        per_page=per_page,
    )
    total_pages = max(1, math.ceil(total_rows / per_page)) if total_rows else 1
    if page > total_pages:
        page = total_pages
        rows, total_rows = list_admin_public_transactions(
            status=status,
            search_query=search_query,
            date_from=date_from,
            date_to=date_to,
            page=page,
            per_page=per_page,
        )

    return render_template(
        "daftar_tamu/admin_public_transactions.html",
        rows=rows,
        status=status,
        search_query=search_query,
        page=page,
        per_page=per_page,
        total_rows=total_rows,
        total_pages=total_pages,
        date_from_str=date_from.isoformat() if date_from else "",
        date_to_str=date_to.isoformat() if date_to else "",
    )


@daftar_tamu_bp.route("/admin/tujuan-kunjungan", methods=["GET", "POST"])
@role_required("admin")
def admin_purpose_keywords() -> Response:
    user = current_user()
    if request.method == "POST":
        action = (request.form.get("action") or "").strip().lower()
        if action == "add":
            keyword = (request.form.get("keyword") or "").strip()
            try:
                upsert_purpose_keyword(keyword=keyword, created_by=user.get("id"))
            except ValueError as exc:
                flash(str(exc), "danger")
            else:
                flash("Kata kunci tujuan disimpan.", "success")
        elif action == "toggle":
            keyword_id = _to_int(request.form.get("keyword_id"), 0)
            active_raw = (request.form.get("active") or "true").strip().lower()
            active = active_raw in ("1", "true", "yes", "on")
            if keyword_id <= 0:
                flash("Data kata kunci tidak valid.", "danger")
            else:
                ok = set_purpose_keyword_active(keyword_id=keyword_id, active=active)
                if ok:
                    flash("Status kata kunci diperbarui.", "success")
                else:
                    flash("Kata kunci tidak ditemukan.", "danger")
        else:
            flash("Aksi tidak dikenali.", "warning")
        return redirect(url_for("daftar_tamu.admin_purpose_keywords"))

    rows = list_purpose_keyword_rows(limit=400)
    return render_template(
        "daftar_tamu/admin_purpose_keywords.html",
        rows=rows,
    )


@daftar_tamu_bp.route("/admin/kontak-prioritas", methods=["GET", "POST"])
@role_required("admin")
def admin_contact_priority() -> Response:
    if request.method == "POST":
        keyword_id = _to_int(request.form.get("keyword_id"), 0)
        sort_order = _to_int(request.form.get("sort_order"), 0)
        active = (request.form.get("active") or "").strip().lower() == "true"
        if keyword_id <= 0 or sort_order <= 0:
            flash("Data prioritas tidak valid.", "danger")
        else:
            ok = update_contact_priority(keyword_id=keyword_id, sort_order=sort_order, active=active)
            if ok:
                flash("Prioritas kontak diperbarui.", "success")
            else:
                flash("Prioritas kontak tidak ditemukan.", "danger")
        return redirect(url_for("daftar_tamu.admin_contact_priority"))

    rows = list_contact_priority_rows(limit=50)
    return render_template(
        "daftar_tamu/admin_contact_priority.html",
        rows=rows,
    )


@daftar_tamu_bp.route("/admin/umum-transactions/<int:transaction_id>/approve", methods=["POST"])
@role_required("admin")
def admin_public_transaction_approve(transaction_id: int) -> Response:
    user = current_user()
    reviewer_notes = (request.form.get("reviewer_notes") or "").strip()
    try:
        ok = update_public_transaction_status(
            transaction_id=transaction_id,
            status="approved",
            reviewer_id=user["id"],
            reviewer_notes=reviewer_notes or None,
        )
    except ValueError:
        ok = False
    if not ok:
        flash("Transaksi tidak ditemukan.", "danger")
        return redirect(request.referrer or url_for("daftar_tamu.admin_public_transactions"))
    try:
        record_admin_action(
            user_id=user.get("id"),
            feature_key="daftar_tamu",
            action="PUBLIC_VERIFY_APPROVE",
            target_type="PUBLIC_GUESTBOOK_TRANSACTION",
            target_id=transaction_id,
            target_name=f"Transaksi Umum #{transaction_id}",
            metadata={"status": "approved", "reviewer_notes": reviewer_notes or None},
        )
    except Exception:
        current_app.logger.exception("Gagal mencatat action admin transaksi tamu umum.")
    flash("Transaksi tamu umum disetujui.", "success")
    return redirect(request.referrer or url_for("daftar_tamu.admin_public_transactions"))


@daftar_tamu_bp.route("/admin/umum-transactions/<int:transaction_id>/reject", methods=["POST"])
@role_required("admin")
def admin_public_transaction_reject(transaction_id: int) -> Response:
    user = current_user()
    reviewer_notes = (request.form.get("reviewer_notes") or "").strip()
    try:
        ok = update_public_transaction_status(
            transaction_id=transaction_id,
            status="rejected",
            reviewer_id=user["id"],
            reviewer_notes=reviewer_notes or None,
        )
    except ValueError:
        ok = False
    if not ok:
        flash("Transaksi tidak ditemukan.", "danger")
        return redirect(request.referrer or url_for("daftar_tamu.admin_public_transactions"))
    try:
        record_admin_action(
            user_id=user.get("id"),
            feature_key="daftar_tamu",
            action="PUBLIC_VERIFY_REJECT",
            target_type="PUBLIC_GUESTBOOK_TRANSACTION",
            target_id=transaction_id,
            target_name=f"Transaksi Umum #{transaction_id}",
            metadata={"status": "rejected", "reviewer_notes": reviewer_notes or None},
        )
    except Exception:
        current_app.logger.exception("Gagal mencatat action admin transaksi tamu umum.")
    flash("Transaksi tamu umum ditolak.", "success")
    return redirect(request.referrer or url_for("daftar_tamu.admin_public_transactions"))


@daftar_tamu_bp.route("/admin/umum-transactions/<int:transaction_id>/reopen", methods=["POST"])
@role_required("admin")
def admin_public_transaction_reopen(transaction_id: int) -> Response:
    user = current_user()
    try:
        ok = update_public_transaction_status(
            transaction_id=transaction_id,
            status="pending",
            reviewer_id=user["id"],
            reviewer_notes=None,
        )
    except ValueError:
        ok = False
    if not ok:
        flash("Transaksi tidak ditemukan.", "danger")
        return redirect(request.referrer or url_for("daftar_tamu.admin_public_transactions"))
    try:
        record_admin_action(
            user_id=user.get("id"),
            feature_key="daftar_tamu",
            action="PUBLIC_VERIFY_REOPEN",
            target_type="PUBLIC_GUESTBOOK_TRANSACTION",
            target_id=transaction_id,
            target_name=f"Transaksi Umum #{transaction_id}",
            metadata={"status": "pending"},
        )
    except Exception:
        current_app.logger.exception("Gagal mencatat action reopen transaksi tamu umum.")
    flash("Transaksi dibuka kembali untuk diverifikasi.", "success")
    return redirect(request.referrer or url_for("daftar_tamu.admin_public_transactions"))


@daftar_tamu_bp.route("/admin/umum/<int:guest_id>/verify", methods=["POST"])
@role_required("admin")
def admin_verify_general_guest(guest_id: int) -> Response:
    user = current_user()
    is_verified_raw = (request.form.get("is_verified") or "true").strip().lower()
    is_verified = is_verified_raw in ("1", "true", "yes", "on")
    with get_cursor(commit=True) as cur:
        cur.execute(
            """
            UPDATE daftar_tamu_general_guests
            SET is_verified = %s,
                verified_by = %s,
                verified_at = CASE WHEN %s THEN NOW() ELSE NULL END,
                updated_at = NOW()
            WHERE id = %s
              AND is_deleted = FALSE
            """,
            (is_verified, user.get("id"), is_verified, guest_id),
        )
        if cur.rowcount == 0:
            return jsonify({"success": False, "message": "Tamu umum tidak ditemukan atau sudah dihapus."}), 404
    try:
        record_admin_action(
            user_id=user.get("id"),
            feature_key="daftar_tamu",
            action="VERIFY_GENERAL_GUEST",
            target_type="GENERAL_GUEST",
            target_id=guest_id,
            target_name=f"Tamu Umum #{guest_id}",
            metadata={"is_verified": is_verified},
        )
    except Exception:
        current_app.logger.exception("Gagal mencatat verifikasi tamu umum.")
    return jsonify({"success": True, "is_verified": is_verified})


@daftar_tamu_bp.route("/admin/umum/<int:guest_id>/delete", methods=["POST"])
@role_required("admin")
def admin_delete_general_guest(guest_id: int) -> Response:
    user = current_user()
    is_deleted_raw = (request.form.get("is_deleted") or "true").strip().lower()
    is_deleted = is_deleted_raw in ("1", "true", "yes", "on")
    with get_cursor(commit=True) as cur:
        cur.execute(
            """
            UPDATE daftar_tamu_general_guests
            SET is_deleted = %s,
                deleted_by = CASE WHEN %s THEN %s ELSE NULL END,
                deleted_at = CASE WHEN %s THEN NOW() ELSE NULL END,
                updated_at = NOW()
            WHERE id = %s
            """,
            (is_deleted, is_deleted, user.get("id"), is_deleted, guest_id),
        )
        if cur.rowcount == 0:
            return jsonify({"success": False, "message": "Tamu umum tidak ditemukan."}), 404
    try:
        record_admin_action(
            user_id=user.get("id"),
            feature_key="daftar_tamu",
            action="DELETE_GENERAL_GUEST" if is_deleted else "RESTORE_GENERAL_GUEST",
            target_type="GENERAL_GUEST",
            target_id=guest_id,
            target_name=f"Tamu Umum #{guest_id}",
            metadata={"is_deleted": is_deleted},
        )
    except Exception:
        current_app.logger.exception("Gagal mencatat hapus tamu umum.")
    return jsonify({"success": True, "is_deleted": is_deleted})


# ===============================
# Sekolah Guestbook
# ===============================

@daftar_tamu_bp.route("/sekolah")
@role_required("sekolah")
def sekolah_guestbook() -> Response:
    user = current_user()
    school = _fetch_school_for_user(user["id"])
    if not school:
        return render_template(
            "daftar_tamu/sekolah_dashboard.html",
            school=None,
            error_message="Akun sekolah belum terhubung dengan data sekolah. Hubungi admin.",
        )

    return render_template(
        "daftar_tamu/sekolah_dashboard.html",
        school=school,
        user_school=school,
        area_contacts=_build_area_contacts(school),
        purpose_keywords=list_purpose_keywords_by_usage(active_only=True, limit=50),
        error_message=None,
    )


@daftar_tamu_bp.route("/sekolah/umum-web")
@role_required("sekolah")
def sekolah_public_web() -> Response:
    user = current_user()
    school = _fetch_school_for_user(user["id"])
    if not school:
        return render_template(
            "daftar_tamu/sekolah_umum_web.html",
            school=None,
            error_message="Akun sekolah belum terhubung dengan data sekolah. Hubungi admin.",
        )

    public_status = (request.args.get("public_status") or "").strip().lower()
    public_per_page = _to_int(request.args.get("public_per_page"), 5)
    public_per_page = max(3, min(public_per_page, 100))
    public_page = _to_int(request.args.get("public_page"), 1)
    public_page = max(1, public_page)

    public_rows, public_total_rows = list_school_public_transactions(
        school_id=school["id"],
        status=public_status,
        page=public_page,
        per_page=public_per_page,
    )
    public_total_pages = max(1, math.ceil(public_total_rows / public_per_page)) if public_total_rows else 1
    if public_page > public_total_pages:
        public_page = public_total_pages
        public_rows, public_total_rows = list_school_public_transactions(
            school_id=school["id"],
            status=public_status,
            page=public_page,
            per_page=public_per_page,
        )

    guestbook_public_url = f"{_web_aska_base_url()}/buku-tamu/{school.get('npsn')}"
    qr_payload = _get_guestbook_qr_payload(school["id"])
    qr_ready = bool(
        qr_payload
        and qr_payload.get("png_base64")
        and qr_payload.get("url") == guestbook_public_url
    )

    return render_template(
        "daftar_tamu/sekolah_umum_web.html",
        school=school,
        user_school=school,
        guestbook_public_url=guestbook_public_url,
        guestbook_qr_ready=qr_ready,
        public_rows=public_rows,
        public_status=public_status,
        public_page=public_page,
        public_per_page=public_per_page,
        public_total_rows=public_total_rows,
        public_total_pages=public_total_pages,
        error_message=None,
    )


@daftar_tamu_bp.route("/sekolah/umum-web/export")
@role_required("sekolah")
def sekolah_public_web_export() -> Response:
    """Export school public guestbook submissions (verified only)."""
    user = current_user()
    school = _fetch_school_for_user(user["id"])
    if not school:
        return Response("Akun sekolah belum terhubung dengan data sekolah.", status=400)

    # Verified rows in this workflow are transactions that have been approved.
    status_filter = "approved"
    per_page = 100
    rows, total_rows = list_school_public_transactions(
        school_id=school["id"],
        status=status_filter,
        page=1,
        per_page=per_page,
    )
    total_pages = max(1, math.ceil(total_rows / per_page)) if total_rows else 1
    if total_pages > 1:
        for page in range(2, total_pages + 1):
            page_rows, _ = list_school_public_transactions(
                school_id=school["id"],
                status=status_filter,
                page=page,
                per_page=per_page,
            )
            rows.extend(page_rows)

    headers = [
        "Tanggal Kunjungan",
        "Jam Kunjungan",
        "Nama Tamu",
        "Jumlah Tamu",
        "Tujuan",
        "Status Verifikasi",
        "Catatan Verifikator",
    ]
    data_rows: list[list[object]] = []
    for row in rows:
        visit_at = row.get("visit_at")
        visit_date = ""
        visit_time = ""
        if isinstance(visit_at, datetime):
            visit_date = visit_at.strftime("%d/%m/%Y")
            visit_time = visit_at.strftime("%H:%M")
        elif isinstance(visit_at, date):
            visit_date = visit_at.strftime("%d/%m/%Y")
        data_rows.append(
            [
                visit_date,
                visit_time,
                row.get("guest_names") or row.get("guest_display") or "",
                row.get("guest_count") or 0,
                row.get("purpose") or "",
                "Disetujui",
                row.get("reviewer_notes") or "",
            ]
        )

    file_format = (request.args.get("format") or "excel").strip().lower()
    school_npsn = (school.get("npsn") or "sekolah").strip()
    if file_format in {"excel", "xlsx"}:
        filename = f"daftar_tamu_umum_terverifikasi_{school_npsn}_{_today_jakarta().isoformat()}.xlsx"
        return _build_xlsx_response(headers, data_rows, filename)

    filename = f"daftar_tamu_umum_terverifikasi_{school_npsn}_{_today_jakarta().isoformat()}.csv"
    return _build_csv_response(headers, data_rows, filename)


@daftar_tamu_bp.route("/sekolah/riwayat")
@role_required("sekolah")
def sekolah_riwayat() -> Response:
    user = current_user()
    school = _fetch_school_for_user(user["id"])
    if not school:
        return render_template(
            "daftar_tamu/sekolah_riwayat.html",
            school=None,
            error_message="Akun sekolah belum terhubung dengan data sekolah. Hubungi admin.",
        )

    status = (request.args.get("status") or "").strip().lower()
    guest_scope = _parse_guest_scope(request.args.get("guest_scope"), default="all")
    per_page = _to_int(request.args.get("per_page"), 10)
    per_page = max(5, min(per_page, 100))
    page = _to_int(request.args.get("page"), 1)
    page = max(1, page)

    rows, total_rows = list_school_transactions(
        school_id=school["id"],
        status=status,
        guest_scope=guest_scope,
        page=page,
        per_page=per_page,
    )
    total_pages = max(1, math.ceil(total_rows / per_page)) if total_rows else 1
    if page > total_pages:
        page = total_pages
        rows, total_rows = list_school_transactions(
            school_id=school["id"],
            status=status,
            guest_scope=guest_scope,
            page=page,
            per_page=per_page,
        )

    return render_template(
        "daftar_tamu/sekolah_riwayat.html",
        school=school,
        user_school=school,
        rows=rows,
        status=status,
        guest_scope=guest_scope,
        page=page,
        per_page=per_page,
        total_rows=total_rows,
        total_pages=total_pages,
        error_message=None,
    )


@daftar_tamu_bp.route("/sekolah/riwayat/harian")
@role_required("sekolah")
def sekolah_riwayat_harian() -> Response:
    user = current_user()
    school = _fetch_school_for_user(user["id"])
    if not school:
        return render_template(
            "daftar_tamu/sekolah_riwayat_harian.html",
            school=None,
            error_message="Akun sekolah belum terhubung dengan data sekolah. Hubungi admin.",
        )

    date_from = _parse_iso_date(request.args.get("date_from"))
    date_to = _parse_iso_date(request.args.get("date_to"))
    if date_from and date_to and date_from > date_to:
        date_from, date_to = date_to, date_from

    search_query = (request.args.get("q") or "").strip()
    guest_scope = _parse_guest_scope(request.args.get("guest_scope"), default="sudin")
    per_page = _to_int(request.args.get("per_page"), 10)
    per_page = max(5, min(per_page, 100))
    page = _to_int(request.args.get("page"), 1)
    page = max(1, page)

    rows, total_rows = fetch_school_visit_days(
        school_id=school["id"],
        page=page,
        per_page=per_page,
        search_query=search_query,
        date_from=date_from,
        date_to=date_to,
        guest_scope=guest_scope,
    )
    total_pages = max(1, math.ceil(total_rows / per_page)) if total_rows else 1
    if page > total_pages:
        page = total_pages
        rows, total_rows = fetch_school_visit_days(
            school_id=school["id"],
            page=page,
            per_page=per_page,
            search_query=search_query,
            date_from=date_from,
            date_to=date_to,
            guest_scope=guest_scope,
        )

    return render_template(
        "daftar_tamu/sekolah_riwayat_harian.html",
        school=school,
        user_school=school,
        rows=rows,
        search_query=search_query,
        date_from_str=date_from.isoformat() if date_from else "",
        date_to_str=date_to.isoformat() if date_to else "",
        guest_scope=guest_scope,
        page=page,
        per_page=per_page,
        total_rows=total_rows,
        total_pages=total_pages,
        error_message=None,
    )


@daftar_tamu_bp.route("/sekolah/riwayat/harian/<visit_date>/guests")
@role_required("sekolah")
def sekolah_riwayat_harian_guests(visit_date: str) -> Response:
    user = current_user()
    school = _fetch_school_for_user(user["id"])
    if not school:
        return jsonify({"success": False, "message": "Akun sekolah belum terhubung."}), 400

    parsed_visit_date = _parse_iso_date(visit_date)
    if not parsed_visit_date:
        return jsonify({"success": False, "message": "Tanggal kunjungan tidak valid."}), 400

    page = _to_int(request.args.get("page"), 1)
    page = max(1, page)
    per_page = _to_int(request.args.get("per_page"), 100)
    per_page = max(5, min(per_page, 100))
    search_query = (request.args.get("q") or "").strip()
    guest_scope = _parse_guest_scope(request.args.get("guest_scope"), default="sudin")

    rows, total_rows = fetch_school_visit_day_guests(
        school_id=school["id"],
        visit_date=parsed_visit_date,
        page=page,
        per_page=per_page,
        search_query=search_query,
        guest_scope=guest_scope,
    )
    total_pages = max(1, math.ceil(total_rows / per_page)) if total_rows else 1
    if page > total_pages:
        page = total_pages
        rows, total_rows = fetch_school_visit_day_guests(
            school_id=school["id"],
            visit_date=parsed_visit_date,
            page=page,
            per_page=per_page,
            search_query=search_query,
            guest_scope=guest_scope,
        )

    for row in rows:
        visit_at = row.get("visit_at")
        row["visit_at"] = visit_at.isoformat() if visit_at else None

    return jsonify(
        {
            "success": True,
            "rows": rows,
            "total_rows": total_rows,
            "total_pages": total_pages,
            "page": page,
            "per_page": per_page,
            "visit_date": parsed_visit_date.isoformat(),
        }
    )


@daftar_tamu_bp.route("/sekolah/riwayat/harian/export")
@role_required("sekolah")
def sekolah_riwayat_harian_export() -> Response:
    """Export school daily kedinasan history with active filters."""
    user = current_user()
    school = _fetch_school_for_user(user["id"])
    if not school:
        return Response("Akun sekolah belum terhubung dengan data sekolah.", status=400)

    date_from = _parse_iso_date(request.args.get("date_from"))
    date_to = _parse_iso_date(request.args.get("date_to"))
    if date_from and date_to and date_from > date_to:
        date_from, date_to = date_to, date_from

    search_query = (request.args.get("q") or "").strip()
    guest_scope = _parse_guest_scope(request.args.get("guest_scope"), default="sudin")
    export_format = (request.args.get("format") or "csv").strip().lower()

    rows: list[dict] = []
    export_page = 1
    export_per_page = 100
    while True:
        page_rows, total_rows = fetch_school_visit_days(
            school_id=school["id"],
            page=export_page,
            per_page=export_per_page,
            search_query=search_query,
            date_from=date_from,
            date_to=date_to,
            guest_scope=guest_scope,
        )
        rows.extend(page_rows)
        if len(rows) >= total_rows or not page_rows:
            break
        export_page += 1

    headers = ["Hari", "Tanggal", "Jumlah Kunjungan", "Jumlah Orang", "Nama Pengunjung", "Instansi", "Tujuan"]
    weekday_labels = {
        0: "Senin",
        1: "Selasa",
        2: "Rabu",
        3: "Kamis",
        4: "Jumat",
        5: "Sabtu",
        6: "Minggu",
    }
    data_rows: list[list[object]] = []
    summary_fill_ranges: list[tuple[int, int, int, str]] = []
    for day_row in rows:
        visit_date = day_row.get("visit_date")
        weekday = ""
        date_label = ""
        if visit_date:
            weekday = weekday_labels.get(visit_date.weekday(), "")
            date_label = _format_date_dmy(visit_date)

        guest_rows: list[dict] = []
        guest_page = 1
        guest_per_page = 100
        while visit_date:
            page_guest_rows, guest_total_rows = fetch_school_visit_day_guests(
                school_id=school["id"],
                visit_date=visit_date,
                page=guest_page,
                per_page=guest_per_page,
                search_query=search_query,
                guest_scope=guest_scope,
            )
            guest_rows.extend(page_guest_rows)
            if len(guest_rows) >= guest_total_rows or not page_guest_rows:
                break
            guest_page += 1

        first_excel_row_for_date = len(data_rows) + 2
        summary_fill_ranges.append((first_excel_row_for_date, 1, 4, "D9EAF7"))
        if not guest_rows:
            data_rows.append(
                [
                    weekday,
                    date_label,
                    day_row.get("visit_count") or 0,
                    day_row.get("people_count") or 0,
                    "",
                    "",
                    "",
                ]
            )
            continue

        for guest_index, guest_row in enumerate(guest_rows):
            show_summary = guest_index == 0
            data_rows.append(
                [
                    weekday if show_summary else "",
                    date_label if show_summary else "",
                    (day_row.get("visit_count") or 0) if show_summary else "",
                    (day_row.get("people_count") or 0) if show_summary else "",
                    guest_row.get("guest_name") or "",
                    guest_row.get("instansi") or "",
                    guest_row.get("purpose") or "",
                ]
            )

    school_npsn = school.get("npsn") or school.get("id") or "sekolah"
    today_label = _today_jakarta().isoformat()
    if export_format == "excel":
        filename = f"data_harian_tamu_kedinasan_{school_npsn}_{today_label}.xlsx"
        return _build_xlsx_response(headers, data_rows, filename, fill_ranges=summary_fill_ranges)

    filename = f"data_harian_tamu_kedinasan_{school_npsn}_{today_label}.csv"
    return _build_csv_response(headers, data_rows, filename)


@daftar_tamu_bp.route("/sekolah/riwayat/export")
@role_required("sekolah")
def sekolah_riwayat_export() -> Response:
    """Export school kedinasan history with active filters (status + guest scope)."""
    user = current_user()
    school = _fetch_school_for_user(user["id"])
    if not school:
        return Response("Akun sekolah belum terhubung dengan data sekolah.", status=400)

    status = (request.args.get("status") or "").strip().lower()
    guest_scope = _parse_guest_scope(request.args.get("guest_scope"), default="all")

    per_page = 100
    rows, total_rows = list_school_transactions(
        school_id=school["id"],
        status=status,
        guest_scope=guest_scope,
        page=1,
        per_page=per_page,
    )
    total_pages = max(1, math.ceil(total_rows / per_page)) if total_rows else 1
    if total_pages > 1:
        for page in range(2, total_pages + 1):
            page_rows, _ = list_school_transactions(
                school_id=school["id"],
                status=status,
                guest_scope=guest_scope,
                page=page,
                per_page=per_page,
            )
            rows.extend(page_rows)

    headers = [
        "Tanggal Kunjungan",
        "Jam Kunjungan",
        "Tipe Tamu",
        "Nama Tamu",
        "Tujuan",
        "Status Verifikasi",
        "Catatan Verifikator",
        "Nama Verifikator",
        "Link Foto",
    ]
    data_rows: list[list[object]] = []
    for row in rows:
        visit_at = row.get("visit_at")
        visit_date = ""
        visit_time = ""
        if isinstance(visit_at, datetime):
            visit_date = visit_at.strftime("%d/%m/%Y")
            visit_time = visit_at.strftime("%H:%M")
        elif isinstance(visit_at, date):
            visit_date = visit_at.strftime("%d/%m/%Y")

        status_value = (row.get("status") or "").strip().lower()
        if status_value == "approved":
            status_label = "Disetujui"
        elif status_value == "rejected":
            status_label = "Ditolak"
        else:
            status_label = "Menunggu"

        photo_link = ""
        photo_path = (row.get("photo_path") or "").strip()
        if photo_path:
            photo_name = photo_path.split("uploads/portal/")[-1]
            photo_link = url_for("portal.uploaded_file", filename=photo_name, _external=True)

        guest_type = (row.get("guest_type") or "").strip().lower()
        guest_type_label = "Sudindik JU 2" if guest_type == "sudin" else "Instansi Pemerintah Lainnya"

        data_rows.append(
            [
                visit_date,
                visit_time,
                guest_type_label,
                row.get("guest_names") or row.get("guest_display") or "",
                row.get("purpose") or "",
                status_label,
                row.get("reviewer_notes") or "",
                row.get("reviewer_name") or "",
                photo_link,
            ]
        )

    file_format = (request.args.get("format") or "excel").strip().lower()
    school_npsn = (school.get("npsn") or "sekolah").strip()
    if file_format in {"excel", "xlsx"}:
        filename = f"riwayat_tamu_kedinasan_{school_npsn}_{_today_jakarta().isoformat()}.xlsx"
        return _build_xlsx_response(headers, data_rows, filename)

    filename = f"riwayat_tamu_kedinasan_{school_npsn}_{_today_jakarta().isoformat()}.csv"
    return _build_csv_response(headers, data_rows, filename)


def _extract_user_staff_note(metadata_value: object, user_id: int) -> tuple[str, str, str]:
    metadata = metadata_value
    if isinstance(metadata, str):
        try:
            metadata = json.loads(metadata)
        except json.JSONDecodeError:
            metadata = {}
    if not isinstance(metadata, dict):
        return "", "", ""

    staff_notes = metadata.get("staff_notes")
    if not isinstance(staff_notes, dict):
        return "", "", ""

    entry = staff_notes.get(str(user_id))
    if isinstance(entry, dict):
        note = (entry.get("note") or "").strip()
        updated_at = (entry.get("updated_at") or "").strip()
        level = _normalize_staff_note_level(entry.get("level"), default="tindak_lanjut") or "tindak_lanjut"
        return note, updated_at, level
    if isinstance(entry, str):
        return entry.strip(), "", "tindak_lanjut"
    return "", "", ""


def _build_user_guestbook_history_redirect(source) -> str:
    params = _parse_user_guestbook_history_args(source)

    return url_for(
        "daftar_tamu.user_guestbook_history",
        tab=params["tab"],
        q=params["search_query"],
        sort=params["sort"],
        date_from=params["date_from_str"],
        date_to=params["date_to_str"],
        guest_scope=params["guest_scope"],
        status=params["status"],
        page=params["page"],
        per_page=params["per_page"],
        home_limit=params["home_limit"],
    )


def _normalize_history_tab(value: Optional[str]) -> str:
    tab = (value or "").strip().lower() or "beranda"
    if tab not in _HISTORY_TAB_OPTIONS:
        return "beranda"
    return tab


def _normalize_history_status(value: Optional[str]) -> str:
    status = (value or "").strip().lower()
    if status in {"all", "semua"}:
        return ""
    if status not in _HISTORY_STATUS_OPTIONS:
        return ""
    return status


def _normalize_history_sort(value: Optional[str]) -> str:
    sort = (value or "").strip().lower() or "date_desc"
    if sort not in _HISTORY_SORT_OPTIONS:
        return "date_desc"
    return sort


def _format_staff_note_updated_label(value: str) -> str:
    raw = (value or "").strip()
    if not raw:
        return ""
    try:
        parsed = datetime.fromisoformat(raw.replace("Z", "+00:00"))
        return to_jakarta(parsed).strftime("%d %b %Y, %H:%M WIB")
    except ValueError:
        return raw


def _parse_user_guestbook_history_args(source) -> dict:
    home_limit_max = 80
    tab = _normalize_history_tab(source.get("tab"))
    status = _normalize_history_status(source.get("status"))
    date_from = _parse_iso_date(source.get("date_from"))
    date_to = _parse_iso_date(source.get("date_to"))
    if date_from and date_to and date_from > date_to:
        date_from, date_to = date_to, date_from

    search_query = (source.get("q") or "").strip()
    guest_scope = _parse_guest_scope(source.get("guest_scope"), default="all")

    page = _to_int(source.get("page"), 1)
    page = max(1, page)
    per_page = _to_int(source.get("per_page"), 10)
    per_page = max(5, min(per_page, 100))
    home_limit = _to_int(source.get("home_limit"), 4)
    home_limit = max(4, min(home_limit, home_limit_max))
    if home_limit % 4:
        home_limit = 4 * math.ceil(home_limit / 4)

    sort = _normalize_history_sort(source.get("sort"))
    home_page = max(1, _to_int(source.get("home_page"), 1))
    home_chunk_size = max(4, min(_to_int(source.get("home_chunk_size"), 4), 20))
    if home_chunk_size % 4:
        home_chunk_size = 4 * math.ceil(home_chunk_size / 4)

    return {
        "tab": tab,
        "status": status,
        "date_from": date_from,
        "date_to": date_to,
        "date_from_str": date_from.isoformat() if date_from else "",
        "date_to_str": date_to.isoformat() if date_to else "",
        "search_query": search_query,
        "guest_scope": guest_scope,
        "page": page,
        "per_page": per_page,
        "home_limit": home_limit,
        "sort": sort,
        "home_page": home_page,
        "home_chunk_size": home_chunk_size,
        "home_limit_max": home_limit_max,
    }


def _decorate_user_history_rows(rows: list[dict], user_id: int) -> None:
    for row in rows:
        status_value = (row.get("status") or "").strip().lower()
        row["photo_url"] = _build_photo_url(row.get("photo_path"))
        row["photo_thumb_url"] = _build_photo_thumb_url(row.get("photo_path"), width=420, quality=72)
        row["status_label"] = {
            "approved": "Terverifikasi",
            "rejected": "Ditolak",
        }.get(status_value, "Menunggu")
        row["status_icon"] = {
            "approved": "bi-check-circle-fill",
            "rejected": "bi-x-circle-fill",
        }.get(status_value, "bi-hourglass-split")
        row["status_tone"] = {
            "approved": "success",
            "rejected": "danger",
        }.get(status_value, "warning")

        visit_at = row.get("visit_at")
        if visit_at:
            local_visit = to_jakarta(visit_at)
            row["visit_at_iso"] = local_visit.isoformat(timespec="seconds")
            row["visit_at_label"] = local_visit.strftime("%d %b %Y, %H:%M WIB")
        else:
            row["visit_at_iso"] = ""
            row["visit_at_label"] = ""

        staff_note, staff_note_updated_at, staff_note_level = _extract_user_staff_note(row.get("metadata"), user_id)
        row["staff_note"] = staff_note
        row["staff_note_updated_at_raw"] = staff_note_updated_at
        row["staff_note_updated_at"] = _format_staff_note_updated_label(staff_note_updated_at)
        row["staff_note_level"] = staff_note_level
        row["staff_note_level_label"] = _STAFF_NOTE_LEVEL_LABEL_MAP.get(staff_note_level, "")
        row["staff_note_level_tone"] = _STAFF_NOTE_LEVEL_TONE_MAP.get(staff_note_level, "secondary")
        row["can_add_staff_note"] = status_value == "approved" and bool(row.get("photo_path"))
        row["signature"] = _build_user_history_signature(row)


def _build_user_history_signature(row: dict) -> str:
    reviewed_at = row.get("reviewed_at")
    if isinstance(reviewed_at, datetime):
        reviewed_val = reviewed_at.isoformat(timespec="seconds")
    else:
        reviewed_val = str(reviewed_at or "")
    return "|".join(
        [
            str(row.get("transaction_id") or ""),
            str((row.get("status") or "").strip().lower()),
            reviewed_val,
            str((row.get("staff_note") or "").strip()),
            str((row.get("staff_note_level") or "").strip()),
            str((row.get("staff_note_updated_at_raw") or "").strip()),
            str((row.get("photo_path") or "").strip()),
        ]
    )


def _serialize_user_history_row(row: dict) -> dict:
    return {
        "transaction_id": int(row.get("transaction_id") or 0),
        "status": (row.get("status") or "").strip().lower(),
        "status_label": row.get("status_label") or "",
        "status_icon": row.get("status_icon") or "",
        "status_tone": row.get("status_tone") or "warning",
        "school_name": row.get("school_name") or "",
        "visit_at_label": row.get("visit_at_label") or "",
        "visit_at_iso": row.get("visit_at_iso") or "",
        "purpose": row.get("purpose") or "",
        "guest_display": row.get("guest_display") or "",
        "guest_count": int(row.get("guest_count") or 0),
        "photo_url": row.get("photo_url") or "",
        "photo_thumb_url": row.get("photo_thumb_url") or "",
        "reviewer_notes": row.get("reviewer_notes") or "",
        "staff_note": row.get("staff_note") or "",
        "staff_note_level": row.get("staff_note_level") or "",
        "staff_note_level_label": row.get("staff_note_level_label") or "",
        "staff_note_level_tone": row.get("staff_note_level_tone") or "secondary",
        "staff_note_updated_at": row.get("staff_note_updated_at") or "",
        "can_add_staff_note": bool(row.get("can_add_staff_note")),
        "signature": _build_user_history_signature(row),
    }


def _format_notification_time_label(value: object) -> str:
    if not isinstance(value, datetime):
        return ""
    local_dt = to_jakarta(value) or value
    now_dt = current_jakarta_time()
    delta_seconds = int((now_dt - local_dt).total_seconds())
    if delta_seconds < 60:
        return "Baru saja"
    if delta_seconds < 3600:
        return f"{max(1, delta_seconds // 60)} menit lalu"
    if delta_seconds < 86400:
        return f"{max(1, delta_seconds // 3600)} jam lalu"
    return local_dt.strftime("%d %b %Y, %H:%M WIB")


def _serialize_user_guestbook_notification(row: dict, fallback_link: str) -> dict:
    notification_id = int(row.get("id") or 0)
    category = (row.get("category") or "").strip()
    status_value = (row.get("status") or "").strip().lower()
    metadata = row.get("metadata")
    if not isinstance(metadata, dict):
        metadata = {}
    status_key = str(metadata.get("status") or "").strip().lower()

    icon = "bi-bell-fill"
    if category == "daftar_tamu_status":
        icon = "bi-journal-check"
    elif category == "panbers_reopen_status":
        icon = "bi-arrow-counterclockwise"
    elif category == "panbers_assignment_status":
        icon = "bi-diagram-3-fill"
    elif category == "panbers_team_member_status":
        icon = "bi-people-fill"
    elif category == "panbers_follow_up_status":
        icon = "bi-tools"
    elif category == "hospitality_status":
        icon = "bi-house-heart"

    tone = "secondary"
    if status_key == "approved":
        tone = "success"
    elif status_key == "rejected":
        tone = "danger"
    elif status_key == "pending":
        tone = "warning"
    elif status_key == "selesai":
        tone = "success"
    elif status_key in {"baru", "diproses", "diajukan"}:
        tone = "warning"

    created_at = row.get("created_at")
    created_at_iso = created_at.isoformat(timespec="seconds") if isinstance(created_at, datetime) else ""

    fallback_title = "Notifikasi buku tamu"
    if category == "panbers_reopen_status":
        fallback_title = "Notifikasi reopen PANBERSS"
    elif category == "panbers_assignment_status":
        fallback_title = "Notifikasi penugasan PANBERSS"
    elif category == "panbers_team_member_status":
        fallback_title = "Notifikasi tim PANBERSS"
    elif category == "panbers_follow_up_status":
        fallback_title = "Notifikasi tindak lanjut PANBERSS"
    elif category == "hospitality_status":
        fallback_title = "Notifikasi Hospitality"

    return {
        "id": notification_id,
        "category": category,
        "title": (row.get("title") or "").strip() or fallback_title,
        "message": (row.get("message") or "").strip(),
        "status": status_value or "unread",
        "is_unread": status_value == "unread",
        "link": (row.get("link") or "").strip() or fallback_link,
        "icon": icon,
        "tone": tone,
        "created_at": created_at_iso,
        "created_label": _format_notification_time_label(created_at),
        "reference_id": int(row.get("reference_id") or 0),
    }


def _build_user_guestbook_history_context(user: dict, source) -> dict:
    params = _parse_user_guestbook_history_args(source)
    user_id = int(user.get("id"))

    home_rows: list[dict] = []
    home_total_rows = 0
    home_has_more = False

    detail_rows: list[dict] = []
    detail_total_rows = 0
    detail_total_pages = 1

    if params["tab"] == "detail":
        detail_rows, detail_total_rows = fetch_user_guestbook_history(
            user_id=user_id,
            page=params["page"],
            per_page=params["per_page"],
            sort_key=params["sort"],
            status=params["status"],
            search_query=params["search_query"],
            date_from=params["date_from"],
            date_to=params["date_to"],
            guest_scope=params["guest_scope"],
        )
        detail_total_pages = max(1, math.ceil(detail_total_rows / params["per_page"])) if detail_total_rows else 1
        if params["page"] > detail_total_pages:
            params["page"] = detail_total_pages
            detail_rows, detail_total_rows = fetch_user_guestbook_history(
                user_id=user_id,
                page=params["page"],
                per_page=params["per_page"],
                sort_key=params["sort"],
                status=params["status"],
                search_query=params["search_query"],
                date_from=params["date_from"],
                date_to=params["date_to"],
                guest_scope=params["guest_scope"],
            )
    else:
        home_rows, home_total_rows = fetch_user_guestbook_history(
            user_id=user_id,
            page=1,
            per_page=params["home_limit"],
            sort_key="date_desc",
            status=params["status"],
            search_query=params["search_query"],
            date_from=params["date_from"],
            date_to=params["date_to"],
            guest_scope=params["guest_scope"],
        )
        home_has_more = home_total_rows > len(home_rows) and params["home_limit"] < params["home_limit_max"]

    _decorate_user_history_rows(home_rows, user_id)
    _decorate_user_history_rows(detail_rows, user_id)

    feed_base_url = url_for(
        "daftar_tamu.user_guestbook_history_feed",
        tab="beranda",
        q=params["search_query"],
        sort="date_desc",
        date_from=params["date_from_str"],
        date_to=params["date_to_str"],
        guest_scope=params["guest_scope"],
        status=params["status"],
        home_chunk_size=4,
    )

    stream_url = url_for(
        "daftar_tamu.user_guestbook_history_stream",
        tab="beranda",
        q=params["search_query"],
        sort="date_desc",
        date_from=params["date_from_str"],
        date_to=params["date_to_str"],
        guest_scope=params["guest_scope"],
        status=params["status"],
    )

    return {
        "tab": params["tab"],
        "status": params["status"],
        "home_rows": home_rows,
        "home_total_rows": home_total_rows,
        "home_limit": params["home_limit"],
        "home_has_more": home_has_more,
        "detail_rows": detail_rows,
        "detail_total_rows": detail_total_rows,
        "detail_total_pages": detail_total_pages,
        "page": params["page"],
        "per_page": params["per_page"],
        "sort": params["sort"],
        "search_query": params["search_query"],
        "date_from_str": params["date_from_str"],
        "date_to_str": params["date_to_str"],
        "guest_scope": params["guest_scope"],
        "today_str": _today_jakarta().isoformat(),
        "user_profile": user,
        "history_feed_base_url": feed_base_url,
        "history_stream_url": stream_url,
    }


def _fetch_user_guestbook_home_chunk(
    *,
    user_id: int,
    page: int,
    chunk_size: int,
    status: str,
    search_query: str,
    date_from: Optional[date],
    date_to: Optional[date],
    guest_scope: str,
) -> tuple[list[dict], int]:
    rows, total_rows = fetch_user_guestbook_history(
        user_id=user_id,
        page=page,
        per_page=chunk_size,
        sort_key="date_desc",
        status=status,
        search_query=search_query,
        date_from=date_from,
        date_to=date_to,
        guest_scope=guest_scope,
    )
    _decorate_user_history_rows(rows, user_id)
    return rows, total_rows


@daftar_tamu_bp.route("/saya/notifikasi")
@role_required("staff", "coordinator", "sekolah")
def user_guestbook_notifications() -> Response:
    user = current_user()
    user_id = int(user.get("id"))
    limit = max(1, min(_to_int(request.args.get("limit"), 8), 30))
    if user.get("role") == "sekolah":
        fallback_link = url_for("daftar_tamu.sekolah_riwayat")
    else:
        fallback_link = url_for("daftar_tamu.user_guestbook_history", tab="detail")
    categories = list(USER_APP_NOTIFICATION_CATEGORIES)

    try:
        summary = fetch_user_notification_summary(user_id=user_id, categories=categories)
        rows = list_user_notifications(user_id=user_id, limit=limit, categories=categories)
    except Exception:
        current_app.logger.exception("Gagal mengambil notifikasi buku tamu pengguna aplikasi.")
        return jsonify(
            {
                "success": False,
                "items": [],
                "unread_count": 0,
                "total_count": 0,
                "generated_at": current_jakarta_time().isoformat(timespec="seconds"),
            }
        )

    return jsonify(
        {
            "success": True,
            "items": [_serialize_user_guestbook_notification(row, fallback_link) for row in rows],
            "unread_count": int(summary.get("unread_count") or 0),
            "total_count": int(summary.get("total_count") or 0),
            "generated_at": current_jakarta_time().isoformat(timespec="seconds"),
        }
    )


@daftar_tamu_bp.route("/saya/notifikasi/tandai-dibaca", methods=["POST"])
@role_required("staff", "coordinator", "sekolah")
def user_guestbook_notifications_mark_read() -> Response:
    user = current_user()
    user_id = int(user.get("id"))

    payload = request.get_json(silent=True) if request.is_json else {}
    if not isinstance(payload, dict):
        payload = {}

    mark_all_raw = payload.get("all", request.form.get("all"))
    mark_all = str(mark_all_raw or "").strip().lower() in {"1", "true", "yes", "on"}

    raw_ids = payload.get("ids")
    if not isinstance(raw_ids, list):
        raw_ids = request.form.getlist("ids") or request.form.getlist("notification_ids")
    if not raw_ids and request.form.get("id"):
        raw_ids = [request.form.get("id")]

    notification_ids: list[int] = []
    for raw_id in raw_ids or []:
        try:
            notification_ids.append(int(raw_id))
        except (TypeError, ValueError):
            continue

    categories = list(USER_APP_NOTIFICATION_CATEGORIES)
    try:
        updated_count = mark_user_notifications_read(
            user_id=user_id,
            notification_ids=notification_ids,
            mark_all=mark_all,
            categories=categories,
        )
        summary = fetch_user_notification_summary(user_id=user_id, categories=categories)
    except Exception:
        current_app.logger.exception("Gagal memperbarui notifikasi buku tamu pengguna aplikasi.")
        return jsonify({"success": False, "message": "Gagal memperbarui notifikasi."}), 500

    return jsonify(
        {
            "success": True,
            "updated_count": int(updated_count or 0),
            "unread_count": int(summary.get("unread_count") or 0),
            "total_count": int(summary.get("total_count") or 0),
        }
    )


@daftar_tamu_bp.route("/saya/riwayat")
@role_required("staff", "coordinator")
def user_guestbook_history() -> Response:
    """Show current user guestbook history with mobile-first home + detail tabs."""
    user = current_user()
    photo_redirect = _require_profile_photo_redirect(user)
    if photo_redirect:
        return photo_redirect
    context = _build_user_guestbook_history_context(user, request.args)
    return render_template("daftar_tamu/user_guestbook_history.html", **context)


@daftar_tamu_bp.route("/saya/riwayat/feed")
@role_required("staff", "coordinator")
def user_guestbook_history_feed() -> Response:
    """Return JSON feed chunk for super-app timeline."""
    user = current_user()
    photo_redirect = _require_profile_photo_redirect(user)
    if photo_redirect:
        return photo_redirect
    user_id = int(user.get("id"))
    params = _parse_user_guestbook_history_args(request.args)

    home_rows, home_total_rows = _fetch_user_guestbook_home_chunk(
        user_id=user_id,
        page=params["home_page"],
        chunk_size=params["home_chunk_size"],
        status=params["status"],
        search_query=params["search_query"],
        date_from=params["date_from"],
        date_to=params["date_to"],
        guest_scope=params["guest_scope"],
    )
    total_pages = max(1, math.ceil(home_total_rows / params["home_chunk_size"])) if home_total_rows else 1
    has_more = params["home_page"] < total_pages

    items_html = render_template(
        "daftar_tamu/partials/_user_guestbook_home_timeline_items.html",
        rows=home_rows,
        search_query=params["search_query"],
        date_from_str=params["date_from_str"],
        date_to_str=params["date_to_str"],
        guest_scope=params["guest_scope"],
        status=params["status"],
        per_page=params["per_page"],
        home_limit=params["home_limit"],
    )

    return jsonify(
        {
            "success": True,
            "home_page": params["home_page"],
            "home_chunk_size": params["home_chunk_size"],
            "home_total_rows": home_total_rows,
            "home_total_pages": total_pages,
            "has_more": has_more,
            "next_page": params["home_page"] + 1 if has_more else None,
            "items_count": len(home_rows),
            "items": [_serialize_user_history_row(row) for row in home_rows],
            "items_html": items_html,
            "generated_at": current_jakarta_time().isoformat(timespec="seconds"),
        }
    )


@daftar_tamu_bp.route("/saya/riwayat/stream")
@role_required("staff", "coordinator")
def user_guestbook_history_stream() -> Response:
    """Server-Sent Events for realtime row updates in super-app history."""
    user = current_user()
    photo_redirect = _require_profile_photo_redirect(user)
    if photo_redirect:
        return photo_redirect
    user_id = int(user.get("id"))
    params = _parse_user_guestbook_history_args(request.args)
    poll_limit = 40

    def _snapshot() -> tuple[list[dict], dict[int, str]]:
        rows, _ = _fetch_user_guestbook_home_chunk(
            user_id=user_id,
            page=1,
            chunk_size=poll_limit,
            status=params["status"],
            search_query=params["search_query"],
            date_from=params["date_from"],
            date_to=params["date_to"],
            guest_scope=params["guest_scope"],
        )
        signatures = {int(row.get("transaction_id") or 0): _build_user_history_signature(row) for row in rows if row.get("transaction_id")}
        return rows, signatures

    def _event_payload(event_name: str, payload: dict) -> str:
        return f"event: {event_name}\ndata: {json.dumps(payload, ensure_ascii=False)}\n\n"

    @stream_with_context
    def _generate():
        _, last_signatures = _snapshot()
        started_at = time.monotonic()
        heartbeat_counter = 0
        yield "retry: 4500\n\n"
        while time.monotonic() - started_at < 55:
            time.sleep(4)
            rows, next_signatures = _snapshot()
            changed_rows = []
            for row in rows:
                tx_id = int(row.get("transaction_id") or 0)
                if not tx_id:
                    continue
                if last_signatures.get(tx_id) != next_signatures.get(tx_id):
                    changed_rows.append(_serialize_user_history_row(row))
            removed_ids = [tx_id for tx_id in last_signatures.keys() if tx_id not in next_signatures]
            if changed_rows or removed_ids:
                payload = {
                    "rows": changed_rows,
                    "removed_ids": removed_ids,
                    "generated_at": current_jakarta_time().isoformat(timespec="seconds"),
                }
                yield _event_payload("history_update", payload)
                last_signatures = next_signatures
                heartbeat_counter = 0
                continue

            heartbeat_counter += 1
            if heartbeat_counter >= 3:
                yield _event_payload("ping", {"ok": True})
                heartbeat_counter = 0

    response = Response(_generate(), mimetype="text/event-stream")
    response.headers["Cache-Control"] = "no-cache, no-store, must-revalidate"
    response.headers["X-Accel-Buffering"] = "no"
    response.headers["Connection"] = "keep-alive"
    return response


def _normalize_metric_events(raw_events) -> dict[str, int]:
    if not isinstance(raw_events, dict):
        return {}
    normalized: dict[str, int] = {}
    for key, raw_value in raw_events.items():
        event_name = str(key or "").strip().lower()
        if not event_name:
            continue
        try:
            value = int(raw_value)
        except (TypeError, ValueError):
            continue
        if value <= 0:
            continue
        normalized[event_name] = min(value, 1_000_000)
    return normalized


def _build_guestbook_ux_summary(rows: list[dict]) -> dict:
    now_jakarta = current_jakarta_time()
    event_totals: dict[str, int] = {}
    user_totals: dict[str, dict] = {}
    sessions: list[dict] = []

    for row in rows:
        payload = row.get("payload")
        if isinstance(payload, str):
            try:
                payload = json.loads(payload)
            except json.JSONDecodeError:
                payload = {}
        if not isinstance(payload, dict):
            payload = {}

        events = _normalize_metric_events(payload.get("events"))
        session_total = sum(events.values())
        for event_name, event_count in events.items():
            event_totals[event_name] = event_totals.get(event_name, 0) + event_count

        user_id = int(row.get("user_id") or 0)
        if user_id:
            entry = user_totals.setdefault(
                str(user_id),
                {
                    "user_id": user_id,
                    "full_name": row.get("full_name") or row.get("email") or f"User {user_id}",
                    "role": row.get("role") or "",
                    "events": 0,
                },
            )
            entry["events"] += session_total

        updated_at = row.get("updated_at")
        updated_label = ""
        if isinstance(updated_at, datetime):
            local_dt = to_jakarta(updated_at)
            updated_label = local_dt.strftime("%d %b %Y, %H:%M WIB")
        sessions.append(
            {
                "user_id": user_id,
                "full_name": row.get("full_name") or row.get("email") or "-",
                "role": row.get("role") or "-",
                "session_key": row.get("session_key") or "",
                "page_path": row.get("page_path") or "",
                "events": session_total,
                "last_event": payload.get("last_event") or "-",
                "updated_at_label": updated_label,
                "updated_at": updated_at,
            }
        )

    top_events = sorted(event_totals.items(), key=lambda item: item[1], reverse=True)[:12]
    top_users = sorted(user_totals.values(), key=lambda item: item.get("events", 0), reverse=True)[:10]
    sessions.sort(key=lambda item: item.get("updated_at") or datetime.min, reverse=True)
    recent_sessions = sessions[:25]
    total_events = sum(event_totals.values())

    active_24h = 0
    for row in rows:
        updated_at = row.get("updated_at")
        if isinstance(updated_at, datetime) and (now_jakarta - to_jakarta(updated_at)).total_seconds() <= 86400:
            active_24h += 1

    return {
        "total_sessions": len(rows),
        "active_users": len(user_totals),
        "active_sessions_24h": active_24h,
        "total_events": total_events,
        "top_events": top_events,
        "top_users": top_users,
        "recent_sessions": recent_sessions,
    }


@daftar_tamu_bp.route("/saya/riwayat/ux-metrics", methods=["POST"])
@role_required("staff", "coordinator")
def user_guestbook_ux_metrics_ingest() -> Response:
    """Persist UX telemetry snapshot for super-app history page."""
    user = current_user()
    payload = request.get_json(silent=True) or {}
    if not payload:
        payload = request.form.to_dict()
    if not isinstance(payload, dict):
        return jsonify({"success": False, "message": "Payload metrik tidak valid."}), 400

    metrics_raw = payload.get("metrics")
    if isinstance(metrics_raw, str):
        try:
            metrics_raw = json.loads(metrics_raw)
        except json.JSONDecodeError:
            metrics_raw = {}
    if not isinstance(metrics_raw, dict):
        metrics_json_raw = payload.get("metrics_json")
        if isinstance(metrics_json_raw, str):
            try:
                metrics_raw = json.loads(metrics_json_raw)
            except json.JSONDecodeError:
                metrics_raw = {}
        else:
            metrics_raw = {}

    session_key = (payload.get("session_key") or "").strip()
    if not session_key:
        return jsonify({"success": False, "message": "session_key wajib diisi."}), 400

    events = _normalize_metric_events(metrics_raw.get("events") if isinstance(metrics_raw, dict) else payload.get("events"))
    metric_payload = {
        "events": events,
        "updated_at": metrics_raw.get("updated_at") if isinstance(metrics_raw, dict) else payload.get("updated_at"),
        "last_event": metrics_raw.get("last_event") if isinstance(metrics_raw, dict) else payload.get("last_event"),
        "last_payload": metrics_raw.get("last_payload") if isinstance(metrics_raw, dict) else payload.get("last_payload"),
        "tab": (payload.get("tab") or "").strip(),
    }

    try:
        upsert_guestbook_ux_metrics(
            user_id=int(user.get("id")),
            session_key=session_key,
            payload=metric_payload,
            page_path=(payload.get("page_path") or request.path).strip(),
        )
    except ValueError as exc:
        return jsonify({"success": False, "message": str(exc)}), 400
    except Exception:
        current_app.logger.exception("Gagal menyimpan metrik UX daftar tamu.")
        return jsonify({"success": False, "message": "Gagal menyimpan metrik."}), 500

    return jsonify({"success": True})


@daftar_tamu_bp.route("/admin/ux-metrics")
@role_required("admin")
def admin_guestbook_ux_metrics() -> Response:
    """Admin dashboard for super-app UX telemetry."""
    days = max(1, min(_to_int(request.args.get("days"), 14), 90))
    rows = fetch_guestbook_ux_metric_rows(days=days, limit=800)
    summary = _build_guestbook_ux_summary(rows)
    return render_template(
        "daftar_tamu/admin_ux_metrics.html",
        days=days,
        summary=summary,
        today_str=_today_jakarta().isoformat(),
    )


@daftar_tamu_bp.route("/saya/riwayat/<int:transaction_id>/catatan-staf", methods=["POST"])
@role_required("staff", "coordinator")
def user_guestbook_staff_note(transaction_id: int) -> Response:
    user = current_user()
    note = (request.form.get("staff_note") or "").strip()
    staff_note_level = _normalize_staff_note_level(request.form.get("staff_note_level"), default="tindak_lanjut") or "tindak_lanjut"
    is_ajax = (
        request.headers.get("X-Requested-With") == "XMLHttpRequest"
        or "application/json" in (request.headers.get("Accept") or "").lower()
    )
    if len(note) > 500:
        if is_ajax:
            return jsonify({"success": False, "message": "Catatan staf maksimal 500 karakter."}), 400
        flash("Catatan staf maksimal 500 karakter.", "warning")
        return redirect(_build_user_guestbook_history_redirect(request.form))

    ok = upsert_transaction_staff_note(
        transaction_id=transaction_id,
        user_id=int(user.get("id")),
        note=note,
        level=staff_note_level,
    )
    if ok:
        message = "Catatan staf berhasil disimpan." if note else "Catatan staf dihapus."
        if is_ajax:
            updated_label = current_jakarta_time().strftime("%d %b %Y, %H:%M WIB") if note else ""
            signature = "|".join(
                [
                    str(transaction_id),
                    "approved",
                    "",
                    note,
                    staff_note_level if note else "",
                    updated_label,
                    "",
                ]
            )
            return jsonify(
                {
                    "success": True,
                    "message": message,
                    "transaction_id": transaction_id,
                    "status": "approved",
                    "staff_note": note,
                    "staff_note_level": staff_note_level if note else "",
                    "staff_note_level_label": _STAFF_NOTE_LEVEL_LABEL_MAP.get(staff_note_level, "") if note else "",
                    "staff_note_level_tone": _STAFF_NOTE_LEVEL_TONE_MAP.get(staff_note_level, "secondary") if note else "secondary",
                    "staff_note_updated_at": updated_label,
                    "signature": signature,
                }
            )
        flash(message, "success")
    else:
        if is_ajax:
            return jsonify(
                {
                    "success": False,
                    "message": "Catatan hanya bisa diisi jika foto transaksi sudah diverifikasi.",
                }
            ), 400
        flash("Catatan hanya bisa diisi jika foto transaksi sudah diverifikasi.", "warning")
    return redirect(_build_user_guestbook_history_redirect(request.form))


@daftar_tamu_bp.route("/sekolah/guest-search")
@role_required("sekolah")
def sekolah_guest_search() -> Response:
    query = (request.args.get("q") or "").strip()
    limit = _to_int(request.args.get("limit"), 20)
    results = list_guest_candidates(query, limit=limit)
    for item in results:
        item["profile_photo_url"] = _build_photo_url(item.get("profile_photo_path"), external=False)
    return jsonify({"success": True, "results": results})


@daftar_tamu_bp.route("/sekolah/guest-search-umum")
@role_required("sekolah")
def sekolah_general_guest_search() -> Response:
    query = (request.args.get("q") or "").strip()
    limit = _to_int(request.args.get("limit"), 20)
    results = list_general_guest_candidates(query, limit=limit)
    return jsonify({"success": True, "results": results})


@daftar_tamu_bp.route("/umum", methods=["POST"])
@role_required("sekolah", "admin")
def create_general_guest() -> Response:
    user = current_user()
    full_name = (request.form.get("full_name") or "").strip()
    email = (request.form.get("email") or "").strip().lower()
    phone = (request.form.get("phone") or "").strip()
    instansi = (request.form.get("instansi") or "").strip()
    jabatan = (request.form.get("jabatan") or "").strip()
    auto_verify = user.get("role") == "admin"
    if not full_name:
        return jsonify({"success": False, "message": "Nama tamu wajib diisi."}), 400
    if not instansi:
        return jsonify({"success": False, "message": "Instansi wajib diisi."}), 400
    if not jabatan:
        return jsonify({"success": False, "message": "Jabatan wajib diisi."}), 400
    if phone:
        phone = _sanitize_phone(phone)

    with get_cursor(commit=True) as cur:
        cur.execute(
            """
            INSERT INTO daftar_tamu_general_guests (
                full_name, email, phone, instansi, jabatan, created_by,
                is_verified, verified_by, verified_at
            )
            VALUES (%s, %s, %s, %s, %s, %s, %s, %s, CASE WHEN %s THEN NOW() ELSE NULL END)
            RETURNING id, full_name, email, phone, instansi, jabatan, is_verified
            """,
            (
                full_name,
                email or None,
                phone or None,
                instansi or None,
                jabatan or None,
                user.get("id"),
                auto_verify,
                user.get("id") if auto_verify else None,
                auto_verify,
            ),
        )
        row = cur.fetchone()
        guest = dict(row) if row else {}

        cur.execute(
            """
            SELECT COUNT(*) AS name_count,
                   MAX(CASE WHEN is_verified THEN 1 ELSE 0 END) AS has_verified
            FROM daftar_tamu_general_guests
            WHERE lower(full_name) = lower(%s)
            """,
            (full_name,),
        )
        stats = dict(cur.fetchone() or {})

    guest["has_duplicate"] = int(stats.get("name_count") or 0) > 1
    guest["has_verified"] = bool(stats.get("has_verified"))
    return jsonify({"success": True, "guest": guest})


@daftar_tamu_bp.route("/admin/umum/<int:guest_id>/update", methods=["POST"])
@role_required("admin")
def admin_update_general_guest(guest_id: int) -> Response:
    full_name = (request.form.get("full_name") or "").strip()
    phone = (request.form.get("phone") or "").strip()
    instansi = (request.form.get("instansi") or "").strip()
    jabatan = (request.form.get("jabatan") or "").strip()
    email = (request.form.get("email") or "").strip().lower()

    if not full_name:
        return jsonify({"success": False, "message": "Nama tamu wajib diisi."}), 400
    if not instansi:
        return jsonify({"success": False, "message": "Instansi wajib diisi."}), 400
    if not jabatan:
        return jsonify({"success": False, "message": "Jabatan wajib diisi."}), 400
    if phone:
        phone = _sanitize_phone(phone)

    with get_cursor(commit=True) as cur:
        cur.execute(
            """
            UPDATE daftar_tamu_general_guests
            SET full_name = %s,
                phone = %s,
                instansi = %s,
                jabatan = %s,
                email = %s,
                updated_at = NOW()
            WHERE id = %s
            """,
            (
                full_name,
                phone or None,
                instansi or None,
                jabatan or None,
                email or None,
                guest_id,
            ),
        )
        if cur.rowcount == 0:
            return jsonify({"success": False, "message": "Tamu umum tidak ditemukan."}), 404
    return jsonify({"success": True})


@daftar_tamu_bp.route("/sekolah/area-contacts")
@role_required("sekolah")
def sekolah_area_contacts() -> Response:
    """Return area contacts for the current school (used by guestbook modal)."""
    user = current_user()
    school = _fetch_school_for_user(user.get("id"))
    if not school:
        return jsonify({"success": False, "message": "Akun sekolah belum terhubung."}), 400
    message = None
    transaction_id_raw = request.args.get("transaction_id")
    if transaction_id_raw:
        try:
            transaction_id = int(transaction_id_raw)
        except (TypeError, ValueError):
            transaction_id = None
        if transaction_id:
            detail = get_transaction_detail(transaction_id)
            if detail and detail.get("school_id") == school.get("id"):
                guest_names = [
                    (g.get("full_name") or g.get("email") or "").strip()
                    for g in (detail.get("guests") or [])
                ]
                guest_names = [name for name in guest_names if name]
                guest_text = ", ".join(guest_names) if guest_names else "-"
                photo_url = None
                photo_path = detail.get("photo_path")
                if photo_path:
                    photo_name = photo_path.split("uploads/portal/")[-1]
                    photo_url = url_for("portal.uploaded_file", filename=photo_name, _external=True)
                message_lines = [
                    f"Halo, kami dari {school.get('name')} (NPSN {school.get('npsn')}) baru mengisi buku tamu.",
                    f"Tamu: {guest_text}",
                ]
                if photo_url:
                    message_lines.append(f"Foto: {photo_url}")
                message_lines.append("Mohon bantuan percepatan verifikasi.")
                message = "\n".join(message_lines)
    contacts = _build_area_contacts(school, message=message)
    return jsonify({"success": True, "contacts": contacts, "user_school": school})


@daftar_tamu_bp.route("/sekolah/umum-transactions/<int:transaction_id>/approve", methods=["POST"])
@role_required("sekolah")
def sekolah_approve_public_transaction(transaction_id: int) -> Response:
    user = current_user()
    school = _fetch_school_for_user(user.get("id"))
    if not school:
        return jsonify({"success": False, "message": "Akun sekolah belum terhubung."}), 400
    reviewer_notes = (request.form.get("reviewer_notes") or "").strip()
    success = update_public_transaction_status(
        transaction_id=transaction_id,
        status="approved",
        reviewer_id=user.get("id"),
        reviewer_notes=reviewer_notes or None,
        school_id=school.get("id"),
    )
    if not success:
        return jsonify({"success": False, "message": "Transaksi tidak ditemukan."}), 404
    try:
        _notify_guestbook_status_change(
            transaction_id=transaction_id,
            status="approved",
            actor=user,
            is_public=True,
        )
    except Exception:
        current_app.logger.exception("Gagal mengirim notifikasi Telegram status buku tamu umum.")
    return redirect(url_for("daftar_tamu.sekolah_public_web", _anchor="publicGuestbook"))


@daftar_tamu_bp.route("/sekolah/umum-transactions/<int:transaction_id>/reject", methods=["POST"])
@role_required("sekolah")
def sekolah_reject_public_transaction(transaction_id: int) -> Response:
    user = current_user()
    school = _fetch_school_for_user(user.get("id"))
    if not school:
        return jsonify({"success": False, "message": "Akun sekolah belum terhubung."}), 400
    reviewer_notes = (request.form.get("reviewer_notes") or "").strip()
    success = update_public_transaction_status(
        transaction_id=transaction_id,
        status="rejected",
        reviewer_id=user.get("id"),
        reviewer_notes=reviewer_notes or None,
        school_id=school.get("id"),
    )
    if not success:
        return jsonify({"success": False, "message": "Transaksi tidak ditemukan."}), 404
    try:
        _notify_guestbook_status_change(
            transaction_id=transaction_id,
            status="rejected",
            actor=user,
            is_public=True,
        )
    except Exception:
        current_app.logger.exception("Gagal mengirim notifikasi Telegram status buku tamu umum.")
    return redirect(url_for("daftar_tamu.sekolah_public_web", _anchor="publicGuestbook"))


@daftar_tamu_bp.route("/sekolah/umum-transactions/bulk-approve", methods=["POST"])
@role_required("sekolah")
def sekolah_bulk_approve_public_transactions() -> Response:
    user = current_user()
    school = _fetch_school_for_user(user.get("id"))
    if not school:
        flash("Akun sekolah belum terhubung dengan data sekolah.", "warning")
        return redirect(url_for("daftar_tamu.sekolah_public_web"))

    raw_ids = request.form.getlist("transaction_ids")
    ids: list[int] = []
    for raw in raw_ids:
        try:
            ids.append(int(raw))
        except (TypeError, ValueError):
            continue
    if not ids:
        flash("Pilih transaksi yang ingin disetujui.", "warning")
        return redirect(url_for("daftar_tamu.sekolah_public_web"))

    reviewer_notes = (request.form.get("reviewer_notes") or "").strip()
    success_count = 0
    for tx_id in ids:
        try:
            ok = update_public_transaction_status(
                transaction_id=tx_id,
                status="approved",
                reviewer_id=user.get("id"),
                reviewer_notes=reviewer_notes or None,
                school_id=school.get("id"),
            )
        except Exception:
            ok = False
        if ok:
            success_count += 1
            try:
                _notify_guestbook_status_change(
                    transaction_id=tx_id,
                    status="approved",
                    actor=user,
                    is_public=True,
                )
            except Exception:
                current_app.logger.exception("Gagal mengirim notifikasi Telegram status buku tamu umum.")

    if success_count:
        flash(f"{success_count} pengajuan berhasil disetujui.", "success")
    else:
        flash("Tidak ada pengajuan yang berhasil disetujui.", "warning")

    return redirect(url_for("daftar_tamu.sekolah_public_web"))


@daftar_tamu_bp.route("/sekolah/umum-transactions/bulk-reject", methods=["POST"])
@role_required("sekolah")
def sekolah_bulk_reject_public_transactions() -> Response:
    user = current_user()
    school = _fetch_school_for_user(user.get("id"))
    if not school:
        flash("Akun sekolah belum terhubung dengan data sekolah.", "warning")
        return redirect(url_for("daftar_tamu.sekolah_public_web"))

    raw_ids = request.form.getlist("transaction_ids")
    ids: list[int] = []
    for raw in raw_ids:
        try:
            ids.append(int(raw))
        except (TypeError, ValueError):
            continue
    if not ids:
        flash("Pilih transaksi yang ingin ditolak.", "warning")
        return redirect(url_for("daftar_tamu.sekolah_public_web"))

    reviewer_notes = (request.form.get("reviewer_notes") or "").strip()
    success_count = 0
    for tx_id in ids:
        try:
            ok = update_public_transaction_status(
                transaction_id=tx_id,
                status="rejected",
                reviewer_id=user.get("id"),
                reviewer_notes=reviewer_notes or None,
                school_id=school.get("id"),
            )
        except Exception:
            ok = False
        if ok:
            success_count += 1
            try:
                _notify_guestbook_status_change(
                    transaction_id=tx_id,
                    status="rejected",
                    actor=user,
                    is_public=True,
                )
            except Exception:
                current_app.logger.exception("Gagal mengirim notifikasi Telegram status buku tamu umum.")

    if success_count:
        flash(f"{success_count} pengajuan berhasil ditolak.", "success")
    else:
        flash("Tidak ada pengajuan yang berhasil ditolak.", "warning")

    return redirect(url_for("daftar_tamu.sekolah_public_web"))


@daftar_tamu_bp.route("/sekolah/qr")
@role_required("sekolah")
def sekolah_guestbook_qr() -> Response:
    user = current_user()
    school = _fetch_school_for_user(user.get("id"))
    if not school:
        return jsonify({"success": False, "message": "Akun sekolah belum terhubung."}), 400

    fmt = (request.args.get("format") or "png").strip().lower()
    paper = (request.args.get("paper") or "a4").strip().lower()
    if fmt not in {"png", "pdf"}:
        fmt = "png"
    if paper not in {"a4", "a5"}:
        paper = "a4"

    target_url = f"{_web_aska_base_url()}/buku-tamu/{school.get('npsn')}"
    qr_payload = _get_guestbook_qr_payload(school.get("id"))
    qr_base64 = None
    if qr_payload and qr_payload.get("png_base64") and qr_payload.get("url") == target_url:
        qr_base64 = qr_payload.get("png_base64")
    if not qr_base64:
        return jsonify({"success": False, "message": "QR belum dibuat. Silakan generate terlebih dahulu."}), 400
    try:
        qr_bytes = base64.b64decode(qr_base64)
        qr_img = Image.open(BytesIO(qr_bytes)).convert("RGBA")
    except Exception as exc:
        return jsonify({"success": False, "message": f"QR tersimpan tidak valid: {exc}"}), 500

    if fmt == "pdf":
        template_path = Path(__file__).resolve().parent.parent / "static" / "qr" / "new_template.png"
        if not template_path.exists():
            return jsonify({"success": False, "message": "Template QR tidak ditemukan."}), 500

        canvas = Image.open(template_path).convert("RGBA")
        base_w, base_h = 4419, 6250  # target high resolution
        
        # Upscale template if it's smaller than target base resolution
        # This ensures text, logos, and QR are rendered at high definition
        if canvas.width < base_w:
             canvas = canvas.resize((base_w, base_h), Image.LANCZOS)
             
        scale = canvas.width / base_w

        # Posisi & ukuran elemen (berdasarkan contoh.svg, diskalakan)
        qr_box_bbox = (1054, 2092, 3364, 4427)  # extracted from template.png
        qr_box_w = qr_box_bbox[2] - qr_box_bbox[0]
        qr_box_h = qr_box_bbox[3] - qr_box_bbox[1]
        qr_size_base = int(min(qr_box_w, qr_box_h) * 0.78)  # keep margin
        qr_center_x_base = (qr_box_bbox[0] + qr_box_bbox[2]) / 2
        qr_center_y_base = (qr_box_bbox[1] + qr_box_bbox[3]) / 2
        name_y_base = 4720
        web_label_y_base = 5440
        web_value_y_base = 5535
        ig_label_y_base = 5440
        ig_value_y_base = 5535
        web_x_base = 1480
        ig_x_base = 2920
        logo_center_base = (3800, 500)
        logo_diameter_base = 510

        qr_size = int(qr_size_base * scale)
        qr_x = int((qr_center_x_base - qr_size_base / 2) * scale)
        qr_y = int((qr_center_y_base - qr_size_base / 2) * scale)
        qr_resized = qr_img.resize((qr_size, qr_size), Image.LANCZOS).convert("RGBA")
        canvas.alpha_composite(qr_resized, (qr_x, qr_y))

        draw = ImageDraw.Draw(canvas)
        def _font(path: str, size: int):
            # Try finding bundled font first for server compatibility
            root_dir = Path(__file__).resolve().parent.parent.parent
            bundled_font = root_dir / "dashboard" / "static" / "fonts" / "Roboto-Bold.ttf"
            
            # List of potential font paths to try
            font_candidates = [
                bundled_font,
                path,
                f"/System/Library/Fonts/Supplemental/{path}",
                f"/System/Library/Fonts/{path}",
                f"/Library/Fonts/{path}",
                "/usr/share/fonts/truetype/dejavu/DejaVuSans-Bold.ttf", # Linux typical
                "/usr/share/fonts/truetype/liberation/LiberationSans-Bold.ttf" # Linux typical
            ]
            
            for font_path in font_candidates:
                try:
                    if os.path.exists(str(font_path)):
                        return ImageFont.truetype(str(font_path), int(size * scale))
                except Exception:
                    continue
            
            # Fallback if nothing works
            print("WARNING: All font paths failed, using default font (tiny)")
            return ImageFont.load_default()

        # Use generic name, but function will pick Roboto-Bold.ttf if available
        font_name = _font("Arial Bold.ttf", 220)
        font_small_value = _font("Arial Bold.ttf", 80)

        cx = canvas.width // 2
        draw.text((cx, int(name_y_base * scale)), (school.get("name") or "Nama Sekolah").upper(),
                  fill=(60, 70, 180, 255), font=font_name, anchor="mm")

        meta = school.get("metadata") or {}
        if isinstance(meta, str):
            try:
                meta = json.loads(meta)
            except json.JSONDecodeError:
                meta = {}
        def _shorten(text: str, max_len: int = 40) -> str:
            text = text or "-"
            return text if len(text) <= max_len else text[: max_len - 3] + "..."

        website_text = _shorten(meta.get("website") or "-")
        # Remove protocol and www
        for prefix in ["https://", "http://", "www."]:
             if website_text.lower().startswith(prefix):
                 website_text = website_text[len(prefix):]
        # Remove trailing slash
        website_text = website_text.rstrip("/")

        instagram_text = _shorten(meta.get("instagram") or meta.get("ig") or "-")
        if instagram_text != "-" and not instagram_text.startswith("@"):
             instagram_text = f"@{instagram_text}"

        web_value_y_base = 5190
        ig_value_y_base = 5190

        # Adjust X for left alignment (lm anchor)
        # web_x_base was 1150 (left), changing to 1200 (slightly right)
        # ig_x_base was 2600 (left)
        web_x_base = 1200
        ig_x_base = 2600

        draw.text((int(web_x_base * scale), int(web_value_y_base * scale)), website_text,
                  fill=(0, 0, 0, 255), font=font_small_value, anchor="lm")
        draw.text((int(ig_x_base * scale), int(ig_value_y_base * scale)), instagram_text,
                  fill=(0, 0, 0, 255), font=font_small_value, anchor="lm")

        logo_img = _load_school_logo(school)
        if logo_img:
            # Logo dimensions
            logo_size = int(logo_diameter_base * scale)
            
            # Process logo
            logo_img = logo_img.convert("RGBA")
            
            # Create ultra high quality circular mask
            # Render at 4x size then downscale for maximum anti-aliasing
            mask_scale_factor = 4
            mask_size = logo_size * mask_scale_factor
            
            mask = Image.new("L", (mask_size, mask_size), 0)
            mask_draw = ImageDraw.Draw(mask)
            mask_draw.ellipse((0, 0, mask_size, mask_size), fill=255)
            
            # Downscale mask perfectly
            mask = mask.resize((logo_size, logo_size), Image.LANCZOS)
            
            # Resize logo to fit
            # First resize to mask_size (upscale if needed) then downscale with LANCZOS
            logo_img = logo_img.resize((mask_size, mask_size), Image.LANCZOS)
            logo_img = logo_img.resize((logo_size, logo_size), Image.LANCZOS)
            
            # Enhance sharpness slightly after downscaling to prevent blur
            from PIL import ImageFilter
            logo_img = logo_img.filter(ImageFilter.SHARPEN)
            
            # Calculate position to center in the circle
            logo_pos = (
                int(logo_center_base[0] * scale - logo_size / 2),
                int(logo_center_base[1] * scale - logo_size / 2)
            )
            
            # Paste logo with mask
            canvas.paste(logo_img, logo_pos, mask)

        buf = BytesIO()
        canvas.convert("RGB").save(buf, format="PDF")
        buf.seek(0)
        filename = f"qr_buku_tamu_{school.get('npsn')}_{paper}.pdf"
        return send_file(
            buf,
            mimetype="application/pdf",
            as_attachment=True,
            download_name=filename,
        )

    buf = BytesIO()
    qr_img.convert("RGB").save(buf, format="PNG")
    buf.seek(0)
    filename = f"qr_buku_tamu_{school.get('npsn')}.png"
    return send_file(
        buf,
        mimetype="image/png",
        as_attachment=True,
        download_name=filename,
    )


@daftar_tamu_bp.route("/sekolah/qr/generate", methods=["POST"])
@role_required("sekolah")
def sekolah_generate_guestbook_qr() -> Response:
    user = current_user()
    school = _fetch_school_for_user(user.get("id"))
    if not school:
        flash("Akun sekolah belum terhubung.", "danger")
        return redirect(url_for("daftar_tamu.sekolah_public_web"))

    target_url = f"{_web_aska_base_url()}/buku-tamu/{school.get('npsn')}"
    
    # Resolve logo path
    logo_path = None
    logo_url = school.get("logo_url")
    if logo_url and "/portal/uploads/" in logo_url:
        # Resolve URL to local path
        # URL format: /portal/uploads/<filename>
        # Local upload dir: base_dir/uploads/portal
        try:
            filename = logo_url.split("/portal/uploads/")[-1]
            base_dir = Path(__file__).resolve().parents[2]
            upload_folder = base_dir / "uploads" / "portal"
            candidate_path = upload_folder / filename
            if candidate_path.exists():
                logo_path = candidate_path
        except Exception:
            pass

    qr_img = _build_guestbook_qr(target_url, size=1024, logo_path=logo_path)
    buf = BytesIO()
    qr_img.convert("RGB").save(buf, format="PNG")
    qr_base64 = base64.b64encode(buf.getvalue()).decode("ascii")

    payload = {
        "url": target_url,
        "png_base64": qr_base64,
        "generated_at": current_jakarta_time().isoformat(),
    }
    _store_guestbook_qr_payload(school.get("id"), payload)
    flash("QR Buku Tamu berhasil dibuat.", "success")
    return redirect(url_for("daftar_tamu.sekolah_public_web"))


@daftar_tamu_bp.route("/sekolah/transactions", methods=["POST"])
@role_required("sekolah")
def sekolah_create_transaction() -> Response:
    user = current_user()
    school = _fetch_school_for_user(user["id"])
    if not school:
        return jsonify({"success": False, "message": "Akun sekolah belum terhubung."}), 400

    sudin_ids, umum_ids = _parse_guest_payload(request.form.get("guest_payload"))
    if not sudin_ids and not umum_ids:
        sudin_ids = _parse_guest_ids(request.form.get("guest_ids"))
    if not sudin_ids and not umum_ids:
        return jsonify({"success": False, "message": "Pilih minimal satu tamu."}), 400

    purpose = (request.form.get("purpose") or "").strip()
    notes = (request.form.get("notes") or "").strip()
    duplicate_confirmed = _is_truthy(request.form.get("duplicate_confirmed"))

    latitude_raw = request.form.get("latitude")
    longitude_raw = request.form.get("longitude")
    accuracy_raw = request.form.get("accuracy")
    try:
        latitude = float(latitude_raw)
        longitude = float(longitude_raw)
    except (TypeError, ValueError):
        return jsonify({"success": False, "message": "Lokasi GPS tidak valid."}), 400

    if not (-90 <= latitude <= 90 and -180 <= longitude <= 180):
        return jsonify({"success": False, "message": "Lokasi GPS di luar jangkauan."}), 400

    if "photo" not in request.files:
        return jsonify({"success": False, "message": "Foto wajib diunggah."}), 400

    photo = request.files["photo"]
    if not photo or not photo.filename:
        return jsonify({"success": False, "message": "Foto wajib diunggah."}), 400

    visit_at = current_jakarta_time()
    duplicate_rows = _find_sudin_same_day_approved_duplicates(
        school_id=int(school["id"]),
        sudin_ids=sudin_ids,
        visit_at=visit_at,
    )
    duplicate_repeat_count = 0
    if duplicate_rows:
        try:
            duplicate_repeat_count = max(
                int(row.get("approved_count") or 0) + 1
                for row in duplicate_rows
            )
        except Exception:
            duplicate_repeat_count = 0
    if duplicate_rows and not duplicate_confirmed:
        warning_message = _build_sudin_duplicate_warning_message(
            school_name=school.get("name"),
            duplicate_rows=duplicate_rows,
        )
        return jsonify(
            {
                "success": False,
                "requires_confirmation": True,
                "message": warning_message,
                "duplicate_guest_count": len(duplicate_rows),
            }
        ), 409

    try:
        stamp_result = stamp_guestbook_photo(
            file_storage=photo,
            latitude=latitude,
            longitude=longitude,
            captured_at=visit_at,
            school_label=school.get("name"),
        )
    except Exception as exc:
        return jsonify({"success": False, "message": f"Gagal memproses foto: {exc}"}), 500

    metadata = {
        "accuracy": float(accuracy_raw) if accuracy_raw else None,
        "user_agent": request.headers.get("User-Agent"),
        "map_provider": stamp_result.get("map_provider"),
        "map_error": stamp_result.get("map_error"),
    }

    has_sudin = bool(sudin_ids)
    has_umum = bool(umum_ids)
    pending_transaction_id = None
    approved_transaction_id = None

    def _insert_transaction(
        cur,
        *,
        guest_type: str,
        guest_ids: list[int],
        status_value: str,
        reviewed_by: Optional[int],
        reviewed_at: Optional[datetime],
        reviewer_notes: Optional[str],
    ) -> Optional[int]:
        if not guest_ids:
            return None
        cur.execute(
            """
            INSERT INTO daftar_tamu_transactions (
                school_id,
                visit_at,
                purpose,
                notes,
                photo_path,
                photo_raw_path,
                latitude,
                longitude,
                status,
                reviewed_by,
                reviewed_at,
                reviewer_notes,
                created_by,
                metadata
            )
            VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
            RETURNING id
            """,
            (
                school["id"],
                visit_at,
                purpose or None,
                notes or None,
                stamp_result["stamped_path"],
                stamp_result.get("raw_path"),
                latitude,
                longitude,
                status_value,
                reviewed_by,
                reviewed_at,
                reviewer_notes,
                user["id"],
                Json(metadata),
            ),
        )
        tx_row = cur.fetchone()
        transaction_id = int(tx_row["id"]) if tx_row else None
        if not transaction_id:
            return None
        if guest_type == "sudin":
            for guest_id in guest_ids:
                cur.execute(
                    """
                    INSERT INTO daftar_tamu_transaction_guests (transaction_id, guest_type, user_id)
                    VALUES (%s, 'sudin', %s)
                    ON CONFLICT (transaction_id, user_id) DO NOTHING
                    """,
                    (transaction_id, guest_id),
                )
        else:
            for guest_id in guest_ids:
                cur.execute(
                    """
                    INSERT INTO daftar_tamu_transaction_guests (transaction_id, guest_type, general_guest_id)
                    VALUES (%s, 'umum', %s)
                    ON CONFLICT (transaction_id, general_guest_id) DO NOTHING
                    """,
                    (transaction_id, guest_id),
                )
        return transaction_id

    with get_cursor(commit=True) as cur:
        if has_sudin:
            pending_transaction_id = _insert_transaction(
                cur,
                guest_type="sudin",
                guest_ids=sudin_ids,
                status_value="pending",
                reviewed_by=None,
                reviewed_at=None,
                reviewer_notes=None,
            )
        if has_umum:
            approved_transaction_id = _insert_transaction(
                cur,
                guest_type="umum",
                guest_ids=umum_ids,
                status_value="approved",
                reviewed_by=user.get("id"),
                reviewed_at=visit_at,
                reviewer_notes="Auto konfirmasi sekolah",
            )

    transaction_id = pending_transaction_id or approved_transaction_id

    if pending_transaction_id:
        try:
            from dashboard.telegram_notifications import notify_guestbook_request
            import threading

            detail = get_transaction_detail(pending_transaction_id)
            photo_links = _build_guestbook_photo_links(
                transaction_id=pending_transaction_id,
                detail=detail,
            )
            guest_names = _extract_guest_names_from_detail(detail)
            
            # Use current_app.app_context() inside the thread
            app = current_app._get_current_object()

            def _send_notification():
                with app.app_context():
                    try:
                        notify_guestbook_request(
                            transaction_id=pending_transaction_id,
                            school_name=school.get("name") or "Sekolah",
                            npsn=None,
                            visit_at=visit_at,
                            guest_summary=None,
                            guest_names=guest_names,
                            duplicate_repeat_count=duplicate_repeat_count or None,
                            purpose=purpose or None,
                            notes=notes or None,
                            photo_links=photo_links,
                            photo_file_path=(detail or {}).get("photo_path"),
                        )
                    except Exception:
                        app.logger.exception("Gagal mengirim notifikasi buku tamu di background thread.")

            threading.Thread(target=_send_notification, daemon=True).start()

        except Exception:
            current_app.logger.exception("Gagal menyiapkan notifikasi buku tamu.")

    return jsonify({"success": True, "transaction_id": transaction_id})
