"""Queries for the Laporan (Form Reports) system."""
from __future__ import annotations

import json
from io import BytesIO
from datetime import datetime, time
from typing import Optional

from ..db_access import get_cursor


# ─────────────────────────────────────────────
# Forms
# ─────────────────────────────────────────────

def list_all_forms(include_inactive: bool = False) -> list[dict]:
    """List all forms (for admin)."""
    with get_cursor() as cur:
        where = "" if include_inactive else "WHERE f.is_active = TRUE AND f.status = 'published'"
        cur.execute(
            f"""
            SELECT f.id, f.title, f.description, f.target_scope, f.target_jenjang,
                   f.allow_multiple, f.allow_late, f.very_late_after_minutes,
                   f.no_submission_after_minutes, f.no_submission_jenjangs, f.no_submission_statuses,
                   f.is_active, f.is_paused, f.status, f.repeat_policy,
                   f.repeat_until_at, f.repeat_deadline_time, f.repeat_deadline_day,
                   f.deadline_at, f.created_at,
                   u.full_name AS created_by_name,
                   (SELECT COUNT(*) FROM laporan_submissions s WHERE s.form_id = f.id AND s.status = 'submitted') AS submission_count
            FROM laporan_forms f
            LEFT JOIN dashboard_users u ON u.id = f.created_by
            {where}
            ORDER BY f.created_at DESC
            """
        )
        return [dict(r) for r in cur.fetchall()]


def get_form(form_id: int) -> Optional[dict]:
    """Fetch one form by id."""
    with get_cursor() as cur:
        cur.execute(
            """
            SELECT f.*, u.full_name AS created_by_name
            FROM laporan_forms f
            LEFT JOIN dashboard_users u ON u.id = f.created_by
            WHERE f.id = %s
            """,
            (form_id,),
        )
        row = cur.fetchone()
        return dict(row) if row else None


def get_form_fields(form_id: int) -> list[dict]:
    """Fetch all fields for a form, ordered by sort_order."""
    with get_cursor() as cur:
        cur.execute(
            """
            SELECT id, form_id, label, field_type, options_json, required, sort_order
            FROM laporan_form_fields
            WHERE form_id = %s
            ORDER BY sort_order, id
            """,
            (form_id,),
        )
        rows = []
        for r in cur.fetchall():
            d = dict(r)
            if d.get("options_json") and isinstance(d["options_json"], str):
                try:
                    d["options_json"] = json.loads(d["options_json"])
                except Exception:
                    d["options_json"] = []
            raw_options = d.get("options_json")
            d["field_key"] = f"db_{d['id']}"
            if isinstance(raw_options, dict):
                d["field_key"] = raw_options.get("field_key") or d["field_key"]
                if raw_options.get("kind") == "formula":
                    d["field_type"] = "formula"
                    d["options_json"] = {
                        "left_key": raw_options.get("left_key") or "",
                        "operator": raw_options.get("operator") or "subtract",
                        "right_key": raw_options.get("right_key") or "",
                    }
                elif raw_options.get("kind") == "link":
                    d["field_type"] = "link"
                    d["options_json"] = {
                        "url": raw_options.get("url") or "",
                        "button_text": raw_options.get("button_text") or "Buka Link",
                    }
                else:
                    d["options_json"] = raw_options.get("choices") or []
            rows.append(d)
        return rows


def get_form_target_school_ids(form_id: int) -> list[int]:
    """Get list of school_ids targeted by a specific form."""
    with get_cursor() as cur:
        cur.execute(
            "SELECT school_id FROM laporan_form_targets WHERE form_id = %s",
            (form_id,),
        )
        return [r["school_id"] for r in cur.fetchall()]


def list_forms_for_school(school_id: int, jenjang: Optional[str] = None) -> list[dict]:
    """
    Return active forms visible to a particular school.
    Includes scope=all, scope=jenjang (matching school jenjang), scope=specific (this school targeted).
    """
    with get_cursor() as cur:
        cur.execute(
            """
            SELECT DISTINCT f.id, f.title, f.description, f.target_scope,
                   f.allow_multiple, f.allow_late, f.very_late_after_minutes,
                   f.no_submission_after_minutes, f.no_submission_jenjangs, f.no_submission_statuses,
                   f.is_paused, f.repeat_policy, f.repeat_until_at,
                   f.repeat_deadline_time, f.repeat_deadline_day, f.deadline_at, f.created_at,
                   (
                       SELECT COUNT(*) FROM laporan_submissions s
                       WHERE s.form_id = f.id AND s.school_id = %s AND s.status = 'submitted'
                   ) AS submission_count
            FROM laporan_forms f
            LEFT JOIN laporan_form_targets ft ON ft.form_id = f.id
            WHERE f.is_active = TRUE
              AND f.status = 'published'
              AND (
                  f.target_scope = 'all'
                  OR (f.target_scope = 'jenjang' AND f.target_jenjang = %s)
                  OR (f.target_scope = 'specific' AND ft.school_id = %s)
              )
            ORDER BY f.created_at DESC
            """,
            (school_id, jenjang or "", school_id),
        )
        return [dict(r) for r in cur.fetchall()]


def can_school_access_form(form_id: int, school_id: int, jenjang: Optional[str] = None) -> bool:
    """Check if a school is allowed to access a specific form based on its target_scope."""
    with get_cursor() as cur:
        cur.execute(
            """
            SELECT 1 FROM laporan_forms f
            LEFT JOIN laporan_form_targets ft ON ft.form_id = f.id
            WHERE f.id = %s
              AND f.is_active = TRUE
              AND f.status = 'published'
              AND (
                  f.target_scope = 'all'
                  OR (f.target_scope = 'jenjang' AND f.target_jenjang = %s)
                  OR (f.target_scope = 'specific' AND ft.school_id = %s)
              )
            LIMIT 1
            """,
            (form_id, jenjang or "", school_id),
        )
        return cur.fetchone() is not None


def create_form(
    title: str,
    description: str,
    target_scope: str,
    target_jenjang: Optional[str],
    allow_multiple: bool,
    allow_late: bool,
    very_late_after_minutes: int,
    no_submission_after_minutes: Optional[int],
    no_submission_jenjangs: Optional[str],
    no_submission_statuses: Optional[str],
    is_active: bool,
    deadline_at: Optional[datetime],
    created_by: int,
    status: str = "published",
    repeat_policy: str = "once",
    repeat_until_at: Optional[datetime] = None,
    repeat_deadline_time: Optional[time] = None,
    repeat_deadline_day: Optional[int] = None,
) -> dict:
    """Insert new laporan form and return the created row."""
    with get_cursor(commit=True) as cur:
        cur.execute(
            """
            INSERT INTO laporan_forms
                (title, description, target_scope, target_jenjang, allow_multiple, allow_late,
                 very_late_after_minutes, no_submission_after_minutes, no_submission_jenjangs,
                 no_submission_statuses,
                 is_active, status, repeat_policy, repeat_until_at, repeat_deadline_time,
                 repeat_deadline_day, deadline_at, created_by, updated_by)
            VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
            RETURNING id, title, status, created_at
            """,
            (
                title,
                description,
                target_scope,
                target_jenjang,
                allow_multiple,
                allow_late,
                very_late_after_minutes,
                no_submission_after_minutes,
                no_submission_jenjangs,
                no_submission_statuses,
                is_active,
                status,
                repeat_policy,
                repeat_until_at,
                repeat_deadline_time,
                repeat_deadline_day,
                deadline_at,
                created_by,
                created_by,
            ),
        )
        return dict(cur.fetchone())


def update_form(
    form_id: int,
    title: str,
    description: str,
    target_scope: str,
    target_jenjang: Optional[str],
    allow_multiple: bool,
    allow_late: bool,
    very_late_after_minutes: int,
    no_submission_after_minutes: Optional[int],
    no_submission_jenjangs: Optional[str],
    no_submission_statuses: Optional[str],
    is_active: bool,
    deadline_at: Optional[datetime],
    updated_by: int,
    status: str = "published",
    repeat_policy: str = "once",
    repeat_until_at: Optional[datetime] = None,
    repeat_deadline_time: Optional[time] = None,
    repeat_deadline_day: Optional[int] = None,
) -> None:
    """Update laporan form metadata."""
    with get_cursor(commit=True) as cur:
        cur.execute(
            """
            UPDATE laporan_forms SET
                title=%s, description=%s, target_scope=%s, target_jenjang=%s,
                allow_multiple=%s, allow_late=%s, very_late_after_minutes=%s,
                no_submission_after_minutes=%s, no_submission_jenjangs=%s,
                no_submission_statuses=%s,
                is_active=%s, status=%s,
                repeat_policy=%s, repeat_until_at=%s, repeat_deadline_time=%s,
                repeat_deadline_day=%s, deadline_at=%s,
                updated_by=%s, updated_at=NOW()
            WHERE id=%s
            """,
            (
                title,
                description,
                target_scope,
                target_jenjang,
                allow_multiple,
                allow_late,
                very_late_after_minutes,
                no_submission_after_minutes,
                no_submission_jenjangs,
                no_submission_statuses,
                is_active,
                status,
                repeat_policy,
                repeat_until_at,
                repeat_deadline_time,
                repeat_deadline_day,
                deadline_at,
                updated_by,
                form_id,
            ),
        )


def delete_form(form_id: int) -> None:
    with get_cursor(commit=True) as cur:
        cur.execute("DELETE FROM laporan_forms WHERE id = %s", (form_id,))


def set_form_paused(form_id: int, is_paused: bool, updated_by: int) -> None:
    """Pause/unpause a published form without changing active/draft status."""
    with get_cursor(commit=True) as cur:
        cur.execute(
            """
            UPDATE laporan_forms
            SET is_paused = %s, updated_by = %s, updated_at = NOW()
            WHERE id = %s
            """,
            (is_paused, updated_by, form_id),
        )


def set_form_targets(form_id: int, school_ids: list[int]) -> None:
    """Replace all specific targets for a form."""
    with get_cursor(commit=True) as cur:
        cur.execute("DELETE FROM laporan_form_targets WHERE form_id = %s", (form_id,))
        if school_ids:
            values = [(form_id, sid) for sid in school_ids]
            cur.executemany(
                "INSERT INTO laporan_form_targets (form_id, school_id) VALUES (%s, %s) ON CONFLICT DO NOTHING",
                values,
            )


def replace_form_fields(form_id: int, fields: list[dict]) -> None:
    """
    Replace/Update all fields for a form without wiping out existing database answers.
    Fields with matching field_key are updated, deleted ones are removed, new ones are inserted.
    """
    with get_cursor(commit=True) as cur:
        # Get existing fields for this form
        cur.execute("SELECT id, options_json->>'field_key' AS field_key FROM laporan_form_fields WHERE form_id = %s", (form_id,))
        existing = {row["field_key"]: row["id"] for row in cur.fetchall() if row.get("field_key")}
        
        # Track which IDs are kept
        kept_ids = []
        
        for i, f in enumerate(fields):
            field_type = f.get("field_type", "text")
            db_field_type = "number" if field_type == "formula" else field_type
            field_key = f.get("field_key") or f"f_{i}"
            raw_options = f.get("options_json")
            if field_type == "formula":
                raw_options = raw_options if isinstance(raw_options, dict) else {}
                options = {
                    "kind": "formula",
                    "field_key": field_key,
                    "left_key": raw_options.get("left_key") or "",
                    "operator": raw_options.get("operator") or "subtract",
                    "right_key": raw_options.get("right_key") or "",
                }
            elif field_type in {"radio", "checkbox", "dropdown"}:
                options = {"field_key": field_key, "choices": raw_options or []}
            elif field_type == "link":
                raw_options = raw_options if isinstance(raw_options, dict) else {}
                options = {
                    "kind": "link",
                    "field_key": field_key,
                    "url": raw_options.get("url") or "",
                    "button_text": raw_options.get("button_text") or "Buka Link",
                }
            else:
                options = {"field_key": field_key}
            options = json.dumps(options, ensure_ascii=False)
            
            existing_id = existing.get(field_key)
            if existing_id:
                # Update existing
                cur.execute(
                    """
                    UPDATE laporan_form_fields
                    SET label = %s, field_type = %s, options_json = %s::jsonb, required = %s, sort_order = %s
                    WHERE id = %s
                    """,
                    (
                        f["label"],
                        db_field_type,
                        options,
                        f.get("required", True),
                        f.get("sort_order", i),
                        existing_id,
                    ),
                )
                kept_ids.append(existing_id)
            else:
                # Insert new
                cur.execute(
                    """
                    INSERT INTO laporan_form_fields (form_id, label, field_type, options_json, required, sort_order)
                    VALUES (%s, %s, %s, %s::jsonb, %s, %s)
                    RETURNING id
                    """,
                    (
                        form_id,
                        f["label"],
                        db_field_type,
                        options,
                        f.get("required", True),
                        f.get("sort_order", i),
                    ),
                )
                new_id = cur.fetchone()["id"]
                kept_ids.append(new_id)

        # Delete fields that are no longer in the form
        if kept_ids:
            cur.execute(
                "DELETE FROM laporan_form_fields WHERE form_id = %s AND id NOT IN %s",
                (form_id, tuple(kept_ids)),
            )
        else:
            cur.execute(
                "DELETE FROM laporan_form_fields WHERE form_id = %s",
                (form_id,),
            )


# ─────────────────────────────────────────────
# Submissions
# ─────────────────────────────────────────────

def school_has_submitted(form_id: int, school_id: int) -> bool:
    """Check if a school already has a submitted (not draft) submission for a form."""
    with get_cursor() as cur:
        cur.execute(
            """
            SELECT 1 FROM laporan_submissions
            WHERE form_id=%s AND school_id=%s AND status='submitted'
            LIMIT 1
            """,
            (form_id, school_id),
        )
        return cur.fetchone() is not None


def school_has_submitted_for_period(form_id: int, school_id: int, repeat_period_key: str) -> bool:
    """Check if a school already submitted a form for the given repeat period."""
    if not repeat_period_key:
        return False
    with get_cursor() as cur:
        cur.execute(
            """
            SELECT 1 FROM laporan_submissions
            WHERE form_id=%s
              AND school_id=%s
              AND status='submitted'
              AND repeat_period_key=%s
            LIMIT 1
            """,
            (form_id, school_id, repeat_period_key),
        )
        return cur.fetchone() is not None


def create_submission(
    form_id: int,
    school_id: int,
    submitted_by: int,
    is_late: bool = False,
    late_days: int = 0,
    late_minutes: int = 0,
    repeat_period_key: Optional[str] = None,
    repeat_period_label: Optional[str] = None,
) -> dict:
    """Create a new submission record (submitted status) and return it."""
    with get_cursor(commit=True) as cur:
        cur.execute(
            """
            INSERT INTO laporan_submissions (
                form_id, school_id, submitted_by, status, submitted_at,
                is_late, late_days, late_minutes, repeat_period_key, repeat_period_label
            )
            VALUES (%s, %s, %s, 'submitted', NOW(), %s, %s, %s, %s, %s)
            RETURNING id, form_id, school_id, submitted_at
            """,
            (
                form_id,
                school_id,
                submitted_by,
                is_late,
                late_days,
                late_minutes,
                repeat_period_key,
                repeat_period_label,
            ),
        )
        return dict(cur.fetchone())


def save_answer(submission_id: int, field_id: int, answer_text: Optional[str], answer_json=None) -> int:
    """Upsert an answer for a submission+field. Returns answer id."""
    with get_cursor(commit=True) as cur:
        json_val = None
        if answer_json is not None:
            json_val = json.dumps(answer_json, ensure_ascii=False) if not isinstance(answer_json, str) else answer_json
        cur.execute(
            """
            INSERT INTO laporan_submission_answers (submission_id, field_id, answer_text, answer_json)
            VALUES (%s, %s, %s, %s::jsonb)
            ON CONFLICT (submission_id, field_id)
            DO UPDATE SET answer_text=EXCLUDED.answer_text, answer_json=EXCLUDED.answer_json
            RETURNING id
            """,
            (submission_id, field_id, answer_text, json_val),
        )
        return cur.fetchone()["id"]


def save_file(answer_id: int, file_path: str, original_name: str, mime_type: str, size_bytes: int) -> None:
    with get_cursor(commit=True) as cur:
        cur.execute(
            """
            INSERT INTO laporan_submission_files (answer_id, file_path, original_name, mime_type, size_bytes)
            VALUES (%s, %s, %s, %s, %s)
            """,
            (answer_id, file_path, original_name, mime_type, size_bytes),
        )


def get_submission_with_answers(submission_id: int) -> Optional[dict]:
    """Fetch one submission + all its answers."""
    with get_cursor() as cur:
        cur.execute(
            """
            SELECT s.*, COALESCE(s.late_minutes, s.late_days * 1440, 0) AS late_minutes,
                   f.very_late_after_minutes,
                   f.deadline_at AS form_deadline_at,
                   f.repeat_policy AS form_repeat_policy,
                   f.repeat_deadline_time AS form_repeat_deadline_time,
                   f.repeat_deadline_day AS form_repeat_deadline_day,
                   sc.name AS school_name, sc.npsn, sc.jenjang, u.full_name AS submitted_by_name
            FROM laporan_submissions s
            JOIN laporan_forms f ON f.id = s.form_id
            JOIN portal_schools sc ON sc.id = s.school_id
            LEFT JOIN dashboard_users u ON u.id = s.submitted_by
            WHERE s.id = %s
            """,
            (submission_id,),
        )
        row = cur.fetchone()
        if not row:
            return None
        sub = dict(row)

        cur.execute(
            """
            SELECT a.id AS answer_id, a.field_id, a.answer_text, a.answer_json,
                   ff.label, ff.field_type, ff.sort_order
            FROM laporan_submission_answers a
            JOIN laporan_form_fields ff ON ff.id = a.field_id
            WHERE a.submission_id = %s
            ORDER BY ff.sort_order, ff.id
            """,
            (submission_id,),
        )
        answers = []
        for r in cur.fetchall():
            d = dict(r)
            if d.get("answer_json") and isinstance(d["answer_json"], str):
                try:
                    d["answer_json"] = json.loads(d["answer_json"])
                except Exception:
                    pass
            # Fetch files for this answer
            cur.execute(
                "SELECT id, file_path, original_name, mime_type FROM laporan_submission_files WHERE answer_id = %s",
                (d["answer_id"],),
            )
            d["files"] = [dict(f) for f in cur.fetchall()]
            answers.append(d)
        sub["answers"] = answers
        return sub


def delete_submitted_submission(form_id: int, submission_id: int) -> Optional[dict]:
    """
    Delete one submitted laporan history row and return metadata for cleanup/flash.
    Draft and no_submission rows are intentionally ignored here.
    """
    with get_cursor(commit=True) as cur:
        cur.execute(
            """
            SELECT s.id, s.form_id, s.school_id, s.status, s.repeat_period_key,
                   s.repeat_period_label, sc.name AS school_name,
                   ARRAY_REMOVE(ARRAY_AGG(sf.file_path), NULL) AS file_paths
            FROM laporan_submissions s
            JOIN portal_schools sc ON sc.id = s.school_id
            LEFT JOIN laporan_submission_answers a ON a.submission_id = s.id
            LEFT JOIN laporan_submission_files sf ON sf.answer_id = a.id
            WHERE s.id = %s AND s.form_id = %s AND s.status = 'submitted'
            GROUP BY s.id, s.form_id, s.school_id, s.status, s.repeat_period_key,
                     s.repeat_period_label, sc.name
            """,
            (submission_id, form_id),
        )
        row = cur.fetchone()
        if not row:
            return None
        deleted = dict(row)
        cur.execute(
            "DELETE FROM laporan_submissions WHERE id = %s AND form_id = %s AND status = 'submitted'",
            (submission_id, form_id),
        )
        return deleted


def delete_empty_submitted_submissions(form_id: int, repeat_period_key: Optional[str] = None) -> int:
    """Delete submitted rows that have no saved answers, optionally scoped to one period."""
    params = [form_id]
    period_filter = ""
    if repeat_period_key:
        period_filter = "AND s.repeat_period_key = %s"
        params.append(repeat_period_key)

    with get_cursor(commit=True) as cur:
        cur.execute(
            f"""
            DELETE FROM laporan_submissions s
            WHERE s.form_id = %s
              AND s.status = 'submitted'
              {period_filter}
              AND NOT EXISTS (
                  SELECT 1
                  FROM laporan_submission_answers a
                  WHERE a.submission_id = s.id
              )
            """,
            tuple(params),
        )
        return cur.rowcount


def list_school_submissions(school_id: int) -> list[dict]:
    """List all submissions by a school (for school's own history view)."""
    with get_cursor() as cur:
        cur.execute(
            """
            SELECT s.id, s.form_id, s.status, s.submitted_at, s.created_at,
                   s.is_late, s.late_days, COALESCE(s.late_minutes, s.late_days * 1440, 0) AS late_minutes,
                   s.repeat_period_key, s.repeat_period_label,
                   f.title AS form_title, f.description AS form_description,
                   f.very_late_after_minutes,
                   f.deadline_at AS form_deadline_at,
                   f.repeat_policy AS form_repeat_policy,
                   f.repeat_deadline_time AS form_repeat_deadline_time,
                   f.repeat_deadline_day AS form_repeat_deadline_day
            FROM laporan_submissions s
            JOIN laporan_forms f ON f.id = s.form_id
            WHERE s.school_id = %s
            ORDER BY s.submitted_at DESC NULLS LAST, s.created_at DESC
            """,
            (school_id,),
        )
        return [dict(r) for r in cur.fetchall()]


def get_last_submission_answers(form_id: int, school_id: int) -> Optional[dict]:
    """
    Fetch the most recent submitted answers for a given form+school.
    Returns a dict with:
      - period_label: label periode submission tersebut
      - answers: {field_id -> {answer_text, answer_json, field_type}}
    Returns None if no previous submission exists.
    Uses only 2 queries (no N+1).
    """
    with get_cursor() as cur:
        # Query 1: ambil submission terakhir
        cur.execute(
            """
            SELECT s.id, s.repeat_period_label
            FROM laporan_submissions s
            WHERE s.form_id = %s AND s.school_id = %s AND s.status = 'submitted'
            ORDER BY s.submitted_at DESC NULLS LAST, s.id DESC
            LIMIT 1
            """,
            (form_id, school_id),
        )
        row = cur.fetchone()
        if not row:
            return None
        submission_id = row["id"]
        period_label = row["repeat_period_label"]

        # Query 2: batch ambil semua jawaban sekaligus
        cur.execute(
            """
            SELECT a.field_id, a.answer_text, a.answer_json,
                   ff.field_type, ff.options_json
            FROM laporan_submission_answers a
            JOIN laporan_form_fields ff ON ff.id = a.field_id
            WHERE a.submission_id = %s
            ORDER BY ff.sort_order, ff.id
            """,
            (submission_id,),
        )
        answers = {}
        for r in cur.fetchall():
            d = dict(r)
            answer_json = d.get("answer_json")
            if answer_json and isinstance(answer_json, str):
                try:
                    answer_json = json.loads(answer_json)
                except Exception:
                    answer_json = None
            answers[str(d["field_id"])] = {
                "answer_text": d.get("answer_text"),
                "answer_json": answer_json,
                "field_type": d.get("field_type"),
            }

        return {
            "period_label": period_label,
            "answers": answers,
        }




def list_form_submissions(form_id: int) -> list[dict]:
    """List all submissions for a given form (admin view)."""
    with get_cursor() as cur:
        cur.execute(
            """
            SELECT s.id, s.school_id, s.status, s.submitted_at, s.created_at,
                   s.is_late, s.late_days, COALESCE(s.late_minutes, s.late_days * 1440, 0) AS late_minutes,
                   s.repeat_period_key, s.repeat_period_label,
                   f.deadline_at AS form_deadline_at,
                   f.repeat_policy AS form_repeat_policy,
                   f.repeat_deadline_time AS form_repeat_deadline_time,
                   f.repeat_deadline_day AS form_repeat_deadline_day,
                   f.very_late_after_minutes,
                   sc.name AS school_name, sc.npsn, sc.jenjang,
                   u.full_name AS submitted_by_name
            FROM laporan_submissions s
            JOIN laporan_forms f ON f.id = s.form_id
            JOIN portal_schools sc ON sc.id = s.school_id
            LEFT JOIN dashboard_users u ON u.id = s.submitted_by
            WHERE s.form_id = %s AND s.status IN ('submitted', 'no_submission')
            ORDER BY s.repeat_period_key DESC NULLS LAST, s.submitted_at DESC NULLS LAST, s.created_at DESC, sc.name ASC
            """,
            (form_id,),
        )
        return [dict(r) for r in cur.fetchall()]


# ─────────────────────────────────────────────
# Export Excel
# ─────────────────────────────────────────────

def _answer_export_value(field: dict, answer: Optional[dict]) -> str:
    if not answer:
        return ""
    ftype = field["field_type"]
    if ftype == "file":
        files = answer.get("files") or []
        return ", ".join(fi.get("original_name", "") for fi in files)
    if ftype == "checkbox" and answer.get("answer_json"):
        val = answer["answer_json"]
        return ", ".join(val) if isinstance(val, list) else str(val)
    return answer.get("answer_text", "") or ""


def export_form_xlsx(form_id: int, filter_period: Optional[str] = None) -> tuple[str, bytes]:
    """
    Export all submitted answers for a form as a styled Excel workbook.
    Returns (filename, xlsx_bytes).
    """
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    form = get_form(form_id)
    if not form:
        return "laporan.xlsx", b""

    fields = [f for f in get_form_fields(form_id) if f.get("field_type") not in {"header", "info"}]
    submissions = list_form_submissions(form_id)
    if filter_period and filter_period != "all":
        submissions = [s for s in submissions if s.get("repeat_period_key") == filter_period]
    header = ["No", "Sekolah", "NPSN", "Jenjang", "Disubmit Oleh", "Periode", "Waktu Submit"]
    for f in fields:
        header.append(f["label"])

    wb = Workbook()
    ws = wb.active
    ws.title = "Jawaban"
    ws.append(header)

    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for idx, sub in enumerate(submissions, 1):
        sub_detail = get_submission_with_answers(sub["id"])
        answers_map = {a["field_id"]: a for a in (sub_detail.get("answers") or [])} if sub_detail else {}
        row = [
            idx,
            sub.get("school_name", ""),
            sub.get("npsn", ""),
            sub.get("jenjang", ""),
            sub.get("submitted_by_name", "") or "",
            sub.get("repeat_period_label", "") or "",
            sub.get("submitted_at").strftime("%d/%m/%Y %H:%M") if sub.get("submitted_at") else ("Tidak Mengumpulkan" if sub.get("status") == "no_submission" else ""),
        ]
        for f in fields:
            row.append(_answer_export_value(f, answers_map.get(f["id"])))
        ws.append(row)

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    for col_idx, column_cells in enumerate(ws.columns, 1):
        max_len = 0
        for cell in column_cells:
            value = str(cell.value or "")
            max_len = max(max_len, min(len(value), 60))
        ws.column_dimensions[get_column_letter(col_idx)].width = max(12, min(max_len + 2, 45))

    submitted_subs = [s for s in submissions if s.get("status") == "submitted"]
    no_subs = [s for s in submissions if s.get("status") == "no_submission"]

    summary = wb.create_sheet("Ringkasan")
    summary.append(["Form", form.get("title", "")])
    summary.append(["Total sasaran laporan", len(submissions)])
    summary.append(["Terkirim", len(submitted_subs)])
    summary.append(["Tidak mengumpulkan", len(no_subs)])
    summary.append(["Sekolah berbeda (yang mengirim)", len({s.get("school_id") for s in submitted_subs if s.get("school_id")} )])
    summary.append(["Tepat waktu (dari terkirim)", sum(1 for s in submitted_subs if not s.get("is_late"))])
    summary.append(["Terlambat (dari terkirim)", sum(1 for s in submitted_subs if s.get("is_late"))])
    summary.append([])
    summary.append(["No", "Pertanyaan", "Tipe", "Terisi", "Kosong"])
    for cell in summary[9]:
        cell.fill = header_fill
        cell.font = header_font
    answer_counts = {f["id"]: 0 for f in fields}
    for sub in submissions:
        sub_detail = get_submission_with_answers(sub["id"])
        if sub_detail:
            for answer in sub_detail.get("answers") or []:
                if answer.get("answer_text") or answer.get("answer_json") or answer.get("files"):
                    answer_counts[answer["field_id"]] = answer_counts.get(answer["field_id"], 0) + 1
    for idx, field in enumerate(fields, 1):
        summary.append([
            idx,
            field.get("label", ""),
            field.get("field_type", ""),
            answer_counts.get(field["id"], 0),
            max(len(submissions) - answer_counts.get(field["id"], 0), 0),
        ])
    summary.column_dimensions["A"].width = 8
    summary.column_dimensions["B"].width = 45
    summary.column_dimensions["C"].width = 16
    summary.column_dimensions["D"].width = 12
    summary.column_dimensions["E"].width = 12

    safe_title = "".join(c if c.isalnum() or c in " _-" else "_" for c in (form["title"] or "laporan"))
    filename = f"laporan_{safe_title[:40]}.xlsx"
    output = BytesIO()
    wb.save(output)
    return filename, output.getvalue()


def export_form_csv(form_id: int) -> tuple[str, bytes]:
    """Backward-compatible alias; now returns a native Excel workbook."""
    return export_form_xlsx(form_id)


def export_no_submissions_xlsx(form_id: int, filter_period: Optional[str] = None) -> tuple[str, bytes]:
    """
    Export the list of schools that did not submit for a form as an Excel workbook.
    Returns (filename, xlsx_bytes).
    """
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter
    from io import BytesIO

    form = get_form(form_id)
    if not form:
        return "tidak_mengumpulkan.xlsx", b""

    # Fetch submissions with status 'no_submission'
    submissions = list_form_submissions(form_id)
    if filter_period and filter_period != "all":
        submissions = [s for s in submissions if s.get("repeat_period_key") == filter_period]
    no_subs = [s for s in submissions if s.get("status") == "no_submission"]

    header = ["No", "Sekolah", "NPSN", "Jenjang", "Periode", "Status"]

    wb = Workbook()
    ws = wb.active
    ws.title = "Tidak Mengumpulkan"
    ws.append(header)

    header_fill = PatternFill("solid", fgColor="C00000")  # Dark red for "Tidak Mengumpulkan"
    header_font = Font(color="FFFFFF", bold=True)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for idx, sub in enumerate(no_subs, 1):
        row = [
            idx,
            sub.get("school_name", ""),
            sub.get("npsn", ""),
            sub.get("jenjang", ""),
            sub.get("repeat_period_label", "Sekali isi") or "Sekali isi",
            "Tidak Mengumpulkan",
        ]
        ws.append(row)

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    for col_idx, column_cells in enumerate(ws.columns, 1):
        max_len = 0
        for cell in column_cells:
            value = str(cell.value or "")
            max_len = max(max_len, min(len(value), 60))
        ws.column_dimensions[get_column_letter(col_idx)].width = max(12, min(max_len + 2, 45))

    safe_title = "".join(c if c.isalnum() or c in " _-" else "_" for c in (form["title"] or "laporan"))
    filename = f"tidak_mengumpulkan_{safe_title[:40]}.xlsx"
    output = BytesIO()
    wb.save(output)
    return filename, output.getvalue()


# ─────────────────────────────────────────────
# School list helper (for admin target selection)
# ─────────────────────────────────────────────

def list_all_schools_simple() -> list[dict]:
    with get_cursor() as cur:
        cur.execute(
            "SELECT id, name, npsn, jenjang FROM portal_schools WHERE active=TRUE ORDER BY jenjang, name"
        )
        return [dict(r) for r in cur.fetchall()]


# ─────────────────────────────────────────────
# KPI (Admin)
# ─────────────────────────────────────────────

def fetch_laporan_kpi_schools() -> list[dict]:
    """
    Fetch KPIs for reporting tardiness grouped by school.
    Rules:
      - is_late = false -> Tepat Waktu
      - is_late = true AND late_minutes <= form threshold -> Terlambat
      - is_late = true AND late_minutes > form threshold -> Sangat Terlambat
    """
    with get_cursor() as cur:
        cur.execute(
            """
            SELECT sc.id AS school_id, sc.name AS school_name, sc.npsn, sc.jenjang,
                   COUNT(s.id) AS total_submissions,
                   SUM(CASE WHEN s.is_late = FALSE THEN 1 ELSE 0 END) AS on_time_count,
                   SUM(CASE
                       WHEN s.is_late = TRUE
                        AND COALESCE(s.late_minutes, s.late_days * 1440, 0)
                            <= COALESCE(f.very_late_after_minutes, 180)
                       THEN 1 ELSE 0
                   END) AS late_count,
                   SUM(CASE
                       WHEN s.is_late = TRUE
                        AND COALESCE(s.late_minutes, s.late_days * 1440, 0)
                            > COALESCE(f.very_late_after_minutes, 180)
                       THEN 1 ELSE 0
                   END) AS very_late_count
            FROM portal_schools sc
            LEFT JOIN laporan_submissions s ON s.school_id = sc.id AND s.status = 'submitted'
            LEFT JOIN laporan_forms f ON f.id = s.form_id
            WHERE sc.active = TRUE
            GROUP BY sc.id, sc.name, sc.npsn, sc.jenjang
            ORDER BY total_submissions DESC, sc.name
            """
        )
        return [dict(r) for r in cur.fetchall()]
