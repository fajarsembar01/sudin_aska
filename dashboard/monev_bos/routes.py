from flask import Blueprint, current_app, render_template, request, jsonify, flash, redirect, url_for, session, abort, send_file
from dashboard.auth import role_required, current_user
from dashboard.queries import record_admin_action
from . import monev_bos_bp, queries
from .bop_claims import get_school_bop_claim, is_bop_claim_period, recommend_expense_type
from .external_photos import access_token_matches, validate_external_identity, validate_external_nip
import psycopg2
import base64
import shutil
import uuid
import time

MAX_ACTIVITY_FIELD_PHOTOS = 3
LOCKED_SCHOOL_REPORT_STATUSES = {"completed", "completed_with_notes"}
REPORT_AUDIT_STATUSES = {
    "submitted",
    "in_review",
    "needs_revision",
    "completed_with_notes",
    "completed",
}


def _resolve_school_report_vendor(vendor_id_raw):
    """Resolve a globally selectable vendor while its verification is pending."""
    if not vendor_id_raw:
        return None, None, None
    if not str(vendor_id_raw).isdigit():
        return None, None, "Vendor / narasumber yang dipilih tidak valid."

    vendor_id = int(vendor_id_raw)
    vendor = queries.get_vendor_by_id(vendor_id)
    if not vendor:
        return None, None, "Vendor / narasumber tidak ditemukan."
    if vendor.get("status") not in {"pending", "verified"}:
        return None, None, "Vendor / narasumber yang ditolak tidak dapat dimasukkan ke laporan."

    display_name = queries.get_vendor_display_name(vendor)
    return vendor_id, display_name, None


def _resolve_school_report_vendors(vendor_id_values):
    vendor_ids = []
    vendor_names = []
    for raw_vendor_id in vendor_id_values:
        vendor_id, vendor_name, error = _resolve_school_report_vendor(raw_vendor_id)
        if error:
            return [], [], error
        if vendor_id is not None and vendor_id not in vendor_ids:
            vendor_ids.append(vendor_id)
            vendor_names.append(vendor_name)
    return vendor_ids, vendor_names, None


def _activity_account_code_data(form) -> dict:
    """Normalize a multi-select account code submission while preserving order."""
    selected_codes = []
    for raw_code in form.getlist("account_code"):
        code = str(raw_code or "").strip()
        if code and code not in selected_codes:
            selected_codes.append(code)
    return {
        "account_code": ", ".join(selected_codes) or None,
        "primary_account_code": selected_codes[0] if selected_codes else None,
    }


def _school_bku_number(raw_value):
    """Validate the three-digit BKU number used by school activity forms."""
    value = str(raw_value or "").strip()
    if len(value) != 3 or not value.isdigit():
        return None, "Nomor BKU harus terdiri dari 3 angka, misalnya 001."
    return value, None


def _activity_has_vendor(activity):
    """Return True when an activity has at least one vendor reference."""
    return bool(
        activity.get("vendors")
        or activity.get("vendor_id")
        or (activity.get("vendor_name") or "").strip()
    )


def _activity_vendor_is_unverified(activity):
    """Return True when the required vendor is missing or not fully verified."""
    if activity.get("vendors"):
        return any(vendor.get("status") != "verified" for vendor in activity["vendors"])
    if not _activity_has_vendor(activity):
        return True
    return activity.get("vendor_status") != "verified"


def _requested_activity_photos():
    files = [
        file
        for file in request.files.getlist("doc_field_photo") + request.files.getlist("field_photo")
        if file and file.filename
    ]
    camera_photos = [value for value in request.form.getlist("field_photo_data") if value]
    raw_story_ids = request.form.getlist("story_post_ids") or request.form.getlist("story_post_id")
    story_ids = list(dict.fromkeys(
        int(value) for value in raw_story_ids if value and value.isdigit()
    ))
    return files, camera_photos, story_ids


def _school_activity_mutation_error(report, activity):
    """Return the reason a school cannot edit/delete one of its activities."""
    if report.get("status") in LOCKED_SCHOOL_REPORT_STATUSES:
        return "Laporan yang sudah selesai tidak bisa diubah."
    if not activity or activity.get("report_id") != report.get("id"):
        return "Kegiatan tidak ditemukan pada laporan sekolah ini."
    if activity.get("status") == "valid":
        return "Kegiatan berstatus Sesuai tidak bisa diedit atau dihapus."
    return None


def _record_monev_admin_action(
    action: str,
    target_type: str,
    *,
    target_id: int = None,
    target_name: str = None,
    metadata: dict = None,
    allow_staff: bool = False,
) -> None:
    """Record successful Monev mutations without breaking the main action."""
    user = current_user() or {}
    if user.get("role") != "admin" and not (allow_staff and user.get("role") == "staff"):
        return
    try:
        record_admin_action(
            user_id=user.get("id"),
            feature_key="monev_bos",
            action=action,
            target_type=target_type,
            target_id=target_id,
            target_name=target_name,
            metadata=metadata,
        )
    except Exception:
        current_app.logger.exception(
            "Failed to record Monev BOS admin action %s on %s #%s",
            action,
            target_type,
            target_id,
        )

@monev_bos_bp.route("/")
@role_required("admin", "staff", "sekolah")
def index():
    user = current_user()
    if user.get("role") == "admin":
        return redirect(url_for("monev_bos.admin_dashboard"))
    elif user.get("role") == "sekolah":
        return redirect(url_for("monev_bos.sekolah_dashboard"))
    else:
        return redirect(url_for("monev_bos.staff_dashboard"))

@monev_bos_bp.before_request
def enforce_headmaster_info_for_sekolah():
    user = current_user()
    if user and user.get("role") == "sekolah":
        try:
            from dashboard.portal.routes import _fetch_user_school, _normalize_metadata
            school = _fetch_user_school(user["id"])
            if school:
                meta = _normalize_metadata(school.get("metadata"))
                h_name = (meta.get("headmaster_name") or "").strip()
                h_nip = (meta.get("headmaster_nip") or "").strip()
                if not h_name or not h_nip or h_name == "-" or h_nip == "-":
                    flash("Mohon lengkapi Data Kepala Sekolah (Nama & NIP) pada Profil Sekolah terlebih dahulu untuk mengakses Monev BOS/BOP.", "warning")
                    return redirect(url_for("portal.sekolah_profile"))
        except Exception:
            pass

@monev_bos_bp.context_processor
def inject_monev_bos_context() -> dict:
    user = current_user()
    user_app_notifications = {"unread_count": 0, "total_count": 0}
    if user and user.get("role") in ["staff", "coordinator", "sekolah"]:
        try:
            from dashboard.daftar_tamu.queries import fetch_user_notification_summary, USER_APP_NOTIFICATION_CATEGORIES
            user_app_notifications = fetch_user_notification_summary(
                user_id=int(user["id"]),
                categories=list(USER_APP_NOTIFICATION_CATEGORIES)
            )
        except Exception:
            pass

    admin_pending = {
        "pending_users": 0,
        "pending_assignment_requests": 0,
        "pending_team_member_requests": 0,
        "pending_reopen_requests": 0,
        "pending_guestbook": 0,
        "pending_monev_edit_requests": 0,
        "pending_call_center": 0,
        "total": 0,
    }
    if user and user.get("role") == "admin":
        try:
            from dashboard.portal.queries import fetch_admin_pending_summary
            admin_pending = fetch_admin_pending_summary()
        except Exception:
            pass

    submitted_reports_count = 0
    if user and user.get("role") in ["admin", "staff"]:
        try:
            submitted_reports_count = queries.get_submitted_reports_count()
        except Exception:
            pass

    active_story_groups = []
    if user and user.get("role") in ["admin", "sekolah"]:
        try:
            school_user_id = int(user["id"]) if user.get("role") == "sekolah" else None
            active_story_groups = queries.list_active_story_groups(school_user_id=school_user_id)
        except Exception:
            current_app.logger.exception("Failed to load active Monev school stories")

    return {
        "app_title": "MONEV BOS/BOP",
        "admin_pending": admin_pending,
        "user_app_notifications": user_app_notifications,
        "submitted_reports_count": submitted_reports_count,
        "active_story_groups": active_story_groups,
        "undo_window_seconds": 7,
    }

from datetime import date, datetime


ADMIN_ANALYTICS_METRICS = {
    "all": ("Semua Sekolah", "Seluruh sekolah yang ditugaskan atau memiliki laporan."),
    "assigned": ("Sekolah Ditugaskan", "Daftar sekolah yang sudah ditempatkan ke tim verifikator."),
    "reports": ("Laporan Masuk", "Sekolah yang sudah membuat laporan, termasuk laporan draft."),
    "review": ("Perlu Diverifikasi", "Laporan yang sudah disubmit atau sedang diverifikasi."),
    "completed": ("Verifikasi Selesai", "Laporan yang verifikasinya sudah diselesaikan."),
    "unreported": ("Belum Melapor", "Sekolah penugasan yang belum membuat laporan."),
    "revision": ("Perlu Revisi", "Laporan yang dikembalikan kepada sekolah untuk diperbaiki."),
    "draft": ("Laporan Draft", "Laporan yang masih disusun dan belum dikirim untuk verifikasi."),
}


def _selected_admin_period(periods, requested_period_id=None):
    if requested_period_id:
        selected = next(
            (period for period in periods if int(period["id"]) == int(requested_period_id)),
            None,
        )
        if selected:
            return selected
    default_period = queries.get_active_period()
    if default_period:
        return next(
            (period for period in periods if int(period["id"]) == int(default_period["id"])),
            default_period,
        )
    if periods:
        return max(periods, key=lambda period: (int(period["year"]), int(period["tw"])))
    return None


def _usable_activity_photos(photo_rows):
    """Exclude stale database records whose local image file is missing or empty."""
    dashboard_root = os.path.realpath(os.path.join(monev_bos_bp.root_path, ".."))
    static_root = os.path.realpath(os.path.join(dashboard_root, "static"))
    usable = []
    for photo in photo_rows:
        relative_path = str(photo.get("file_path") or "").lstrip("/")
        if not relative_path.startswith("static/uploads/monev_bos/"):
            continue
        absolute_path = os.path.realpath(os.path.join(dashboard_root, relative_path))
        try:
            if os.path.commonpath([static_root, absolute_path]) != static_root:
                continue
            if not os.path.isfile(absolute_path) or os.path.getsize(absolute_path) <= 0:
                continue
        except (OSError, ValueError):
            continue
        photo["photo_url"] = "/" + relative_path
        usable.append(photo)
    return usable

def _attach_team_progress_metrics(team):
    """Attach reporting and activity-status rates used by admin team progress bars."""
    assigned = int(team.get("assigned_schools") or 0)
    reports = int(team.get("total_reports") or 0)
    total = int(team.get("total_activities") or 0)
    pending_count = int(team.get("pending_activities") or 0)
    in_review_count = int(team.get("in_review_activities") or 0)
    status_counts = {
        "valid": int(team.get("valid_activities") or 0),
        "invalid": int(team.get("invalid_activities") or 0),
        "pending": pending_count + in_review_count,
    }
    status_counts["other"] = max(total - sum(status_counts.values()), 0)

    team["audited_activities"] = status_counts["valid"] + status_counts["invalid"]
    team["reporting_rate"] = min(round(reports / assigned * 100), 100) if assigned else 0
    team["verification_rate"] = min(round(team["audited_activities"] / total * 100), 100) if total else 0
    active_statuses = [status for status, count in status_counts.items() if count]
    allocated_rate = 0.0
    for status, count in status_counts.items():
        team[f"{status}_activities"] = count
        if not total or not count:
            rate = 0
        elif status == active_statuses[-1]:
            rate = round(100.0 - allocated_rate, 1)
        else:
            rate = round(count / total * 100, 1)
            allocated_rate += rate
        team[f"{status}_rate"] = rate
    return team


@monev_bos_bp.route("/admin")
@role_required("admin")
def admin_dashboard():
    periods = queries.list_periods()
    active_period = _selected_admin_period(periods, request.args.get("period_id", type=int))
    photo_order = (request.args.get("photo_order") or "newest").lower()
    if photo_order not in {"newest", "random"}:
        photo_order = "newest"
    overview = {}
    recent_reports = []
    team_performance = []
    activity_photos = []
    days_remaining = None
    if active_period:
        overview = queries.get_admin_dashboard_overview(active_period["id"])
        recent_reports = queries.list_recent_period_reports(active_period["id"])
        team_performance = queries.list_admin_team_performance(active_period["id"])
        all_activity_photos = _usable_activity_photos(
            queries.list_admin_activity_photos(None, limit=500, order=photo_order)
        )
        activity_photos = all_activity_photos[:12]
        activity_photo_total = len(all_activity_photos)
        for team in team_performance:
            _attach_team_progress_metrics(team)
        end_date = active_period.get("end_date")
        if isinstance(end_date, datetime):
            end_date = end_date.date()
        if isinstance(end_date, date):
            days_remaining = (end_date - date.today()).days

    assigned_schools = int(overview.get("assigned_schools") or 0)
    total_reports = int(overview.get("total_reports") or 0)
    total_activities = int(overview.get("total_activities") or 0)
    audited_activities = int(overview.get("valid_activities") or 0) + int(overview.get("invalid_activities") or 0)
    dashboard_metrics = {
        "unreported_schools": max(assigned_schools - total_reports, 0),
        "reporting_rate": min(round(total_reports / assigned_schools * 100), 100) if assigned_schools else 0,
        "verification_rate": min(round(audited_activities / total_activities * 100), 100) if total_activities else 0,
        "audited_activities": audited_activities,
        "review_queue": int(overview.get("submitted_reports") or 0) + int(overview.get("in_review_reports") or 0),
        "finished_reports": int(overview.get("completed_reports") or 0) + int(overview.get("completed_with_notes_reports") or 0),
    }
    return render_template(
        "monev_bos/admin/dashboard.html",
        active_period=active_period,
        overview=overview,
        dashboard_metrics=dashboard_metrics,
        recent_reports=recent_reports,
        days_remaining=days_remaining,
        periods=periods,
        team_performance=team_performance,
        activity_photos=activity_photos,
        photo_order=photo_order,
        activity_photo_total=activity_photo_total if active_period else 0,
    )


@monev_bos_bp.route("/admin/analytics/<metric>")
@role_required("admin")
def admin_analytics(metric):
    if metric not in ADMIN_ANALYTICS_METRICS:
        metric = "all"

    periods = queries.list_periods()
    selected_period = _selected_admin_period(periods, request.args.get("period_id", type=int))
    if not selected_period:
        flash("Belum ada periode MONEV yang tersedia.", "warning")
        return redirect(url_for("monev_bos.admin_dashboard"))

    overview = queries.get_admin_dashboard_overview(selected_period["id"])
    all_school_rows = queries.list_admin_period_school_analytics(selected_period["id"])
    team_performance = queries.list_admin_team_performance(selected_period["id"])
    selected_team_id = request.args.get("team_id", type=int)
    selected_team = next(
        (team for team in team_performance if int(team["team_id"]) == selected_team_id),
        None,
    ) if selected_team_id else None
    if selected_team_id and not selected_team:
        selected_team_id = None

    scoped_school_rows = [
        row for row in all_school_rows
        if selected_team_id is None or row.get("team_id") == selected_team_id
    ]

    def matches_metric(row):
        status = row.get("report_status")
        if metric == "assigned":
            return bool(row.get("is_assigned"))
        if metric == "reports":
            return bool(row.get("report_id"))
        if metric == "review":
            return status in {"submitted", "in_review"}
        if metric == "completed":
            return status in {"completed", "completed_with_notes"}
        if metric == "unreported":
            return bool(row.get("is_assigned")) and not row.get("report_id")
        if metric == "revision":
            return status == "needs_revision"
        if metric == "draft":
            return status == "draft"
        return True

    school_rows = [row for row in scoped_school_rows if matches_metric(row)]
    search_query = (request.args.get("q") or "").strip()
    if search_query:
        search_normalized = search_query.casefold()
        school_rows = [
            row for row in school_rows
            if search_normalized in " ".join([
                str(row.get("school_name") or ""),
                str(row.get("npsn") or ""),
                str(row.get("team_name") or ""),
            ]).casefold()
        ]

    for row in all_school_rows:
        audited = int(row.get("valid_activities") or 0) + int(row.get("invalid_activities") or 0)
        total = int(row.get("total_activities") or 0)
        row["audited_activities"] = audited
        row["verification_rate"] = min(round(audited / total * 100), 100) if total else 0

    for team in team_performance:
        _attach_team_progress_metrics(team)

    metric_counts = {
        key: sum(1 for row in scoped_school_rows if (
            (key == "all")
            or (key == "assigned" and row.get("is_assigned"))
            or (key == "reports" and row.get("report_id"))
            or (key == "review" and row.get("report_status") in {"submitted", "in_review"})
            or (key == "completed" and row.get("report_status") in {"completed", "completed_with_notes"})
            or (key == "unreported" and row.get("is_assigned") and not row.get("report_id"))
            or (key == "revision" and row.get("report_status") == "needs_revision")
            or (key == "draft" and row.get("report_status") == "draft")
        ))
        for key in ADMIN_ANALYTICS_METRICS
    }

    bos_receipts = float(overview.get("bosp_receipts") or 0)
    bop_receipts = float(overview.get("bop_receipts") or 0)
    bos_realized = float(overview.get("bos_realized") or 0)
    bop_realized = float(overview.get("bop_realized") or 0)
    financial_analytics = {
        "bos_absorption": round(bos_realized / bos_receipts * 100, 1) if bos_receipts else 0,
        "bop_absorption": round(bop_realized / bop_receipts * 100, 1) if bop_receipts else 0,
        "total_absorption": round((bos_realized + bop_realized) / (bos_receipts + bop_receipts) * 100, 1) if (bos_receipts + bop_receipts) else 0,
    }

    return render_template(
        "monev_bos/admin/analytics.html",
        active_period=selected_period,
        periods=periods,
        metric=metric,
        metric_config=ADMIN_ANALYTICS_METRICS[metric],
        metric_configs=ADMIN_ANALYTICS_METRICS,
        metric_counts=metric_counts,
        school_rows=school_rows,
        overview=overview,
        team_performance=team_performance,
        financial_analytics=financial_analytics,
        search_query=search_query,
        selected_team_id=selected_team_id,
        selected_team=selected_team,
    )


@monev_bos_bp.route("/admin/activity-gallery")
@role_required("admin")
def admin_activity_gallery():
    periods = queries.list_periods()
    raw_period_id = (request.args.get("period_id") or "all").lower()
    all_periods = raw_period_id == "all"
    requested_period_id = int(raw_period_id) if raw_period_id.isdigit() else None
    selected_period = _selected_admin_period(periods, requested_period_id)
    if not selected_period:
        flash("Belum ada periode MONEV yang tersedia.", "warning")
        return redirect(url_for("monev_bos.admin_dashboard"))

    team_performance = queries.list_admin_team_performance(selected_period["id"])
    requested_team_id = request.args.get("team_id", type=int)
    team_id = requested_team_id if any(
        int(team["team_id"]) == requested_team_id for team in team_performance
    ) else None
    fund_source = (request.args.get("fund_source") or "").upper()
    if fund_source not in {"BOS", "BOP"}:
        fund_source = ""
    photo_status = (request.args.get("photo_status") or "").lower()
    if photo_status not in {"valid", "invalid"}:
        photo_status = ""
    photo_order = (request.args.get("order") or "newest").lower()
    if photo_order not in {"newest", "oldest", "random"}:
        photo_order = "newest"
    search_query = (request.args.get("q") or "").strip()

    photos = _usable_activity_photos(queries.list_admin_activity_photos(
        None if all_periods else selected_period["id"],
        limit=500,
        team_id=team_id,
        fund_source=fund_source or None,
        photo_status=photo_status or None,
        search_query=search_query or None,
        order=photo_order,
    ))
    return render_template(
        "monev_bos/admin/activity_gallery.html",
        active_period=selected_period,
        periods=periods,
        team_performance=team_performance,
        team_id=team_id,
        fund_source=fund_source,
        photo_status=photo_status,
        photo_order=photo_order,
        search_query=search_query,
        photos=photos,
        all_periods=all_periods,
    )

@monev_bos_bp.route("/admin/periods", methods=["GET", "POST"])
@role_required("admin")
def admin_periods():
    current_year = datetime.now().year
    
    if request.method == "POST":
        action = request.form.get("action")
        if action == "generate_year":
            year = int(request.form.get("year"))
            queries.ensure_periods_for_year(year)
            _record_monev_admin_action(
                "GENERATE",
                "MONEV_PERIOD_YEAR",
                target_name=f"Periode {year}",
                metadata={"year": year},
            )
            flash(f"Periode TW 1-4 Tahun {year} berhasil disiapkan.", "success")
        elif action == "set_active":
            period_id = int(request.form.get("period_id"))
            queries.set_active_period(period_id)
            _record_monev_admin_action("ACTIVATE", "MONEV_PERIOD", target_id=period_id)
            flash("Periode berhasil diaktifkan.", "success")
        elif action == "deactivate":
            period_id = int(request.form.get("period_id"))
            queries.deactivate_period(period_id)
            _record_monev_admin_action("DEACTIVATE", "MONEV_PERIOD", target_id=period_id)
            flash("Periode berhasil dinonaktifkan.", "success")
        elif action == "set_deadline":
            period_id = int(request.form.get("period_id"))
            deadline = datetime.strptime(request.form.get("deadline"), "%Y-%m-%d").date()
            queries.update_period_deadline(period_id, deadline)
            _record_monev_admin_action(
                "UPDATE_DEADLINE",
                "MONEV_PERIOD",
                target_id=period_id,
                metadata={"deadline": deadline.isoformat()},
            )
            flash("Deadline berhasil diperbarui.", "success")
        return redirect(url_for("monev_bos.admin_periods"))
    
    # Auto-generate current year if empty
    available_years = queries.get_available_years()
    if not available_years:
        queries.ensure_periods_for_year(current_year)
        available_years = [current_year]
    
    selected_year = request.args.get("year", available_years[0] if available_years else current_year, type=int)
    
    # Always ensure all 4 TW exist for selected year (fixes missing or wrong dates)
    queries.ensure_periods_for_year(selected_year)
    periods = queries.list_periods_by_year(selected_year)
    queries.attach_period_admin_input_names(periods)
    available_years = queries.get_available_years()
    activity_logs = queries.list_admin_action_history(["MONEV_PERIOD_YEAR", "MONEV_PERIOD"])
    
    return render_template("monev_bos/admin/periods.html", 
                           periods=periods, 
                           available_years=available_years,
                           selected_year=selected_year,
                           current_year=current_year,
                           activity_logs=activity_logs)

@monev_bos_bp.route("/admin/checklists", methods=["GET", "POST"])
@role_required("admin")
def admin_checklists():
    if request.method == "POST":
        action = request.form.get("action")
        expense_type_ids = [int(x) for x in request.form.getlist("expense_type_ids") if x.isdigit()]
        if action == "create":
            name = request.form.get("name")
            description = request.form.get("description")
            sort_order = int(request.form.get("sort_order", 0))
            checklist_id = queries.create_checklist(name, description, sort_order, expense_type_ids=expense_type_ids)
            _record_monev_admin_action(
                "CREATE",
                "MONEV_CHECKLIST",
                target_id=checklist_id,
                target_name=name,
                metadata={"expense_type_ids": expense_type_ids},
            )
            flash("Checklist berhasil ditambahkan", "success")
        elif action == "update":
            checklist_id = int(request.form.get("checklist_id"))
            name = request.form.get("name")
            description = request.form.get("description")
            sort_order = int(request.form.get("sort_order", 0))
            is_active = request.form.get("is_active") == "on"
            queries.update_checklist(checklist_id, name, description, sort_order, is_active, expense_type_ids=expense_type_ids)
            _record_monev_admin_action(
                "UPDATE",
                "MONEV_CHECKLIST",
                target_id=checklist_id,
                target_name=name,
                metadata={"is_active": is_active, "expense_type_ids": expense_type_ids},
            )
            flash("Checklist berhasil diupdate", "success")
        elif action == "delete":
            checklist_id = int(request.form.get("checklist_id"))
            queries.delete_checklist(checklist_id)
            _record_monev_admin_action("DELETE", "MONEV_CHECKLIST", target_id=checklist_id)
            flash("Checklist berhasil dihapus", "success")
        return redirect(url_for("monev_bos.admin_checklists"))

    checklists = queries.list_checklists(include_inactive=True)
    queries.attach_admin_input_names(checklists, "MONEV_CHECKLIST")
    expense_types = queries.list_expense_types(include_inactive=False)
    activity_logs = queries.list_admin_action_history(["MONEV_CHECKLIST"])
    return render_template(
        "monev_bos/admin/checklists.html",
        checklists=checklists,
        expense_types=expense_types,
        activity_logs=activity_logs,
    )


@monev_bos_bp.route("/admin/master-activities", methods=["GET", "POST"])
@role_required("admin")
def admin_master_activities():
    if request.method == "POST":
        action = request.form.get("action")
        if action == "create":
            name = request.form.get("name")
            code_prefix = request.form.get("code_prefix")
            fund_source = request.form.get("fund_source", "ALL")
            if name:
                if queries.check_master_activity_exists(name):
                    flash(f"Gagal! Nama Kegiatan '{name}' sudah ada. Tidak boleh ada duplikasi.", "danger")
                else:
                    try:
                        master_id = queries.create_master_activity(name, code_prefix, fund_source)
                        _record_monev_admin_action(
                            "CREATE",
                            "MONEV_MASTER_ACTIVITY",
                            target_id=master_id,
                            target_name=name,
                            metadata={"code_prefix": code_prefix, "fund_source": fund_source},
                        )
                        flash("Master Nama Kegiatan berhasil ditambahkan", "success")
                    except psycopg2.errors.UniqueViolation:
                        flash(f"Gagal! Nama Kegiatan '{name}' sudah ada. Tidak boleh ada duplikasi.", "danger")
            else:
                flash("Nama kegiatan wajib diisi.", "warning")
        elif action == "update":
            master_id = int(request.form.get("master_id"))
            name = request.form.get("name")
            code_prefix = request.form.get("code_prefix")
            fund_source = request.form.get("fund_source", "ALL")
            is_active = request.form.get("is_active") == "on"
            if name:
                if queries.check_master_activity_exists(name, exclude_id=master_id):
                    flash(f"Gagal! Nama Kegiatan '{name}' sudah terpakai oleh data lain.", "danger")
                else:
                    try:
                        queries.update_master_activity(master_id, name, code_prefix, fund_source, is_active)
                        _record_monev_admin_action(
                            "UPDATE",
                            "MONEV_MASTER_ACTIVITY",
                            target_id=master_id,
                            target_name=name,
                            metadata={"code_prefix": code_prefix, "fund_source": fund_source, "is_active": is_active},
                        )
                        flash("Master Nama Kegiatan berhasil diperbarui", "success")
                    except psycopg2.errors.UniqueViolation:
                        flash(f"Gagal! Nama Kegiatan '{name}' sudah terpakai oleh data lain.", "danger")
            else:
                flash("Nama kegiatan wajib diisi.", "warning")
        elif action == "delete":
            master_id = int(request.form.get("master_id"))
            queries.delete_master_activity(master_id)
            _record_monev_admin_action("DELETE", "MONEV_MASTER_ACTIVITY", target_id=master_id)
            flash("Master Nama Kegiatan berhasil dihapus", "success")
        return redirect(url_for("monev_bos.admin_master_activities"))

    search_query = request.args.get("q", "").strip()
    master_activities = queries.list_master_activities(
        include_inactive=True,
        search_query=search_query,
    )
    queries.attach_admin_input_names(master_activities, "MONEV_MASTER_ACTIVITY")
    activity_logs = queries.list_admin_action_history(["MONEV_MASTER_ACTIVITY"])
    return render_template(
        "monev_bos/admin/master_activities.html",
        master_activities=master_activities,
        activity_logs=activity_logs,
        search_query=search_query,
    )


@monev_bos_bp.route("/admin/expense-types", methods=["GET", "POST"])
@role_required("admin")
def admin_expense_types():
    if request.method == "POST":
        action = request.form.get("action")
        if action == "create":
            name = (request.form.get("name") or "").strip()
            code = (request.form.get("code") or "").strip()
            description = (request.form.get("description") or "").strip()
            sort_order = int(request.form.get("sort_order", 0) or 0)
            if name:
                expense_type_id = queries.create_expense_type(name, code, description, sort_order)
                _record_monev_admin_action(
                    "CREATE",
                    "MONEV_EXPENSE_TYPE",
                    target_id=expense_type_id,
                    target_name=name,
                    metadata={"code": code, "sort_order": sort_order},
                )
                flash("Jenis Belanja berhasil ditambahkan.", "success")
            else:
                flash("Nama Jenis Belanja wajib diisi.", "warning")
        elif action == "update":
            expense_type_id = int(request.form.get("expense_type_id"))
            name = (request.form.get("name") or "").strip()
            code = (request.form.get("code") or "").strip()
            description = (request.form.get("description") or "").strip()
            sort_order = int(request.form.get("sort_order", 0) or 0)
            is_active = request.form.get("is_active") == "on"
            if name and expense_type_id:
                queries.update_expense_type(expense_type_id, name, code, description, sort_order, is_active)
                _record_monev_admin_action(
                    "UPDATE",
                    "MONEV_EXPENSE_TYPE",
                    target_id=expense_type_id,
                    target_name=name,
                    metadata={"code": code, "sort_order": sort_order, "is_active": is_active},
                )
                flash("Jenis Belanja berhasil diperbarui.", "success")
        elif action == "delete":
            expense_type_id = int(request.form.get("expense_type_id"))
            if expense_type_id:
                queries.delete_expense_type(expense_type_id)
                _record_monev_admin_action("DELETE", "MONEV_EXPENSE_TYPE", target_id=expense_type_id)
                flash("Jenis Belanja berhasil dihapus.", "success")
        return redirect(url_for("monev_bos.admin_expense_types"))

    expense_types = queries.list_expense_types(include_inactive=True)
    queries.attach_admin_input_names(expense_types, "MONEV_EXPENSE_TYPE")
    activity_logs = queries.list_admin_action_history(["MONEV_EXPENSE_TYPE"])
    return render_template(
        "monev_bos/admin/expense_types.html",
        expense_types=expense_types,
        activity_logs=activity_logs,
    )



@monev_bos_bp.route("/admin/account-codes", methods=["GET", "POST"])
@role_required("admin")
def admin_account_codes():
    if request.method == "POST":
        action = request.form.get("action")
        if action == "create":
            code = (request.form.get("code") or "").strip()
            name = (request.form.get("name") or "").strip()
            description = (request.form.get("description") or "").strip()
            if code:
                if queries.check_account_code_exists(code):
                    flash(f"Gagal! Kode Rekening '{code}' sudah ada. Tidak boleh ada duplikasi.", "danger")
                else:
                    try:
                        account_code_id = queries.create_account_code(code, name, description)
                        _record_monev_admin_action(
                            "CREATE",
                            "MONEV_ACCOUNT_CODE",
                            target_id=account_code_id,
                            target_name=f"{code} - {name}" if name else code,
                        )
                        flash("Master Kode Rekening berhasil ditambahkan.", "success")
                    except psycopg2.errors.UniqueViolation:
                        flash(f"Gagal! Kode Rekening '{code}' sudah ada. Tidak boleh ada duplikasi.", "danger")
            else:
                flash("Kode Rekening wajib diisi.", "warning")
        elif action == "update":
            account_code_id = int(request.form.get("account_code_id"))
            code = (request.form.get("code") or "").strip()
            name = (request.form.get("name") or "").strip()
            description = (request.form.get("description") or "").strip()
            is_active = request.form.get("is_active") == "on"
            if code and account_code_id:
                if queries.check_account_code_exists(code, exclude_id=account_code_id):
                    flash(f"Gagal! Kode Rekening '{code}' sudah terpakai oleh data lain.", "danger")
                else:
                    try:
                        queries.update_account_code(account_code_id, code, name, description, is_active)
                        _record_monev_admin_action(
                            "UPDATE",
                            "MONEV_ACCOUNT_CODE",
                            target_id=account_code_id,
                            target_name=f"{code} - {name}" if name else code,
                            metadata={"is_active": is_active},
                        )
                        flash("Master Kode Rekening berhasil diperbarui.", "success")
                    except psycopg2.errors.UniqueViolation:
                        flash(f"Gagal! Kode Rekening '{code}' sudah terpakai oleh data lain.", "danger")
        elif action == "delete":
            account_code_id = int(request.form.get("account_code_id"))
            if account_code_id:
                queries.delete_account_code(account_code_id)
                _record_monev_admin_action("DELETE", "MONEV_ACCOUNT_CODE", target_id=account_code_id)
                flash("Master Kode Rekening berhasil dihapus.", "success")
        return redirect(url_for("monev_bos.admin_account_codes"))

    account_codes = queries.list_account_codes(include_inactive=True)
    queries.attach_admin_input_names(account_codes, "MONEV_ACCOUNT_CODE")
    activity_logs = queries.list_admin_action_history(["MONEV_ACCOUNT_CODE"])
    return render_template(
        "monev_bos/admin/master_account_codes.html",
        account_codes=account_codes,
        activity_logs=activity_logs,
    )

@monev_bos_bp.route("/admin/edit-requests", methods=["GET", "POST"])
@role_required("admin")
def admin_edit_requests():
    user = current_user()
    if request.method == "POST":
        action = request.form.get("action")
        request_id = int(request.form.get("request_id"))
        review_notes = request.form.get("review_notes", "").strip()

        if action == "approve":
            queries.approve_edit_request(request_id, user["id"], review_notes or None)
            _record_monev_admin_action(
                "APPROVE",
                "MONEV_EDIT_REQUEST",
                target_id=request_id,
                metadata={"has_review_notes": bool(review_notes)},
            )
            flash("Pengajuan edit telah disetujui dan data kegiatan berhasil diperbarui.", "success")
        elif action == "reject":
            queries.reject_edit_request(request_id, user["id"], review_notes or None)
            _record_monev_admin_action(
                "REJECT",
                "MONEV_EDIT_REQUEST",
                target_id=request_id,
                metadata={"has_review_notes": bool(review_notes)},
            )
            flash("Pengajuan edit ditolak.", "info")

        return redirect(url_for("monev_bos.admin_edit_requests"))

    status_filter = request.args.get("status", "pending")
    edit_requests = queries.list_edit_requests(status=status_filter if status_filter != "all" else None)
    return render_template("monev_bos/admin/edit_requests.html",
                           edit_requests=edit_requests,
                           status_filter=status_filter)



@monev_bos_bp.route("/admin/teams", methods=["GET", "POST"])
@role_required("admin")
def admin_teams():
    if request.method == "POST":
        action = request.form.get("action")
        if action == "create":
            name = request.form.get("name")
            leader_id = request.form.get("leader_id")
            leader_id = int(leader_id) if leader_id else None
            team_id = queries.create_team(name, leader_id)
            _record_monev_admin_action(
                "CREATE",
                "MONEV_TEAM",
                target_id=team_id,
                target_name=name,
                metadata={"leader_id": leader_id},
            )
            flash("Tim berhasil dibuat", "success")
        elif action == "update_team":
            team_id = int(request.form.get("team_id"))
            name = (request.form.get("name") or "").strip()
            leader_id = request.form.get("leader_id")
            leader_id = int(leader_id) if leader_id else None
            if not name:
                flash("Nama tim wajib diisi.", "warning")
            elif len(name) > 150:
                flash("Nama tim maksimal 150 karakter.", "warning")
            else:
                queries.update_team(team_id, name, leader_id)
                _record_monev_admin_action(
                    "UPDATE",
                    "MONEV_TEAM",
                    target_id=team_id,
                    target_name=name,
                    metadata={"leader_id": leader_id},
                )
                flash("Nama dan ketua tim berhasil diperbarui", "success")
        elif action == "update_leader":
            team_id = int(request.form.get("team_id"))
            leader_id = request.form.get("leader_id")
            leader_id = int(leader_id) if leader_id else None
            queries.update_team_leader(team_id, leader_id)
            _record_monev_admin_action(
                "UPDATE_LEADER",
                "MONEV_TEAM",
                target_id=team_id,
                metadata={"leader_id": leader_id},
            )
            flash("Ketua tim berhasil diupdate", "success")
        elif action == "delete":
            team_id = int(request.form.get("team_id"))
            queries.delete_team(team_id)
            _record_monev_admin_action("DELETE", "MONEV_TEAM", target_id=team_id)
            flash("Tim berhasil dihapus", "success")
        elif action == "assign":
            team_id = int(request.form.get("team_id"))
            school_id = int(request.form.get("school_id"))
            period_value = (request.form.get("period_id") or "").strip()
            active_period_ids = [
                int(period["id"]) for period in queries.get_active_periods()
            ]
            if period_value == "all":
                target_period_ids = active_period_ids
            elif period_value.isdigit() and int(period_value) in active_period_ids:
                target_period_ids = [int(period_value)]
            else:
                target_period_ids = []

            if not target_period_ids:
                flash("Periode penugasan tidak aktif atau tidak ditemukan.", "warning")
            else:
                for period_id in target_period_ids:
                    queries.assign_team_to_school(team_id, school_id, period_id)
                    _record_monev_admin_action(
                        "ASSIGN",
                        "MONEV_SCHOOL_ASSIGNMENT",
                        target_id=school_id,
                        metadata={
                            "team_id": team_id,
                            "period_id": period_id,
                            "all_active_periods": period_value == "all",
                        },
                    )
                if period_value == "all":
                    flash(
                        f"Sekolah berhasil ditugaskan ke tim pada {len(target_period_ids)} periode aktif.",
                        "success",
                    )
                else:
                    flash("Sekolah berhasil ditugaskan ke tim", "success")
        elif action == "unassign":
            assignment_id = int(request.form.get("assignment_id"))
            queries.unassign_school(assignment_id)
            _record_monev_admin_action("UNASSIGN", "MONEV_SCHOOL_ASSIGNMENT", target_id=assignment_id)
            flash("Tugas sekolah berhasil dilepas", "success")
        elif action == "copy_all_assignments":
            source_period_id = int(request.form.get("source_period_id"))
            target_period_id = int(request.form.get("target_period_id"))
            active_period_ids = {
                int(period["id"]) for period in queries.get_active_periods()
            }
            if source_period_id not in active_period_ids:
                flash("Periode asal sudah tidak aktif.", "warning")
            elif target_period_id not in active_period_ids:
                flash("Periode tujuan tidak aktif atau tidak ditemukan.", "warning")
            elif source_period_id == target_period_id:
                flash("Periode tujuan harus berbeda dari periode asal.", "warning")
            else:
                copy_result = queries.copy_all_assignments_between_periods(
                    source_period_id,
                    target_period_id,
                )
                if copy_result["source_count"] == 0:
                    flash("Periode asal belum memiliki penugasan untuk disalin.", "warning")
                else:
                    _record_monev_admin_action(
                        "COPY_ALL_ASSIGNMENTS",
                        "MONEV_SCHOOL_ASSIGNMENT",
                        target_id=target_period_id,
                        metadata={
                            "source_period_id": source_period_id,
                            "target_period_id": target_period_id,
                            **copy_result,
                        },
                    )
                    message = f"{copy_result['copied_count']} penugasan berhasil disalin ke periode tujuan."
                    if copy_result["skipped_count"]:
                        message += f" {copy_result['skipped_count']} sekolah dilewati karena sudah memiliki penugasan di periode tujuan."
                    flash(message, "success")
        return redirect(url_for("monev_bos.admin_teams"))

    teams = queries.list_teams()
    queries.attach_admin_input_names(teams, "MONEV_TEAM")
    staff_users = queries.get_staff_users()
    sekolah_users = queries.get_sekolah_users()
    active_periods = queries.get_active_periods()
    assignments = queries.list_assignments_for_periods(
        [int(period["id"]) for period in active_periods]
    )
    queries.attach_admin_input_names(
        assignments,
        "MONEV_SCHOOL_ASSIGNMENT",
        actions=["ASSIGN"],
        item_id_field="school_id",
    )
    assignment_groups_by_school = {}
    for assignment in assignments:
        school_id = int(assignment["school_id"])
        group = assignment_groups_by_school.setdefault(
            school_id,
            {
                "school_id": school_id,
                "school_name": assignment.get("school_name") or "Sekolah",
                "school_email": assignment.get("school_email"),
                "assignments": [],
                "teams": [],
                "admin_names": [],
            },
        )
        group["assignments"].append(assignment)
        if not any(team["id"] == int(assignment["team_id"]) for team in group["teams"]):
            group["teams"].append(
                {"id": int(assignment["team_id"]), "name": assignment["team_name"]}
            )
        admin_name = assignment.get("input_admin_name") or "Tidak ada"
        if admin_name not in group["admin_names"]:
            group["admin_names"].append(admin_name)
    assignment_groups = sorted(
        assignment_groups_by_school.values(),
        key=lambda group: str(group["school_name"]).lower(),
    )
    team_groups_by_id = {}
    for assignment in assignments:
        team_id = int(assignment["team_id"])
        school_id = int(assignment["school_id"])
        team_school_groups = team_groups_by_id.setdefault(team_id, {})
        school_group = team_school_groups.setdefault(
            school_id,
            {
                "school_id": school_id,
                "school_name": assignment.get("school_name") or "Sekolah",
                "school_email": assignment.get("school_email"),
                "assignments": [],
            },
        )
        school_group["assignments"].append(assignment)
    team_assignment_groups = {
        team_id: sorted(
            school_groups.values(),
            key=lambda group: str(group["school_name"]).lower(),
        )
        for team_id, school_groups in team_groups_by_id.items()
    }
    activity_logs = queries.list_admin_action_history(
        ["MONEV_TEAM", "MONEV_TEAM_MEMBER", "MONEV_SCHOOL_ASSIGNMENT"]
    )

    return render_template(
        "monev_bos/admin/teams.html", 
        teams=teams, 
        staff_users=staff_users,
        sekolah_users=sekolah_users,
        active_periods=active_periods,
        assignments=assignments,
        assignment_groups=assignment_groups,
        team_assignment_groups=team_assignment_groups,
        activity_logs=activity_logs,
    )

@monev_bos_bp.route("/sekolah")
@role_required("sekolah")
def sekolah_dashboard():
    user = current_user()
    school_info = queries.get_school_kecamatan_and_admin_wa(user["id"])
    
    # Available years filter
    available_years = queries.get_available_years()
    current_year = datetime.now().year
    if current_year not in available_years:
        available_years.insert(0, current_year)
    if not available_years:
        available_years = [2026, 2025]
    
    selected_year = request.args.get("year", type=int)
    if not selected_year or selected_year not in available_years:
        selected_year = available_years[0]
        
    existing_periods = queries.list_periods_by_year(selected_year)
    period_map = {p["tw"]: p for p in existing_periods}

    tw_cards = []
    tw_names = {1: "TW I", 2: "TW II", 3: "TW III", 4: "TW IV"}
    tw_months = {1: "Januari - Maret", 2: "April - Juni", 3: "Juli - September", 4: "Oktober - Desember"}

    for tw in range(1, 5):
        period = period_map.get(tw)
        if period:
            period_id = period["id"]
            is_active = period["is_active"]
            report = queries.get_school_report(user["id"], period_id)
        else:
            period_id = None
            is_active = False
            report = None

        bosp_receipt = float(report["bosp_receipt_amount"]) if report and report.get("bosp_receipt_amount") else 0.0
        bop_receipt = float(report["bop_receipt_amount"]) if report and report.get("bop_receipt_amount") else 0.0
        total_receipt = bosp_receipt + bop_receipt

        activities = queries.list_activities(report["id"]) if report else []
        bos_realized = float(sum(a["realized_amount"] for a in activities if a["fund_source"] == "BOS"))
        bop_realized = float(sum(a["realized_amount"] for a in activities if a["fund_source"] == "BOP"))
        total_realized = float(sum(a["realized_amount"] for a in activities))
        remaining_balance = total_receipt - total_realized

        valid_count = sum(1 for a in activities if a["status"] == "valid")
        invalid_count = sum(1 for a in activities if a["status"] == "invalid")
        in_review_count = sum(1 for a in activities if a["status"] == "in_review")
        pending_count = sum(1 for a in activities if a["status"] == "pending")

        status_key = "closed"
        if is_active:
            if not report:
                status_key = "unfilled"
            else:
                status_key = report["status"]

        card = {
            "tw": tw,
            "tw_label": tw_names[tw],
            "tw_months": tw_months[tw],
            "period": period,
            "period_id": period_id,
            "is_active": is_active,
            "report": report,
            "status_key": status_key,
            "bosp_receipt": bosp_receipt,
            "bop_receipt": bop_receipt,
            "total_receipt": total_receipt,
            "bos_realized": bos_realized,
            "bop_realized": bop_realized,
            "total_realized": total_realized,
            "remaining_balance": remaining_balance,
            "total_activities": len(activities),
            "valid_count": valid_count,
            "invalid_count": invalid_count,
            "in_review_count": in_review_count,
            "pending_count": pending_count,
        }
        tw_cards.append(card)

    year_stats = {
        "total_bosp_receipt": sum(c["bosp_receipt"] for c in tw_cards),
        "total_bop_receipt": sum(c["bop_receipt"] for c in tw_cards),
        "total_receipt": sum(c["total_receipt"] for c in tw_cards),
        "total_realized": sum(c["total_realized"] for c in tw_cards),
        "total_remaining": sum(c["remaining_balance"] for c in tw_cards),
        "filled_tw_count": sum(1 for c in tw_cards if c["report"]),
        "completed_tw_count": sum(1 for c in tw_cards if c["report"] and c["report"]["status"] in ["completed", "completed_with_notes"]),
        "revision_tw_count": sum(1 for c in tw_cards if c["report"] and c["report"]["status"] == "needs_revision"),
    }

    return render_template(
        "monev_bos/sekolah/dashboard.html",
        selected_year=selected_year,
        available_years=available_years,
        tw_cards=tw_cards,
        year_stats=year_stats,
        school_info=school_info
    )

def _parse_float(val) -> float:
    if not val:
        return 0.0
    val_str = str(val).replace(".", "").replace(",", ".")
    try:
        return float(val_str)
    except ValueError:
        return 0.0

@monev_bos_bp.route("/sekolah/report", methods=["GET", "POST"])
@role_required("sekolah")
def sekolah_report_form():
    user = current_user()
    period_id = request.args.get("period_id", type=int)
    if not period_id:
        flash("Pilih triwulan terlebih dahulu.", "warning")
        return redirect(url_for("monev_bos.sekolah_dashboard"))
    
    active_periods = queries.get_active_periods()
    active_period = next((p for p in active_periods if p["id"] == period_id), None)
    if not active_period:
        flash("Periode tidak ditemukan atau belum dibuka.", "warning")
        return redirect(url_for("monev_bos.sekolah_dashboard"))

    report = queries.get_school_report(user["id"], active_period["id"])
    
    if request.method == "POST":
        if report and report["status"] not in ["draft", "needs_revision"]:
            flash("Laporan sudah tidak bisa diubah.", "warning")
            return redirect(url_for("monev_bos.sekolah_dashboard"))
            
        bosp = _parse_float(request.form.get("bosp_amount"))
        bop = _parse_float(request.form.get("bop_amount"))
        queries.save_school_report_receipts(user["id"], active_period["id"], bosp, bop)
        flash("Data penerimaan dana berhasil disimpan.", "success")
        return redirect(url_for("monev_bos.sekolah_funds", period_id=period_id))

    return render_template("monev_bos/sekolah/form.html", active_period=active_period, report=report)

@monev_bos_bp.route("/sekolah/funds")
@role_required("sekolah")
def sekolah_funds():
    user = current_user()
    period_id = request.args.get("period_id", type=int)
    if not period_id:
        flash("Pilih triwulan terlebih dahulu.", "warning")
        return redirect(url_for("monev_bos.sekolah_dashboard"))
    
    active_periods = queries.get_active_periods()
    active_period = next((p for p in active_periods if p["id"] == period_id), None)
    if not active_period:
        flash("Periode tidak ditemukan atau belum dibuka.", "warning")
        return redirect(url_for("monev_bos.sekolah_dashboard"))

    report = queries.get_school_report(user["id"], active_period["id"])
    if not report:
        flash("Silakan isi data penerimaan dana terlebih dahulu.", "warning")
        return redirect(url_for("monev_bos.sekolah_report_form", period_id=period_id))

    all_activities = queries.list_activities(report["id"])
    
    bos_activities = [a for a in all_activities if a["fund_source"] == "BOS"]
    bop_activities = [a for a in all_activities if a["fund_source"] == "BOP"]
    
    bos_realized = sum(a["realized_amount"] for a in bos_activities)
    bop_realized = sum(a["realized_amount"] for a in bop_activities)

    return render_template("monev_bos/sekolah/funds.html", 
                           active_period=active_period, 
                           report=report,
                           bos_count=len(bos_activities),
                           bop_count=len(bop_activities),
                           bos_realized=bos_realized,
                           bop_realized=bop_realized)

import os
import io
from PIL import Image, ImageOps
from werkzeug.utils import secure_filename

def _compress_image_file(file_storage, target_kb=100) -> bytes:
    """Auto-compress image to <= target_kb (100KB) automatically."""
    target_bytes = target_kb * 1024
    
    file_storage.seek(0)
    img = Image.open(file_storage)
    
    if img.mode in ("RGBA", "P"):
        img = img.convert("RGB")
        
    quality = 80
    output = io.BytesIO()
    img.save(output, format="JPEG", quality=quality, optimize=True)
    
    while output.tell() > target_bytes and quality > 20:
        quality -= 15
        output = io.BytesIO()
        img.save(output, format="JPEG", quality=quality, optimize=True)
        
    width, height = img.size
    while output.tell() > target_bytes and width > 300 and height > 300:
        width = int(width * 0.75)
        height = int(height * 0.75)
        resized_img = img.resize((width, height), Image.Resampling.LANCZOS)
        output = io.BytesIO()
        resized_img.save(output, format="JPEG", quality=45, optimize=True)
        
    return output.getvalue()

def _save_uploaded_file(file, base_dir, sub_path, max_size_bytes=100 * 1024):
    if not file or file.filename == '':
        return None, "File tidak ada"
    
    filename = secure_filename(file.filename)
    ext = os.path.splitext(filename)[1].lower()
    if ext not in [".jpg", ".jpeg", ".png", ".webp", ".bmp"]:
        return None, "Dokumen hanya menerima file gambar (JPG, PNG, WEBP, atau BMP)."
    
    upload_dir = os.path.join(base_dir, sub_path)
    os.makedirs(upload_dir, exist_ok=True)
    
    try:
        compressed_bytes = _compress_image_file(file, target_kb=max_size_bytes // 1024)
        save_name = os.path.splitext(filename)[0] + ".jpg"
        file_path = os.path.join(upload_dir, save_name)
        with open(file_path, "wb") as f:
            f.write(compressed_bytes)
        return f"static/uploads/{sub_path}/{save_name}", None
    except Exception:
        return None, "File gambar gagal diproses. Pastikan file tidak rusak."

def _save_camera_photo(data_url, base_dir, sub_path):
    if not data_url:
        return None, "Foto kegiatan wajib diambil langsung dari kamera."
    if "," not in data_url:
        return None, "Format foto kamera tidak valid."

    try:
        header, encoded = data_url.split(",", 1)
        if not header.startswith("data:image/"):
            return None, "Format foto kamera harus berupa gambar."
        image_bytes = base64.b64decode(encoded)
        img = Image.open(io.BytesIO(image_bytes))
        if img.mode in ("RGBA", "P"):
            img = img.convert("RGB")

        upload_dir = os.path.join(base_dir, sub_path)
        os.makedirs(upload_dir, exist_ok=True)
        filename = f"camera_{uuid.uuid4().hex[:10]}.jpg"
        file_path = os.path.join(upload_dir, filename)
        img.thumbnail((1200, 1200), Image.Resampling.LANCZOS)
        img.save(file_path, format="JPEG", quality=75, optimize=True)
        return f"static/uploads/{sub_path}/{filename}", None
    except Exception:
        return None, "Foto kamera gagal diproses. Silakan ambil ulang foto."


def _save_story_camera_photo(data_url: str, school_user_id: int) -> tuple:
    """Save a camera-only story photo as JPEG capped at 200 KB."""
    if not data_url or "," not in data_url:
        return None, 0, "Foto Live wajib diambil melalui kamera."

    try:
        header, encoded = data_url.split(",", 1)
        if not header.startswith("data:image/"):
            return None, 0, "Format Foto Live tidak valid."
        image_bytes = base64.b64decode(encoded, validate=True)
        image = ImageOps.exif_transpose(Image.open(io.BytesIO(image_bytes)))
        if image.mode != "RGB":
            image = image.convert("RGB")
        image.thumbnail((1600, 1600), Image.Resampling.LANCZOS)

        target_bytes = 200 * 1024
        quality = 85
        output = io.BytesIO()
        while True:
            output = io.BytesIO()
            image.save(output, format="JPEG", quality=quality, optimize=True)
            if output.tell() <= target_bytes:
                break
            if quality > 35:
                quality -= 10
                continue
            width, height = image.size
            if width <= 320 or height <= 320:
                return None, 0, "Foto tidak dapat dikompres hingga maksimal 200 KB."
            image = image.resize(
                (max(320, int(width * 0.82)), max(320, int(height * 0.82))),
                Image.Resampling.LANCZOS,
            )

        relative_dir = f"monev_bos/stories/{school_user_id}"
        upload_root = os.path.join(monev_bos_bp.root_path, "..", "static", "uploads")
        upload_dir = os.path.join(upload_root, relative_dir)
        os.makedirs(upload_dir, exist_ok=True)
        filename = f"story_{uuid.uuid4().hex}.jpg"
        absolute_path = os.path.join(upload_dir, filename)
        with open(absolute_path, "wb") as file_handle:
            file_handle.write(output.getvalue())
        return f"static/uploads/{relative_dir}/{filename}", output.tell(), None
    except Exception:
        current_app.logger.exception("Failed to process Monev school story photo")
        return None, 0, "Foto Live gagal diproses. Silakan ambil ulang."


def _absolute_dashboard_file_path(relative_path: str) -> str:
    return os.path.normpath(os.path.join(monev_bos_bp.root_path, "..", (relative_path or "").lstrip("/")))


@monev_bos_bp.before_app_request
def protect_private_school_story_files():
    """Prevent direct /static access to private Foto Live assets."""
    relative_path = request.path.lstrip("/")
    if not relative_path.startswith("static/uploads/monev_bos/stories/"):
        return None
    post = queries.get_school_post_by_photo_path(relative_path)
    if not post or post.get("is_public"):
        return None
    user = current_user() or {}
    if user.get("role") in {"admin", "staff"}:
        return None
    if user.get("role") == "sekolah":
        if int(user.get("id") or 0) == int(post["school_user_id"]):
            return None
        if post.get("is_active_story"):
            return None
    abort(404)


def _can_view_school_post_photo(post: dict) -> bool:
    if post.get("is_public"):
        return True
    user = current_user() or {}
    if user.get("role") in {"admin", "staff"}:
        return True
    if user.get("role") != "sekolah":
        return False
    if int(user.get("id") or 0) == int(post["school_user_id"]):
        return True
    return bool(post.get("is_active_story"))


def _load_report_font(size: int, bold: bool = False):
    from PIL import ImageFont

    candidates = (
        ["/System/Library/Fonts/Supplemental/Arial Bold.ttf", "/usr/share/fonts/truetype/dejavu/DejaVuSans-Bold.ttf"]
        if bold
        else ["/System/Library/Fonts/Supplemental/Arial.ttf", "/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf"]
    )
    for path in candidates:
        if os.path.isfile(path):
            return ImageFont.truetype(path, size=size)
    return ImageFont.load_default()


def _draw_wrapped_text(draw, text: str, xy: tuple, font, fill, max_width: int, line_gap: int = 8) -> int:
    words = (text or "-").split()
    lines = []
    current = ""
    for word in words:
        candidate = f"{current} {word}".strip()
        if current and draw.textbbox((0, 0), candidate, font=font)[2] > max_width:
            lines.append(current)
            current = word
        else:
            current = candidate
    if current:
        lines.append(current)
    x, y = xy
    line_height = draw.textbbox((0, 0), "Ag", font=font)[3] + line_gap
    for line in lines:
        draw.text((x, y), line, font=font, fill=fill)
        y += line_height
    return y


def _build_school_photo_report_pdf(posts: list, profile: dict) -> io.BytesIO:
    """Build one adaptive A4 page containing one to six authenticated photos."""
    import qrcode
    from PIL import ImageDraw

    page_width, page_height = 1240, 1754
    margin = 50
    page = Image.new("RGB", (page_width, page_height), "white")
    draw = ImageDraw.Draw(page)
    title_font = _load_report_font(31, bold=True)
    school_font = _load_report_font(22, bold=True)
    header_meta_font = _load_report_font(16)

    draw.text((margin, 42), "LAPORAN FOTO LIVE SEKOLAH", font=title_font, fill="#153e75")
    draw.text((margin, 88), profile.get("school_name") or "Sekolah", font=school_font, fill="#111827")
    draw.text((page_width - margin, 55), f"{len(posts)} foto / 1 lembar", font=header_meta_font, fill="#4b5563", anchor="ra")
    draw.text((page_width - margin, 88), "Scan QR pada tiap foto untuk verifikasi", font=header_meta_font, fill="#4b5563", anchor="ra")
    draw.line((margin, 130, page_width - margin, 130), fill="#cbd5e1", width=3)

    if len(posts) == 1:
        columns, rows = 1, 1
    elif len(posts) == 2:
        columns, rows = 1, 2
    elif len(posts) <= 4:
        columns, rows = 2, 2
    else:
        columns, rows = 2, 3

    gap = 18
    content_top = 154
    content_bottom = page_height - 72
    content_width = page_width - (margin * 2)
    content_height = content_bottom - content_top
    card_width = (content_width - (gap * (columns - 1))) // columns
    card_height = (content_height - (gap * (rows - 1))) // rows
    info_height = 190 if rows == 1 else (132 if rows == 2 else 126)
    qr_size = 158 if rows == 1 else (104 if rows == 2 else 100)
    card_title_font = _load_report_font(24 if rows == 1 else (18 if rows == 2 else 15), bold=True)
    card_meta_font = _load_report_font(17 if rows == 1 else (14 if rows == 2 else 12))
    stamp_font = _load_report_font(17 if rows == 1 else (14 if rows == 2 else 12), bold=True)

    def one_line(text: str, font, max_width: int) -> str:
        clean = " ".join((text or "-").split())
        if draw.textbbox((0, 0), clean, font=font)[2] <= max_width:
            return clean
        suffix = "..."
        while clean and draw.textbbox((0, 0), clean + suffix, font=font)[2] > max_width:
            clean = clean[:-1]
        return clean.rstrip() + suffix

    for index, post in enumerate(posts):
        row, column = divmod(index, columns)
        card_x = margin + column * (card_width + gap)
        card_y = content_top + row * (card_height + gap)
        card_right = card_x + card_width
        card_bottom = card_y + card_height
        padding = 12
        draw.rounded_rectangle(
            (card_x, card_y, card_right, card_bottom),
            radius=14,
            fill="#ffffff",
            outline="#cbd5e1",
            width=3,
        )

        photo_x = card_x + padding
        photo_y = card_y + padding
        photo_width = card_width - (padding * 2)
        photo_height = card_height - info_height - (padding * 2)
        draw.rounded_rectangle(
            (photo_x, photo_y, photo_x + photo_width, photo_y + photo_height),
            radius=8,
            fill="#111827",
        )
        source_path = _absolute_dashboard_file_path(post["photo_path"])
        with Image.open(source_path) as source:
            source = ImageOps.exif_transpose(source).convert("RGB")
            photo = ImageOps.contain(source, (photo_width, photo_height), Image.Resampling.LANCZOS)
        paste_x = photo_x + (photo_width - photo.width) // 2
        paste_y = photo_y + (photo_height - photo.height) // 2
        page.paste(photo, (paste_x, paste_y))

        timestamp = post["created_at"].strftime("%d/%m/%Y %H:%M WIB")
        stamp_text = f"{timestamp}"
        stamp_box = draw.textbbox((0, 0), stamp_text, font=stamp_font)
        stamp_width = stamp_box[2] - stamp_box[0] + 20
        stamp_height = stamp_box[3] - stamp_box[1] + 16
        stamp_x = photo_x + 9
        stamp_y = photo_y + photo_height - stamp_height - 9
        draw.rounded_rectangle((stamp_x, stamp_y, stamp_x + stamp_width, stamp_y + stamp_height), radius=6, fill="#111827")
        draw.text((stamp_x + 10, stamp_y + 6), stamp_text, font=stamp_font, fill="white")

        verification_url = url_for(
            "monev_bos.public_school_post_verification",
            public_token=post["public_token"],
            _external=True,
        )
        qr = qrcode.QRCode(version=None, box_size=6, border=2, error_correction=qrcode.constants.ERROR_CORRECT_M)
        qr.add_data(verification_url)
        qr.make(fit=True)
        qr_image = qr.make_image(fill_color="#111827", back_color="white").convert("RGB")
        qr_image = qr_image.resize((qr_size, qr_size), Image.Resampling.NEAREST)
        qr_x = card_right - padding - qr_size
        info_y = photo_y + photo_height + 9
        page.paste(qr_image, (qr_x, info_y))

        text_x = card_x + padding
        text_width = qr_x - text_x - 10
        title = one_line(post.get("title") or "Foto Live", card_title_font, text_width)
        draw.text((text_x, info_y + 2), title, font=card_title_font, fill="#111827")
        authenticity_code = (post.get("photo_sha256") or post.get("public_token") or "")[:12].upper()
        if rows <= 2:
            location = one_line(f"Lokasi: {post.get('location_text') or '-'}", card_meta_font, text_width)
            draw.text((text_x, info_y + (38 if rows == 1 else 29)), location, font=card_meta_font, fill="#374151")
            code_y = info_y + (70 if rows == 1 else 53)
        else:
            code_y = info_y + 25
        draw.text((text_x, code_y), f"Kode: {authenticity_code}", font=card_meta_font, fill="#4b5563")
        draw.text((qr_x + qr_size // 2, min(info_y + qr_size + 2, card_bottom - 18)), "VERIFIKASI", font=card_meta_font, fill="#153e75", anchor="ma")

    footer_font = _load_report_font(13)
    draw.text(
        (page_width // 2, page_height - 40),
        "Foto dipublikasikan atas persetujuan pemilik. Sidik jari SHA-256 tercatat pada halaman verifikasi.",
        font=footer_font,
        fill="#4b5563",
        anchor="ma",
    )

    output = io.BytesIO()
    page.save(output, format="PDF", resolution=150.0)
    output.seek(0)
    return output


@monev_bos_bp.route("/posts")
@role_required("admin", "sekolah")
def school_posts_explore():
    user = current_user()
    search_query = request.args.get("q", "").strip()
    posts = queries.list_school_posts(
        search_query=search_query,
        shared_only=user.get("role") == "sekolah",
        limit=300,
    )
    audit_logs = queries.list_story_audit_logs(limit=100) if user.get("role") == "admin" else []
    return render_template(
        "monev_bos/social/posts.html",
        posts=posts,
        search_query=search_query,
        audit_logs=audit_logs,
    )


@monev_bos_bp.route("/schools/<int:school_user_id>/posts")
@role_required("admin", "sekolah")
def school_posts_profile(school_user_id: int):
    user = current_user()
    is_own_profile = user.get("role") == "sekolah" and int(school_user_id) == int(user["id"])
    profile = queries.get_school_post_profile(school_user_id)
    if not profile:
        flash("Profil sekolah tidak ditemukan.", "warning")
        if user.get("role") == "sekolah":
            return redirect(url_for("monev_bos.sekolah_dashboard"))
        return redirect(url_for("monev_bos.school_posts_explore"))
    search_query = request.args.get("q", "").strip()
    posts = queries.list_school_posts(
        school_user_id=school_user_id,
        search_query=search_query,
        shared_only=user.get("role") == "sekolah" and not is_own_profile,
        limit=300,
    )
    external_photo_links = (
        queries.list_external_photo_links(school_user_id)
        if user.get("role") == "sekolah" and int(user["id"]) == int(school_user_id)
        else []
    )
    external_photo_teachers = (
        queries.list_external_photo_teachers(school_user_id)
        if user.get("role") == "sekolah" and int(user["id"]) == int(school_user_id)
        else []
    )
    return render_template(
        "monev_bos/social/school_profile.html",
        profile=profile,
        posts=posts,
        search_query=search_query,
        external_photo_links=external_photo_links,
        external_photo_teachers=external_photo_teachers,
        can_manage_profile=user.get("role") == "admin" or is_own_profile,
    )


@monev_bos_bp.route("/external-photo-teachers", methods=["POST"])
@role_required("sekolah")
def save_external_photo_teacher():
    user = current_user()
    full_name = (request.form.get("full_name") or "").strip()
    nip = (request.form.get("nip") or "").strip()
    errors = validate_external_identity(full_name, nip)
    if errors:
        for error in errors:
            flash(error, "danger")
    else:
        queries.save_external_photo_teacher(int(user["id"]), full_name, nip, int(user["id"]))
        flash("Guru berhasil didaftarkan untuk Foto Live eksternal.", "success")
    return redirect(url_for("monev_bos.school_posts_profile", school_user_id=user["id"], _anchor="external-photo-teachers"))


@monev_bos_bp.route("/external-photo-teachers/<int:teacher_id>/delete", methods=["POST"])
@role_required("sekolah")
def delete_external_photo_teacher(teacher_id: int):
    user = current_user()
    deleted = queries.delete_external_photo_teacher(teacher_id, int(user["id"]))
    flash(
        "Guru dihapus dari daftar Foto Live eksternal." if deleted else "Data guru tidak ditemukan.",
        "success" if deleted else "warning",
    )
    return redirect(url_for("monev_bos.school_posts_profile", school_user_id=user["id"], _anchor="external-photo-teachers"))


@monev_bos_bp.route("/external-photo-links/create", methods=["POST"])
@role_required("sekolah")
def create_external_photo_link():
    user = current_user()
    if not queries.list_external_photo_teachers(int(user["id"]), limit=1):
        flash("Daftarkan minimal satu guru sebelum membuat link Foto Live eksternal.", "warning")
        return redirect(url_for("monev_bos.school_posts_profile", school_user_id=user["id"], _anchor="external-photo-teachers"))
    queries.create_external_photo_link(int(user["id"]), int(user["id"]))
    flash("Link Foto Live eksternal dibuat dan aktif selama 24 jam.", "success")
    return redirect(url_for("monev_bos.school_posts_profile", school_user_id=user["id"]))


@monev_bos_bp.route("/external-photo-links/<int:link_id>/revoke", methods=["POST"])
@role_required("sekolah")
def revoke_external_photo_link(link_id: int):
    user = current_user()
    revoked = queries.revoke_external_photo_link(link_id, int(user["id"]), int(user["id"]))
    flash(
        "Link Foto Live eksternal berhasil dicabut." if revoked else "Link tidak ditemukan atau sudah dicabut.",
        "success" if revoked else "warning",
    )
    return redirect(url_for("monev_bos.school_posts_profile", school_user_id=user["id"]))


def _external_photo_token_is_rate_limited(public_id: str) -> bool:
    key = f"external_photo_attempt:{public_id}"
    state = session.get(key) or {}
    now = int(time.time())
    started_at = int(state.get("started_at") or now)
    if now - started_at >= 15 * 60:
        session.pop(key, None)
        return False
    return int(state.get("count") or 0) >= 5


def _record_external_photo_token_failure(public_id: str) -> None:
    key = f"external_photo_attempt:{public_id}"
    now = int(time.time())
    state = session.get(key) or {"count": 0, "started_at": now}
    if now - int(state.get("started_at") or now) >= 15 * 60:
        state = {"count": 0, "started_at": now}
    state["count"] = int(state.get("count") or 0) + 1
    session[key] = state


def _external_photo_verification(public_id: str, link_id: int) -> dict:
    state = session.get(f"external_photo_verified:{public_id}") or {}
    if int(state.get("link_id") or 0) != int(link_id):
        return {}
    if not state.get("teacher_id") or not state.get("teacher_name") or not state.get("teacher_nip"):
        return {}
    return state


@monev_bos_bp.route("/foto-eksternal/<public_id>", methods=["GET", "POST"])
def external_photo_capture(public_id: str):
    link = queries.get_external_photo_link(public_id)
    if not link:
        abort(404)

    verification_key = f"external_photo_verified:{public_id}"
    if request.method == "GET" and request.args.get("reset") == "1":
        session.pop(verification_key, None)
    verified_identity = _external_photo_verification(public_id, int(link["id"]))
    current_step = "photo" if request.args.get("step") == "photo" and verified_identity else "identity"

    if not link.get("is_active"):
        session.pop(verification_key, None)
        verified_identity = {}

    if request.method == "POST":
        action = request.form.get("action") or "verify_identity"
        errors = []
        if not link.get("is_active"):
            errors.append("Link ini sudah kedaluwarsa atau telah dicabut oleh sekolah.")

        if action == "verify_identity":
            current_step = "identity"
            teacher_nip = (request.form.get("teacher_nip") or "").strip()
            access_token = (request.form.get("access_token") or "").strip()
            identity_errors = validate_external_nip(teacher_nip)
            errors.extend(identity_errors)
            token_is_valid = False
            if _external_photo_token_is_rate_limited(public_id):
                errors.append("Terlalu banyak percobaan token. Coba kembali dalam 15 menit.")
            elif not access_token_matches(access_token, link.get("access_token") or ""):
                if link.get("is_active"):
                    _record_external_photo_token_failure(public_id)
                errors.append("Token 6 digit tidak sesuai.")
            else:
                token_is_valid = True

            registered_teacher = None
            if token_is_valid and link.get("is_active") and not identity_errors:
                registered_teacher = queries.get_external_photo_teacher(
                    int(link["school_user_id"]), teacher_nip
                )
                if not registered_teacher:
                    errors.append("NIP belum didaftarkan oleh sekolah untuk Foto Live eksternal.")

            if not errors:
                session[verification_key] = {
                    "link_id": int(link["id"]),
                    "teacher_id": int(registered_teacher["id"]),
                    "teacher_name": registered_teacher["full_name"],
                    "teacher_nip": registered_teacher["nip"],
                    "verified_at": int(time.time()),
                }
                session.pop(f"external_photo_attempt:{public_id}", None)
                return redirect(url_for("monev_bos.external_photo_capture", public_id=public_id, step="photo"))
        elif action == "submit_photo":
            current_step = "photo"
            verified_identity = _external_photo_verification(public_id, int(link["id"]))
            if not verified_identity:
                flash("Sesi verifikasi tidak ditemukan. Masukkan identitas dan token kembali.", "warning")
                return redirect(url_for("monev_bos.external_photo_capture", public_id=public_id))
            registered_teacher = queries.get_external_photo_teacher(
                int(link["school_user_id"]), verified_identity["teacher_nip"]
            )
            if not registered_teacher or int(registered_teacher["id"]) != int(verified_identity["teacher_id"]):
                session.pop(verification_key, None)
                flash("NIP sudah tidak terdaftar. Hubungi sekolah untuk mendaftarkannya kembali.", "warning")
                return redirect(url_for("monev_bos.external_photo_capture", public_id=public_id))

            title = (request.form.get("title") or "").strip()
            photo_data = request.form.get("photo_data") or ""
            if not title:
                errors.append("Judul foto wajib diisi.")
            elif len(title) > 200:
                errors.append("Judul foto maksimal 200 karakter.")
            try:
                latitude = float(request.form.get("latitude"))
                longitude = float(request.form.get("longitude"))
                accuracy_raw = request.form.get("location_accuracy")
                location_accuracy = float(accuracy_raw) if accuracy_raw else None
                if not (-90 <= latitude <= 90 and -180 <= longitude <= 180):
                    raise ValueError
            except (TypeError, ValueError):
                errors.append("Lokasi GPS wajib tersedia dan valid sebelum foto dikirim.")

        else:
            abort(400)

        if not errors and action == "submit_photo":
            photo_path, photo_size, photo_error = _save_story_camera_photo(
                photo_data, int(link["school_user_id"])
            )
            if photo_error:
                flash(photo_error, "danger")
            else:
                location_text = f"{latitude:.6f}, {longitude:.6f}"
                try:
                    queries.create_school_post(
                        school_user_id=int(link["school_user_id"]),
                        title=title,
                        photo_path=photo_path,
                        photo_size=photo_size,
                        latitude=latitude,
                        longitude=longitude,
                        location_accuracy=location_accuracy,
                        location_text=location_text,
                        actor_id=None,
                        external_link_id=int(link["id"]),
                        external_photographer_name=verified_identity["teacher_name"],
                        external_photographer_nip=verified_identity["teacher_nip"],
                    )
                except ValueError as exc:
                    try:
                        os.remove(_absolute_dashboard_file_path(photo_path))
                    except OSError:
                        pass
                    flash(str(exc), "danger")
                except Exception:
                    try:
                        os.remove(_absolute_dashboard_file_path(photo_path))
                    except OSError:
                        pass
                    raise
                else:
                    session.pop(verification_key, None)
                    flash("Foto berhasil dikirim ke profil sekolah.", "success")
                    return redirect(url_for("monev_bos.external_photo_capture", public_id=public_id, berhasil=1))

        for error in dict.fromkeys(errors):
            flash(error, "danger")

    return render_template(
        "monev_bos/social/external_photo_capture.html",
        link=link,
        submission_success=request.args.get("berhasil") == "1",
        current_step=current_step,
        verified_identity=verified_identity,
    )


@monev_bos_bp.route("/posts/<int:post_id>/photo")
def school_post_photo(post_id: int):
    post = queries.get_school_post(post_id)
    if not post or not _can_view_school_post_photo(post):
        abort(404)
    path = _absolute_dashboard_file_path(post["photo_path"])
    if not os.path.isfile(path):
        abort(404)
    response = send_file(path, conditional=True, max_age=0)
    response.headers["Cache-Control"] = "public, max-age=3600" if post.get("is_public") else "private, no-store"
    return response


@monev_bos_bp.route("/public/photos/<public_token>")
def public_school_post_verification(public_token: str):
    import hashlib
    import hmac

    post = queries.get_public_school_post(public_token)
    if not post:
        abort(404)
    path = _absolute_dashboard_file_path(post["photo_path"])
    if not os.path.isfile(path):
        abort(404)
    digest = hashlib.sha256()
    with open(path, "rb") as photo_file:
        for chunk in iter(lambda: photo_file.read(1024 * 1024), b""):
            digest.update(chunk)
    fingerprint_matches = bool(post.get("photo_sha256")) and hmac.compare_digest(
        digest.hexdigest(),
        post["photo_sha256"],
    )
    return render_template(
        "monev_bos/social/photo_verification.html",
        post=post,
        fingerprint_matches=fingerprint_matches,
    )


@monev_bos_bp.route("/public/photos/<public_token>/image")
def public_school_post_image(public_token: str):
    post = queries.get_public_school_post(public_token)
    if not post:
        abort(404)
    path = _absolute_dashboard_file_path(post["photo_path"])
    if not os.path.isfile(path):
        abort(404)
    response = send_file(path, conditional=True, max_age=3600)
    response.headers["Cache-Control"] = "public, max-age=3600"
    return response


@monev_bos_bp.route("/schools/<int:school_user_id>/posts/report.pdf", methods=["POST"])
@role_required("admin", "sekolah")
def school_posts_photo_report(school_user_id: int):
    import hashlib

    user = current_user()
    if user.get("role") == "sekolah" and int(user["id"]) != int(school_user_id):
        abort(403)

    raw_ids = request.form.getlist("post_ids")
    post_ids = list(dict.fromkeys(int(value) for value in raw_ids if value.isdigit()))
    if not 1 <= len(post_ids) <= 6:
        abort(400, description="Pilih minimal 1 dan maksimal 6 foto.")

    profile = queries.get_school_post_profile(school_user_id)
    posts = queries.get_school_posts_by_ids(post_ids, school_user_id)
    post_map = {int(post["id"]): post for post in posts}
    if not profile or len(post_map) != len(post_ids):
        abort(404, description="Satu atau beberapa foto tidak ditemukan.")

    photo_hashes = {}
    for post in posts:
        path = _absolute_dashboard_file_path(post["photo_path"])
        if not os.path.isfile(path):
            abort(404, description=f"File foto #{post['id']} tidak ditemukan.")
        digest = hashlib.sha256()
        with open(path, "rb") as photo_file:
            for chunk in iter(lambda: photo_file.read(1024 * 1024), b""):
                digest.update(chunk)
        photo_hashes[int(post["id"])] = digest.hexdigest()

    published_posts = queries.publish_school_posts(
        post_ids,
        school_user_id,
        int(user["id"]),
        photo_hashes=photo_hashes,
    )
    published_map = {int(post["id"]): post for post in published_posts}
    ordered_posts = [published_map[post_id] for post_id in post_ids]
    pdf = _build_school_photo_report_pdf(ordered_posts, profile)
    safe_school_name = secure_filename(profile.get("school_name") or f"sekolah-{school_user_id}")
    filename = f"laporan-foto-{safe_school_name or school_user_id}.pdf"
    return send_file(pdf, mimetype="application/pdf", as_attachment=True, download_name=filename)


@monev_bos_bp.route("/stories/create", methods=["POST"])
@role_required("sekolah")
def create_school_story():
    user = current_user()
    own_profile_url = url_for("monev_bos.school_posts_profile", school_user_id=user["id"])
    title = (request.form.get("title") or "").strip()
    photo_data = request.form.get("photo_data") or ""
    try:
        latitude = float(request.form.get("latitude"))
        longitude = float(request.form.get("longitude"))
        accuracy_raw = request.form.get("location_accuracy")
        location_accuracy = float(accuracy_raw) if accuracy_raw else None
    except (TypeError, ValueError):
        flash("Lokasi GPS wajib tersedia sebelum Foto Live diposting.", "danger")
        return redirect(request.referrer or own_profile_url)

    if not title:
        flash("Judul foto wajib diisi.", "danger")
        return redirect(request.referrer or own_profile_url)
    if len(title) > 200:
        flash("Judul foto maksimal 200 karakter.", "danger")
        return redirect(request.referrer or own_profile_url)
    if not (-90 <= latitude <= 90 and -180 <= longitude <= 180):
        flash("Koordinat GPS tidak valid.", "danger")
        return redirect(request.referrer or own_profile_url)

    photo_path, photo_size, error = _save_story_camera_photo(photo_data, int(user["id"]))
    if error:
        flash(error, "danger")
        return redirect(request.referrer or own_profile_url)

    location_text = f"{latitude:.6f}, {longitude:.6f}"
    try:
        post_id = queries.create_school_post(
            school_user_id=int(user["id"]),
            title=title,
            photo_path=photo_path,
            photo_size=photo_size,
            latitude=latitude,
            longitude=longitude,
            location_accuracy=location_accuracy,
            location_text=location_text,
            actor_id=int(user["id"]),
        )
    except Exception:
        try:
            os.remove(_absolute_dashboard_file_path(photo_path))
        except OSError:
            pass
        raise

    flash("Foto Live berhasil diposting dan tersimpan permanen di profil sekolah.", "success")
    return redirect(url_for("monev_bos.school_posts_profile", school_user_id=user["id"], _anchor=f"post-{post_id}"))


@monev_bos_bp.route("/posts/<int:post_id>/delete", methods=["POST"])
@role_required("admin", "sekolah")
def delete_school_post(post_id: int):
    user = current_user()
    fallback_url = (
        url_for("monev_bos.school_posts_profile", school_user_id=user["id"])
        if user.get("role") == "sekolah"
        else url_for("monev_bos.school_posts_explore")
    )
    post = queries.get_school_post(post_id)
    if not post:
        flash("Postingan tidak ditemukan.", "warning")
        return redirect(fallback_url)
    if user.get("role") == "sekolah" and int(post["school_user_id"]) != int(user["id"]):
        flash("Anda tidak berhak menghapus postingan sekolah lain.", "danger")
        return redirect(fallback_url)

    source_path = _absolute_dashboard_file_path(post["photo_path"])
    links = queries.list_post_activity_links(post_id)
    prepared_copies = []
    try:
        if links and not os.path.isfile(source_path):
            raise FileNotFoundError("File postingan sumber tidak ditemukan.")
        for link in links:
            if not link.get("activity_doc_id"):
                continue
            relative_dir = f"monev_bos/{link['report_id']}/{link['activity_id']}/field_photo"
            destination_dir = os.path.join(
                monev_bos_bp.root_path,
                "..",
                "static",
                "uploads",
                relative_dir,
            )
            os.makedirs(destination_dir, exist_ok=True)
            filename = f"post_{post_id}_{uuid.uuid4().hex[:10]}.jpg"
            destination = os.path.join(destination_dir, filename)
            shutil.copy2(source_path, destination)
            prepared_copies.append(
                (
                    link,
                    destination,
                    f"static/uploads/{relative_dir}/{filename}",
                    os.path.getsize(destination),
                )
            )

        copied_links = [
            {
                "link_id": int(link["id"]),
                "activity_doc_id": int(link["activity_doc_id"]),
                "copied_path": copied_path,
                "copied_size": copied_size,
            }
            for link, _destination, copied_path, copied_size in prepared_copies
        ]
        queries.finalize_school_post_delete(
            post_id,
            int(user["id"]),
            int(post["school_user_id"]),
            copied_links,
            {"title": post["title"], "copied_activity_count": len(prepared_copies)},
        )
    except Exception as exc:
        for _link, destination, _copied_path, _copied_size in prepared_copies:
            try:
                os.remove(destination)
            except OSError:
                pass
        current_app.logger.exception("Failed to delete Monev school post %s", post_id)
        flash(f"Postingan gagal dihapus: {exc}", "danger")
        return redirect(url_for("monev_bos.school_posts_profile", school_user_id=post["school_user_id"]))

    try:
        os.remove(source_path)
    except OSError:
        pass
    _record_monev_admin_action(
        "DELETE",
        "MONEV_SCHOOL_POST",
        target_id=post_id,
        target_name=post["title"],
        metadata={"school_user_id": post["school_user_id"], "copied_activity_count": len(prepared_copies)},
    )
    flash("Postingan berhasil dihapus. Foto kegiatan yang tertaut sudah diamankan.", "success")
    return redirect(url_for("monev_bos.school_posts_profile", school_user_id=post["school_user_id"]))

@monev_bos_bp.route("/sekolah/activities", methods=["GET", "POST"])
@role_required("sekolah")
def sekolah_activities():
    user = current_user()
    period_id = request.args.get("period_id", type=int) or (request.form.get("period_id", type=int) if request.method == "POST" else None)
    fund_source = request.args.get("fund_source", "BOS") or (request.form.get("fund_source", "BOS") if request.method == "POST" else "BOS")
    if fund_source not in ["BOS", "BOP"]:
        fund_source = "BOS"

    if not period_id:
        flash("Pilih triwulan terlebih dahulu.", "warning")
        return redirect(url_for("monev_bos.sekolah_dashboard"))
    
    active_periods = queries.get_active_periods()
    active_period = next((p for p in active_periods if p["id"] == period_id), None)
    if not active_period:
        flash("Periode tidak ditemukan atau belum dibuka.", "warning")
        return redirect(url_for("monev_bos.sekolah_dashboard"))

    report = queries.get_school_report(user["id"], active_period["id"])
    if not report:
        flash("Silakan isi data penerimaan dana terlebih dahulu.", "warning")
        return redirect(url_for("monev_bos.sekolah_report_form", period_id=period_id))

    if request.method == "POST":
        action = request.form.get("action")
        if action == "claim_bop_transactions":
            if fund_source != "BOP":
                flash("Klaim transaksi hanya tersedia untuk sumber dana BOP.", "warning")
            else:
                try:
                    identity = queries.get_school_claim_identity(user["id"])
                    claim = get_school_bop_claim(
                        identity.get("npsn"), active_period["year"], active_period["tw"]
                    )
                    if not claim:
                        flash("Data transaksi BOP untuk NPSN sekolah ini tidak ditemukan.", "warning")
                    else:
                        source_by_code = {
                            item["activity_code"]: item for item in claim["transactions"]
                        }
                        submitted_codes = request.form.getlist("claim_activity_code")
                        activity_names = request.form.getlist("claim_activity_name")
                        expense_type_ids = request.form.getlist("claim_expense_type_id")
                        bku_numbers = request.form.getlist("claim_bku_number")
                        submitted_account_codes = request.form.getlist("claim_account_code")
                        realized_amounts = request.form.getlist("claim_realized_amount")
                        vendor_ids = request.form.getlist("claim_vendor_id")
                        item_names = request.form.getlist("claim_item_name")
                        field_lists = [
                            activity_names,
                            expense_type_ids,
                            bku_numbers,
                            submitted_account_codes,
                            realized_amounts,
                            vendor_ids,
                            item_names,
                        ]
                        if not submitted_codes or any(len(values) != len(submitted_codes) for values in field_lists):
                            raise ValueError("Data step klaim tidak lengkap. Silakan buka ulang wizard klaim.")
                        if len(set(submitted_codes)) != len(submitted_codes):
                            raise ValueError("Terdapat transaksi ganda pada data klaim.")

                        active_account_codes = {
                            str(item["code"]).strip()
                            for item in queries.list_account_codes(include_inactive=False)
                        }

                        adjusted_transactions = []
                        for index, activity_code in enumerate(submitted_codes):
                            source_item = source_by_code.get(activity_code)
                            if not source_item:
                                raise ValueError("Salah satu transaksi tidak berasal dari dataset sekolah ini.")
                            activity_name = activity_names[index].strip()
                            bku_number = bku_numbers[index].strip()
                            account_code = submitted_account_codes[index].strip()
                            expense_type_raw = expense_type_ids[index]
                            realized_amount = _parse_float(realized_amounts[index])
                            if not activity_name or not bku_number or not account_code or not expense_type_raw.isdigit():
                                raise ValueError(f"Data wajib pada step {index + 1} belum lengkap.")
                            if account_code not in active_account_codes:
                                raise ValueError(f"Kode rekening pada step {index + 1} tidak terdaftar atau tidak aktif.")
                            if realized_amount <= 0:
                                raise ValueError(f"Nilai realisasi pada step {index + 1} harus lebih dari 0.")

                            vendor_id, vendor_name, vendor_error = _resolve_school_report_vendor(
                                vendor_ids[index]
                            )
                            if vendor_error:
                                raise ValueError(f"Step {index + 1}: {vendor_error}")

                            adjusted_transactions.append(
                                {
                                    **source_item,
                                    "activity_name": activity_name,
                                    "expense_type_id": int(expense_type_raw),
                                    "bku_number": bku_number,
                                    "account_code": account_code,
                                    "realized_amount": realized_amount,
                                    "vendor_id": vendor_id,
                                    "vendor_name": vendor_name,
                                    "item_name": item_names[index].strip(),
                                }
                            )

                        result = queries.claim_bop_transactions(
                            report["id"], user["id"], adjusted_transactions
                        )
                        if result["inserted"] or result["updated"]:
                            action_parts = []
                            if result["inserted"]:
                                action_parts.append(f"{result['inserted']} transaksi berhasil diklaim")
                            if result["updated"]:
                                action_parts.append(f"{result['updated']} transaksi berhasil diperbarui")
                            flash(
                                f"{' dan '.join(action_parts)} untuk {claim['school_name']}.",
                                "success",
                            )
                        else:
                            flash("Seluruh transaksi BOP sekolah ini sudah pernah diklaim.", "info")
                except (OSError, ValueError) as exc:
                    flash(str(exc), "warning")
            return redirect(
                url_for(
                    "monev_bos.sekolah_activities",
                    period_id=period_id,
                    fund_source="BOP",
                )
            )

        if action == "update_receipts":
            if report and report["status"] not in ["draft", "needs_revision"]:
                flash("Laporan sudah tidak bisa diubah.", "warning")
                return redirect(url_for("monev_bos.sekolah_activities", period_id=period_id, fund_source=fund_source))
            
            bosp = _parse_float(request.form.get("bosp_amount"))
            bop = _parse_float(request.form.get("bop_amount"))
            queries.save_school_report_receipts(user["id"], active_period["id"], bosp, bop)
            flash(f"Total Penerimaan Dana {fund_source} berhasil diperbarui.", "success")
            return redirect(url_for("monev_bos.sekolah_activities", period_id=period_id, fund_source=fund_source))

        if action == "add_activity":
            # Tambahan dari modal selalu masuk ke sumber dana halaman yang sedang dibuka.
            target_fund_source = fund_source
            bku_number, bku_error = _school_bku_number(request.form.get("bku_number"))
            if bku_error:
                flash(bku_error, "warning")
                return redirect(url_for("monev_bos.sekolah_activities", period_id=period_id, fund_source=fund_source))
            account_code_data = _activity_account_code_data(request.form)
            vendor_ids, vendor_names, vendor_error = _resolve_school_report_vendors(
                request.form.getlist("vendor_id")
            )
            if vendor_error:
                flash(vendor_error, "warning")
                return redirect(url_for("monev_bos.sekolah_activities", period_id=period_id, fund_source=fund_source))

            expense_type_id_raw = request.form.get("expense_type_id")
            expense_type_id = int(expense_type_id_raw) if expense_type_id_raw and expense_type_id_raw.isdigit() else None

            data = {
                "activity_code": request.form.get("activity_code") or bku_number,
                "activity_name": request.form.get("activity_name"),
                **account_code_data,
                "expense_type_id": expense_type_id,
                "realized_amount": _parse_float(request.form.get("realized_amount")),
                "vendor_id": vendor_ids[0] if vendor_ids else None,
                "vendor_name": ", ".join(vendor_names) or None,
                "bku_number": bku_number,
                "item_name": request.form.get("item_name"),
                "item_specs": request.form.get("item_specs"),
                "item_quantity": int(request.form.get("item_quantity", 0) or 0)
            }
            duplicate_matches = queries.find_activity_duplicate_matches_for_data(
                report["id"], target_fund_source, data
            )
            duplicate_confirmation = " ".join(
                (request.form.get("duplicate_confirmation") or "").casefold().split()
            )
            if duplicate_matches and duplicate_confirmation != "kegiatan berbeda":
                flash(
                    "Kegiatan terindikasi ganda pada halaman ini. Periksa data pembanding lalu ketik tepat 'kegiatan berbeda' untuk melanjutkan.",
                    "warning",
                )
                return redirect(url_for(
                    "monev_bos.sekolah_activities",
                    period_id=period_id,
                    fund_source=fund_source,
                ))

            field_photo_files, field_photo_data_items, story_post_ids = _requested_activity_photos()
            requested_photo_count = len(field_photo_files) + len(field_photo_data_items) + len(story_post_ids)
            if requested_photo_count > MAX_ACTIVITY_FIELD_PHOTOS:
                flash("Maksimal 3 Foto Kegiatan/Barang per kegiatan, termasuk tautan Foto Live.", "warning")
                return redirect(url_for("monev_bos.sekolah_activities", period_id=period_id, fund_source=fund_source))

            activity_id = queries.create_activity(report["id"], target_fund_source, data)
            queries.set_activity_vendors(activity_id, vendor_ids)
            
            # Handle optional document uploads (Faktur/Kwitansi & Bukti Transfer)
            base_dir = os.path.join(monev_bos_bp.root_path, "..", "static", "uploads")

            # Handle optional camera photo or file upload (Foto Kegiatan / Barang)
            for field_photo_file in field_photo_files:
                sub_path = f"monev_bos/{report['id']}/{activity_id}/field_photo"
                saved_path, err_msg = _save_uploaded_file(field_photo_file, base_dir, sub_path, max_size_bytes=100 * 1024)
                if err_msg:
                    queries.delete_activity(activity_id)
                    flash(err_msg, "danger")
                    return redirect(url_for("monev_bos.sekolah_activities", period_id=period_id, fund_source=fund_source))
                if saved_path:
                    queries.add_activity_doc(activity_id, "field_photo", saved_path, field_photo_file.content_length or 0, user["id"])
            for field_photo_data in field_photo_data_items:
                saved_path, err_msg = _save_camera_photo(field_photo_data, base_dir, f"monev_bos/{report['id']}/{activity_id}/field_photo")
                if err_msg:
                    queries.delete_activity(activity_id)
                    flash(err_msg, "danger")
                    return redirect(url_for("monev_bos.sekolah_activities", period_id=period_id, fund_source=fund_source))
                queries.add_activity_doc(activity_id, "field_photo", saved_path, 0, user["id"])
            if story_post_ids:
                try:
                    for story_post_id in story_post_ids:
                        queries.link_post_to_activity(activity_id, story_post_id, int(user["id"]), int(user["id"]))
                except ValueError as exc:
                    queries.delete_activity(activity_id)
                    flash(str(exc), "danger")
                    return redirect(url_for("monev_bos.sekolah_activities", period_id=period_id, fund_source=fund_source))

            # Handle optional document file uploads
            has_error = False
            for doc_type in ["invoice", "transfer"]:
                file = request.files.get(f"doc_{doc_type}")
                if file and file.filename != '':
                    sub_path = f"monev_bos/{report['id']}/{activity_id}/{doc_type}"
                    saved_path, err_msg = _save_uploaded_file(file, base_dir, sub_path, max_size_bytes=100 * 1024)
                    if err_msg:
                        flash(err_msg, "danger")
                        has_error = True
                    elif saved_path:
                        queries.add_activity_doc(activity_id, doc_type, saved_path, file.content_length or 0, user["id"])

            if not has_error:
                flash("Kegiatan berhasil ditambahkan.", "success")
            
        elif action == "move_activity_fund_source":
            activity_id = request.form.get("activity_id", type=int)
            target_source = (request.form.get("target_fund_source") or "").upper()
            activity = queries.get_activity_by_id(activity_id) if activity_id else None

            if report["status"] not in ["draft", "needs_revision"]:
                flash("Laporan sudah tidak bisa diubah.", "warning")
            elif target_source not in ["BOS", "BOP"]:
                flash("Tujuan migrasi sumber dana tidak valid.", "warning")
            elif not activity or activity.get("report_id") != report["id"]:
                flash("Kegiatan tidak ditemukan pada laporan sekolah ini.", "danger")
            elif activity.get("status") == "valid":
                flash("Kegiatan berstatus Sesuai harus melalui pengajuan perubahan terlebih dahulu.", "warning")
            elif activity.get("fund_source") == target_source:
                flash(f"Kegiatan sudah berada pada sumber dana {target_source}.", "info")
            elif queries.move_activity_fund_source(
                activity_id, report["id"], target_source, user["id"]
            ):
                flash(f"Kegiatan berhasil dipindahkan ke {target_source}.", "success")
                return redirect(url_for(
                    "monev_bos.sekolah_activities",
                    period_id=period_id,
                    fund_source=target_source,
                ))
            else:
                flash("Kegiatan gagal dipindahkan.", "danger")

        elif action == "delete_activity":
            activity_id = request.form.get("activity_id", type=int)
            activity = queries.get_activity_by_id(activity_id) if activity_id else None
            mutation_error = _school_activity_mutation_error(report, activity)
            if mutation_error:
                flash(mutation_error, "warning")
            else:
                queries.delete_activity(activity_id)
                flash("Kegiatan berhasil dihapus.", "success")

        elif action == "request_edit":
            # Pengajuan reopen edit untuk kegiatan yang berstatus Sesuai (valid)
            activity_id = int(request.form.get("activity_id"))
            reason = request.form.get("reason", "").strip()
            act = queries.get_activity_by_id(activity_id)
            if not act:
                flash("Kegiatan tidak ditemukan.", "danger")
            elif not reason:
                flash("Alasan pengajuan reopen wajib diisi.", "warning")
            else:
                # Cancel any previous pending request
                queries.cancel_edit_request(activity_id)
                queries.create_edit_request(activity_id, user["id"], reason)
                
                # Fetch admin/coordinator info for school's kecamatan
                admin_info = queries.get_school_kecamatan_and_admin_wa(user["id"])
                if admin_info.get("admin_phone"):
                    import urllib.parse
                    wa_msg = (
                        f"Yth. Bapak/Ibu {admin_info['admin_name']} (Admin Wilayah Kecamatan {admin_info['kecamatan_name']}),\n\n"
                        f"Saya dari {admin_info['school_name']} mengajukan *Reopen Edit Kegiatan Monev BOS/BOP*:\n"
                        f"• Kode Kegiatan: {act['activity_code']}\n"
                        f"• Uraian: {act['activity_name']}\n"
                        f"• Nominal Realisasi: Rp {act.get('realized_amount', 0):,.0f}\n"
                        f"• Alasan Reopen: \"{reason}\"\n\n"
                        f"Mohon kesediaannya untuk meninjau dan membuka kunci edit kegiatan tersebut. Terima kasih."
                    )
                    wa_url = f"https://wa.me/{admin_info['admin_phone']}?text={urllib.parse.quote(wa_msg)}"
                    session["wa_reopen_prompt"] = {
                        "wa_url": wa_url,
                        "admin_name": admin_info["admin_name"],
                        "kecamatan_name": admin_info["kecamatan_name"],
                        "activity_code": act["activity_code"]
                    }
                
                flash("Pengajuan reopen edit berhasil dikirim ke admin. Silakan konfirmasi ke Admin Wilayah via WhatsApp.", "success")

        elif action == "edit_activity":
            activity_id = request.form.get("activity_id", type=int)
            act = queries.get_activity_by_id(activity_id) if activity_id else None
            mutation_error = _school_activity_mutation_error(report, act)
            if mutation_error:
                flash(mutation_error, "warning")
                return redirect(url_for("monev_bos.sekolah_activities", period_id=period_id, fund_source=fund_source))

            bku_number, bku_error = _school_bku_number(request.form.get("bku_number"))
            if bku_error:
                flash(bku_error, "warning")
                return redirect(url_for("monev_bos.sekolah_activities", period_id=period_id, fund_source=fund_source))

            field_photo_files, field_photo_data_items, story_post_ids = _requested_activity_photos()
            existing_post_ids = {
                int(link["post_id"]) for link in queries.get_activity_post_links(activity_id)
            }
            new_story_post_ids = [post_id for post_id in story_post_ids if post_id not in existing_post_ids]
            valid_photo_count = queries.count_valid_field_photos(activity_id)
            requested_photo_count = len(field_photo_files) + len(field_photo_data_items) + len(new_story_post_ids)
            if valid_photo_count + requested_photo_count > MAX_ACTIVITY_FIELD_PHOTOS:
                remaining = max(0, MAX_ACTIVITY_FIELD_PHOTOS - valid_photo_count)
                flash(
                    f"Maksimal 3 Foto Kegiatan/Barang yang sah per kegiatan. Kuota tambahan saat ini {remaining} foto.",
                    "warning",
                )
                return redirect(url_for("monev_bos.sekolah_activities", period_id=period_id, fund_source=fund_source))

            was_invalid = act and act.get("status") == "invalid"
            # Jika status revisi (invalid), simpan snapshot data lama dulu sebagai history
            if was_invalid:
                queries.save_activity_history(activity_id, user["id"], reason="Perbaikan data oleh sekolah saat status Revisi")

            vendor_ids, vendor_names, vendor_error = _resolve_school_report_vendors(
                request.form.getlist("vendor_id")
            )
            if vendor_error:
                flash(vendor_error, "warning")
                return redirect(url_for("monev_bos.sekolah_activities", period_id=period_id, fund_source=fund_source))

            expense_type_id_raw = request.form.get("expense_type_id")
            expense_type_id = int(expense_type_id_raw) if expense_type_id_raw and expense_type_id_raw.isdigit() else None
            account_code_data = _activity_account_code_data(request.form)

            data = {
                "activity_code": request.form.get("activity_code") or (act and act.get("activity_code")) or bku_number,
                "activity_name": request.form.get("activity_name"),
                **account_code_data,
                "expense_type_id": expense_type_id,
                "realized_amount": _parse_float(request.form.get("realized_amount")),
                "vendor_id": vendor_ids[0] if vendor_ids else None,
                "vendor_name": ", ".join(vendor_names) or None,
                "bku_number": bku_number,
                "item_name": request.form.get("item_name"),
                "item_specs": request.form.get("item_specs"),
                "item_quantity": int(request.form.get("item_quantity", 0) or 0)
            }
            queries.update_activity(activity_id, data)
            queries.set_activity_vendors(activity_id, vendor_ids)
            
            # Kembalikan status kegiatan dari 'invalid' ke 'pending' (menunggu verifikasi ulang)
            queries.update_activity_audit(activity_id, "pending", act.get("audit_notes") or "")

            # Kirim notifikasi in-app & siapkan WA prompt ke Staff Verifikator jika sebelumnya berstatus revisi
            if was_invalid:
                queries.send_revised_activity_notification_to_staff(activity_id)
                staff_wa_info = queries.get_auditor_staff_wa_for_report(report["id"], report["school_id"], period_id, activity_id=activity_id)
                if staff_wa_info.get("staff_phone"):
                    import urllib.parse
                    wa_msg = (
                        f"Yth. Bapak/Ibu {staff_wa_info['staff_name']} (Tim Verifikator Monev BOS/BOP),\n\n"
                        f"Saya dari {report.get('school_name', 'Sekolah')} telah memperbarui/merevisi data kegiatan:\n"
                        f"• Kode Kegiatan: {act['activity_code']}\n"
                        f"• Uraian: {data['activity_name']}\n"
                        f"• Nominal Realisasi: Rp {data['realized_amount']:,.0f}\n\n"
                        f"Mohon kesediaannya untuk melakukan verifikasi ulang. Terima kasih."
                    )
                    wa_url = f"https://wa.me/{staff_wa_info['staff_phone']}?text={urllib.parse.quote(wa_msg)}"
                    session["wa_revision_prompt"] = {
                        "wa_url": wa_url,
                        "staff_name": staff_wa_info["staff_name"],
                        "activity_code": act["activity_code"]
                    }
            
            # Shared upload directory for optional field photo and document replacements.
            base_dir = os.path.join(monev_bos_bp.root_path, "..", "static", "uploads")

            # Handle optional field photo (file upload or camera)
            for field_photo_file in field_photo_files:
                sub_path = f"monev_bos/{report['id']}/{activity_id}/field_photo"
                saved_path, err_msg = _save_uploaded_file(field_photo_file, base_dir, sub_path, max_size_bytes=100 * 1024)
                if err_msg:
                    flash(err_msg, "danger")
                elif saved_path:
                    queries.add_activity_doc(activity_id, "field_photo", saved_path, field_photo_file.content_length or 0, user["id"])
            for field_photo_data in field_photo_data_items:
                saved_path, err_msg = _save_camera_photo(field_photo_data, base_dir, f"monev_bos/{report['id']}/{activity_id}/field_photo")
                if err_msg:
                    flash(err_msg, "danger")
                    return redirect(url_for("monev_bos.sekolah_activities", period_id=period_id, fund_source=fund_source))
                queries.add_activity_doc(activity_id, "field_photo", saved_path, 0, user["id"])
            if story_post_ids:
                try:
                    for story_post_id in story_post_ids:
                        queries.link_post_to_activity(activity_id, story_post_id, int(user["id"]), int(user["id"]))
                except ValueError as exc:
                    flash(str(exc), "danger")
                    return redirect(url_for("monev_bos.sekolah_activities", period_id=period_id, fund_source=fund_source))

            # Handle optional file re-uploads
            for doc_type in ["transfer", "invoice"]:
                file = request.files.get(f"doc_{doc_type}")
                if file and file.filename != '':
                    sub_path = f"monev_bos/{report['id']}/{activity_id}/{doc_type}"
                    saved_path, err_msg = _save_uploaded_file(file, base_dir, sub_path, max_size_bytes=100 * 1024)
                    
                    if err_msg:
                        flash(err_msg, "danger")
                    elif saved_path:
                        queries.add_activity_doc(activity_id, doc_type, saved_path, file.content_length or 0, user["id"])
            
            flash("Perubahan kegiatan berhasil disimpan. Status kegiatan kini kembali Pending dan notifikasi telah dikirim ke Staff Verifikator.", "success")
            
        elif action == "submit_report":
            queries.submit_school_report(report["id"])
            flash("Laporan berhasil disubmit ke tim monev.", "success")
            return redirect(url_for("monev_bos.sekolah_dashboard"))

        return redirect(url_for("monev_bos.sekolah_activities", period_id=period_id, fund_source=fund_source))

    all_activities = queries.list_activities(report["id"])
    activities = [a for a in all_activities if a["fund_source"] == fund_source]
    # Halaman sekolah pertama kali mengikuti urutan kegiatan ditambahkan.
    # Pengurutan BKU diterapkan di browser hanya ketika dipilih pengguna.
    activities.sort(key=lambda activity: int(activity.get("id") or 0))

    bop_claim = None
    bop_claim_supported = fund_source == "BOP" and is_bop_claim_period(
        active_period["year"], active_period["tw"]
    )
    if bop_claim_supported:
        try:
            identity = queries.get_school_claim_identity(user["id"])
            claim = get_school_bop_claim(
                identity.get("npsn"), active_period["year"], active_period["tw"]
            )
            if claim:
                existing_by_code = {
                    activity.get("activity_code"): activity for activity in activities
                }
                claim_transactions = []
                for source_item in claim["transactions"]:
                    existing = existing_by_code.get(source_item["activity_code"])
                    claim_transactions.append(
                        {
                            **source_item,
                            "source_activity_name": source_item["activity_name"],
                            "source_realized_amount": source_item["realized_amount"],
                            "selected_activity_name": existing.get("activity_name") if existing else "",
                            "selected_expense_type_id": existing.get("expense_type_id") if existing else None,
                            "selected_bku_number": existing.get("bku_number") if existing else "",
                            "selected_account_code": existing.get("account_code") if existing else source_item["account_code"],
                            "selected_realized_amount": existing.get("realized_amount") if existing else source_item["realized_amount"],
                            "selected_vendor_id": existing.get("vendor_id") if existing else None,
                            "selected_item_name": existing.get("item_name") if existing else "",
                            "selected_status": existing.get("status") if existing else None,
                            "is_claimed": bool(existing),
                        }
                    )
                pending_transactions = [
                    item for item in claim_transactions if not item["is_claimed"]
                ]
                is_reclaim = not pending_transactions
                reclaim_transactions = [
                    item for item in claim_transactions if item.get("selected_status") != "valid"
                ]
                wizard_transactions = reclaim_transactions if is_reclaim else pending_transactions
                bop_claim = {
                    **claim,
                    "pending_transactions": pending_transactions,
                    "pending_count": len(pending_transactions),
                    "pending_amount": sum(item["realized_amount"] for item in pending_transactions),
                    "claimed_count": len(claim["transactions"]) - len(pending_transactions),
                    "is_reclaim": is_reclaim,
                    "wizard_transactions": wizard_transactions,
                    "wizard_count": len(wizard_transactions),
                    "wizard_amount": sum(item["source_realized_amount"] for item in wizard_transactions),
                    "reclaim_blocked_count": len(claim_transactions) - len(reclaim_transactions),
                }
        except (OSError, ValueError):
            current_app.logger.exception("Failed to load repository-backed BOP claim data")
    
    total_receipt = report["bosp_receipt_amount"] if fund_source == "BOS" else report["bop_receipt_amount"]
    total_realized = sum(a["realized_amount"] for a in activities)
    remaining_balance = total_receipt - total_realized
    
    wa_prompt = session.pop("wa_reopen_prompt", None)
    wa_revision_prompt = session.pop("wa_revision_prompt", None)
    admin_info = queries.get_school_kecamatan_and_admin_wa(user["id"])
    staff_wa_info = queries.get_auditor_staff_wa_for_report(report["id"], report["school_id"], period_id)
    checklist_results_by_activity = queries.get_checklist_results_by_activity_ids(
        [int(act["id"]) for act in activities]
    )

    for act in activities:
        act["account_code_values"] = [
            code.strip()
            for code in str(act.get("account_code") or "").split(",")
            if code.strip()
        ]
        act["checklist_results"] = checklist_results_by_activity.get(int(act["id"]), [])
        act["checklist_notes"] = [
            result
            for result in act["checklist_results"]
            if str(result.get("notes") or "").strip()
        ]
        docs = queries.get_activity_docs(act["id"])
        act["docs"] = {doc["doc_type"]: doc for doc in docs}
        act["field_photos"] = [doc for doc in docs if doc["doc_type"] == "field_photo"]
        act["valid_field_photos"] = [
            doc for doc in act["field_photos"] if doc.get("is_audit_valid", True)
        ]
        act["invalid_field_photos"] = [
            doc for doc in act["field_photos"] if not doc.get("is_audit_valid", True)
        ]
        post_links = queries.get_activity_post_links(act["id"])
        act["linked_story_post_ids"] = [link["post_id"] for link in post_links]
        act["history"] = queries.get_activity_history(act["id"])
        act["pending_edit_request"] = queries.get_edit_request_by_activity(act["id"])
        if act["pending_edit_request"] and admin_info.get("admin_phone"):
            import urllib.parse
            req_reason = act["pending_edit_request"].get("reason") or "-"
            wa_msg = (
                f"Yth. Bapak/Ibu {admin_info['admin_name']} (Admin Wilayah Kecamatan {admin_info['kecamatan_name']}),\n\n"
                f"Saya dari {admin_info['school_name']} mengajukan *Reopen Edit Kegiatan Monev BOS/BOP*:\n"
                f"• Kode Kegiatan: {act['activity_code']}\n"
                f"• Uraian: {act['activity_name']}\n"
                f"• Nominal Realisasi: Rp {act.get('realized_amount', 0):,.0f}\n"
                f"• Alasan Reopen: \"{req_reason}\"\n\n"
                f"Mohon kesediaannya untuk meninjau dan membuka kunci edit kegiatan tersebut. Terima kasih."
            )
            act["wa_url"] = f"https://wa.me/{admin_info['admin_phone']}?text={urllib.parse.quote(wa_msg)}"

        # If activity is pending after revision and staff phone exists
        act_staff_wa = queries.get_auditor_staff_wa_for_report(report["id"], report["school_id"], period_id, activity_id=act["id"])
        if act["status"] == "pending" and act.get("history") and act_staff_wa.get("staff_phone"):
            import urllib.parse
            wa_msg_staff = (
                f"Yth. Bapak/Ibu {act_staff_wa['staff_name']} (Tim Verifikator Monev BOS/BOP),\n\n"
                f"Saya dari {report.get('school_name', 'Sekolah')} telah memperbarui/merevisi data kegiatan:\n"
                f"• Kode Kegiatan: {act['activity_code']}\n"
                f"• Uraian: {act['activity_name']}\n"
                f"• Nominal Realisasi: Rp {act.get('realized_amount', 0):,.0f}\n\n"
                f"Mohon kesediaannya untuk melakukan verifikasi ulang. Terima kasih."
            )
            act["staff_wa_url"] = f"https://wa.me/{act_staff_wa['staff_phone']}?text={urllib.parse.quote(wa_msg_staff)}"

    master_activities = queries.list_master_activities(include_inactive=False, fund_source=fund_source)
    school_story_posts = queries.list_school_posts(school_user_id=int(user["id"]), limit=300)
    expense_types = queries.list_expense_types(include_inactive=False)
    if bop_claim:
        for transaction in bop_claim.get("wizard_transactions", []):
            recommendation = recommend_expense_type(
                transaction.get("source_activity_name"),
                transaction.get("source_realized_amount"),
                expense_types,
            )
            transaction["recommended_expense_type_id"] = recommendation.get("id") if recommendation else None
            transaction["recommended_expense_type_name"] = recommendation.get("name") if recommendation else None
    try:
        active_checklists = queries.list_checklists(include_inactive=False)
        checklist_requirements_by_expense_type = {
            str(expense_type["id"]): [
                checklist["name"]
                for checklist in active_checklists
                if not checklist.get("expense_type_ids")
                or expense_type["id"] in checklist.get("expense_type_ids", [])
            ]
            for expense_type in expense_types
        }
    except Exception:
        current_app.logger.exception("Failed to load checklist requirements by expense type")
        checklist_requirements_by_expense_type = {}
    try:
        account_codes = queries.list_account_codes(include_inactive=False)
    except Exception:
        current_app.logger.exception("Failed to load Monev BOS account codes")
        account_codes = []
    verified_vendors = queries.get_report_selectable_vendors(school_id=int(user["id"]))

    try:
        auditor_team = queries.get_assigned_auditors_for_school(report["school_id"], active_period["id"])
    except Exception:
        current_app.logger.exception("Failed to load assigned Monev BOS verifiers")
        auditor_team = {"team_name": None, "members": []}

    school_info_data = None
    try:
        from dashboard.portal.routes import _fetch_user_school, _normalize_metadata
        school_info_data = _fetch_user_school(user["id"])
        school_meta = _normalize_metadata(school_info_data.get("metadata")) if school_info_data else {}
    except Exception:
        current_app.logger.exception("Failed to load Monev BOS school metadata")
        school_meta = {}
    headmaster_info = {
        "name": school_meta.get("headmaster_name") or "-",
        "nip": school_meta.get("headmaster_nip") or "-"
    }
    school_display_name = (
        (school_info_data or {}).get("name")
        or user.get("full_name")
        or user.get("username")
        or user.get("email")
        or "Sekolah"
    )

    return render_template("monev_bos/sekolah/activities.html", 
                           active_period=active_period, 
                           report=report, 
                           activities=activities,
                           master_activities=master_activities,
                           school_story_posts=school_story_posts,
                           expense_types=expense_types,
                           checklist_requirements_by_expense_type=checklist_requirements_by_expense_type,
                           account_codes=account_codes,
                           verified_vendors=verified_vendors,
                           fund_source=fund_source,
                           total_receipt=total_receipt,
                           total_realized=total_realized,
                           remaining_balance=remaining_balance,
                           wa_prompt=wa_prompt,
                           wa_revision_prompt=wa_revision_prompt,
                           admin_info=admin_info,
                           staff_wa_info=staff_wa_info,
                           auditor_team=auditor_team,
                           headmaster_info=headmaster_info,
                           school_display_name=school_display_name,
                           bop_claim=bop_claim,
                           bop_claim_supported=bop_claim_supported)

@monev_bos_bp.route("/sekolah/activities/export", methods=["GET"])
@role_required("sekolah")
def sekolah_activities_export():
    import io
    from flask import send_file
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from datetime import datetime

    user = current_user()
    period_id = request.args.get("period_id", type=int)
    fund_source = request.args.get("fund_source", "BOS")
    if fund_source not in ["BOS", "BOP"]:
        fund_source = "BOS"

    if not period_id:
        flash("Periode tidak ditemukan.", "warning")
        return redirect(url_for("monev_bos.sekolah_dashboard"))

    active_periods = queries.get_active_periods()
    active_period = next((p for p in active_periods if p["id"] == period_id), None)
    if not active_period:
        flash("Periode tidak ditemukan.", "warning")
        return redirect(url_for("monev_bos.sekolah_dashboard"))

    report = queries.get_school_report(user["id"], active_period["id"])
    if not report:
        flash("Data laporan tidak ditemukan.", "warning")
        return redirect(url_for("monev_bos.sekolah_activities", period_id=period_id, fund_source=fund_source))

    all_activities = queries.list_activities(report["id"])
    activities = [a for a in all_activities if a["fund_source"] == fund_source]

    total_receipt = report["bosp_receipt_amount"] if fund_source == "BOS" else report["bop_receipt_amount"]
    total_realized = sum(a["realized_amount"] for a in activities)
    remaining_balance = total_receipt - total_realized

    wb = Workbook()
    ws = wb.active
    label_fund = "BOSP" if fund_source == "BOS" else "BOP"
    ws.title = f"Laporan {label_fund}"

    title_font = Font(name="Calibri", size=13, bold=True, color="1F4E78")
    header_font = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
    bold_font = Font(name="Calibri", size=10, bold=True)
    normal_font = Font(name="Calibri", size=10)

    header_fill = PatternFill(start_color="1B5E20" if fund_source == "BOS" else "0277BD", end_color="1B5E20" if fund_source == "BOS" else "0277BD", fill_type="solid")
    summary_fill = PatternFill(start_color="F1F8E9" if fund_source == "BOS" else "E0F7FA", end_color="F1F8E9" if fund_source == "BOS" else "E0F7FA", fill_type="solid")

    thin_border = Border(
        left=Side(style='thin', color='D9D9D9'),
        right=Side(style='thin', color='D9D9D9'),
        top=Side(style='thin', color='D9D9D9'),
        bottom=Side(style='thin', color='D9D9D9')
    )

    school_name = user.get("full_name") or user.get("username") or "Sekolah"

    ws["A1"] = f"LAPORAN PERTANGGUNGJAWABAN DANA {label_fund}"
    ws["A1"].font = title_font
    ws["A2"] = f"Nama Sekolah: {school_name}"
    ws["A2"].font = bold_font
    ws["A3"] = f"Periode: Triwulan {active_period['tw']} Tahun {active_period['year']}"
    ws["A3"].font = normal_font
    ws["A4"] = f"Tanggal Unduh: {datetime.now().strftime('%d-%m-%Y %H:%M')}"
    ws["A4"].font = normal_font

    ws["A6"] = f"Total Penerimaan Dana {label_fund}"
    ws["B6"] = total_receipt
    ws["B6"].number_format = '#,##0'

    ws["A7"] = "Total Realisasi Pengeluaran"
    ws["B7"] = total_realized
    ws["B7"].number_format = '#,##0'

    ws["A8"] = "Sisa Saldo"
    ws["B8"] = remaining_balance
    ws["B8"].number_format = '#,##0'

    for row in range(6, 9):
        ws[f"A{row}"].font = bold_font
        ws[f"B{row}"].font = bold_font
        ws[f"A{row}"].fill = summary_fill
        ws[f"B{row}"].fill = summary_fill

    headers = [
        "No",
        "Jenis Belanja",
        "Nama Kegiatan",
        "No. BKU",
        "Kode Rekening",
        "Nama Toko / Vendor",
        "Detail Kegiatan",
        "Nilai Realisasi (Rp)",
        "Status"
    ]

    start_row = 10
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=start_row, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    status_labels = {
        "draft": "Draft",
        "pending": "Menunggu Verifikasi",
        "approved": "Disetujui",
        "needs_revision": "Perlu Perbaikan",
        "rejected": "Ditolak"
    }

    current_row = start_row + 1
    for idx, act in enumerate(activities, 1):
        status_text = status_labels.get(act.get("status", "draft"), act.get("status", "-"))

        ws.cell(row=current_row, column=1, value=idx).alignment = Alignment(horizontal="center")
        ws.cell(row=current_row, column=2, value=act.get("expense_type_name") or "-")
        ws.cell(row=current_row, column=3, value=act.get("activity_name") or "-")
        ws.cell(row=current_row, column=4, value=act.get("bku_number") or act.get("activity_code") or "-").alignment = Alignment(horizontal="center")
        ws.cell(row=current_row, column=5, value=act.get("account_code") or "-").alignment = Alignment(horizontal="center")
        ws.cell(row=current_row, column=6, value=act.get("vendor_display_name") or act.get("vendor_name") or "-")
        ws.cell(row=current_row, column=7, value=act.get("item_name") or act.get("item_specs") or "-")

        realized_cell = ws.cell(row=current_row, column=8, value=float(act.get("realized_amount") or 0))
        realized_cell.number_format = '#,##0'

        ws.cell(row=current_row, column=9, value=status_text).alignment = Alignment(horizontal="center")

        for col_idx in range(1, 10):
            ws.cell(row=current_row, column=col_idx).border = thin_border
            ws.cell(row=current_row, column=col_idx).font = normal_font

        current_row += 1

    ws.cell(row=current_row, column=1, value="")
    ws.cell(row=current_row, column=2, value="")
    ws.cell(row=current_row, column=3, value="")
    ws.cell(row=current_row, column=4, value="")
    ws.cell(row=current_row, column=5, value="")
    ws.cell(row=current_row, column=6, value="")
    total_label_cell = ws.cell(row=current_row, column=7, value="TOTAL REALISASI")
    total_label_cell.font = bold_font
    total_label_cell.alignment = Alignment(horizontal="right")

    total_sum_cell = ws.cell(row=current_row, column=8, value=total_realized)
    total_sum_cell.font = bold_font
    total_sum_cell.number_format = '#,##0'

    ws.cell(row=current_row, column=9, value="")

    for col_idx in range(1, 10):
        cell = ws.cell(row=current_row, column=col_idx)
        cell.border = thin_border
        cell.fill = summary_fill

    # Signature & Verifier Section in Excel Export
    auditor_team = queries.get_assigned_auditors_for_school(report["school_id"], active_period["id"])
    from dashboard.portal.routes import _fetch_user_school, _normalize_metadata
    school_info_data = _fetch_user_school(user["id"])
    school_meta = _normalize_metadata(school_info_data.get("metadata")) if school_info_data else {}
    headmaster_info = {
        "name": school_meta.get("headmaster_name") or "-",
        "nip": school_meta.get("headmaster_nip") or "-"
    }

    sig_start_row = current_row + 3
    date_str = datetime.now().strftime('%d-%m-%Y')
    
    ws.cell(row=sig_start_row, column=6, value=f"Jakarta, {date_str}").font = normal_font
    
    ws.cell(row=sig_start_row + 1, column=2, value="Mengetahui / Menyetujui,").font = normal_font
    ws.cell(row=sig_start_row + 2, column=2, value=f"Kepala {school_name}").font = bold_font
    
    ws.cell(row=sig_start_row + 6, column=2, value=f"( {headmaster_info['name']} )").font = bold_font
    ws.cell(row=sig_start_row + 7, column=2, value=f"NIP. {headmaster_info['nip']}").font = normal_font
    
    team_title = f"Tim Verifikator Monev ({auditor_team['team_name']}):" if auditor_team.get("team_name") else "Tim Verifikator Monev:"
    ws.cell(row=sig_start_row + 1, column=6, value=team_title).font = bold_font
    
    auditor_members = auditor_team.get("members") or []
    if auditor_members:
        for idx, m in enumerate(auditor_members, 1):
            role_badge = " (Ketua Tim)" if m.get("is_leader") else ""
            nip_str = f" - NIP. {m['staff_nip']}" if m.get("staff_nip") else ""
            ws.cell(row=sig_start_row + 1 + idx, column=6, value=f"{idx}. {m['staff_name']}{role_badge}{nip_str}").font = normal_font
    else:
        ws.cell(row=sig_start_row + 2, column=6, value="- Tim Monev Wilayah").font = normal_font

    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            val = str(cell.value or "")
            if len(val) > max_len:
                max_len = len(val)
        ws.column_dimensions[col_letter].width = min(max(max_len + 3, 12), 50)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    clean_school = "".join([c for c in school_name if c.isalnum() or c in [' ', '_']]).rstrip().replace(" ", "_")
    filename = f"Laporan_Monev_{label_fund}_{clean_school}_TW{active_period['tw']}_{active_period['year']}.xlsx"

    return send_file(
        output,
        as_attachment=True,
        download_name=filename,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

@monev_bos_bp.route("/staff")
@role_required("staff", "admin")
def staff_dashboard():
    user = current_user()
    periods = queries.list_periods()
    requested_period_id = request.args.get("period_id", type=int)
    requested_year = request.args.get("year", type=int)
    requested_tw = request.args.get("tw", type=int)

    active_period = None
    if requested_period_id:
        active_period = next(
            (period for period in periods if int(period["id"]) == requested_period_id),
            None,
        )
    if active_period is None and requested_year:
        year_periods = [
            period for period in periods if int(period["year"]) == requested_year
        ]
        if requested_tw:
            active_period = next(
                (
                    period
                    for period in year_periods
                    if int(period["tw"]) == requested_tw
                ),
                None,
            )
        if active_period is None and year_periods:
            active_period = max(year_periods, key=lambda period: int(period["tw"]))
    if active_period is None:
        default_active_period = queries.get_active_period()
        if default_active_period:
            active_period = next(
                (
                    period
                    for period in periods
                    if int(period["id"]) == int(default_active_period["id"])
                ),
                default_active_period,
            )
        elif periods:
            active_period = max(
                periods,
                key=lambda period: (int(period["year"]), int(period["tw"])),
            )

    available_years = sorted(
        {int(period["year"]) for period in periods}, reverse=True
    )
    selected_year_periods = (
        [
            period
            for period in periods
            if int(period["year"]) == int(active_period["year"])
        ]
        if active_period
        else []
    )
    teams = queries.get_teams_for_staff(user["id"])
    
    assigned_schools = []
    if active_period and teams:
        # Untuk simple MVP, ambil sekolah dari tim pertama
        # Idealnya bisa pilih tim jika staff ikut banyak tim
        team_id = teams[0]["id"] 
        assigned_schools = queries.get_schools_for_team(team_id, active_period["id"])
        
    return render_template("monev_bos/staff/dashboard.html", 
                           active_period=active_period, 
                           available_years=available_years,
                           selected_year_periods=selected_year_periods,
                           teams=teams, 
                           assigned_schools=assigned_schools)

@monev_bos_bp.route("/staff/my-team", methods=["GET", "POST"])
@role_required("staff", "admin")
def staff_my_team():
    user = current_user()
    teams = queries.get_teams_for_staff(user["id"])
    
    if not teams:
        flash("Anda belum dimasukkan ke tim manapun.", "warning")
        return redirect(url_for("monev_bos.staff_dashboard"))
        
    team = teams[0] # Ambil tim pertama
    
    if request.method == "POST":
        if not team["is_leader"]:
            flash("Hanya ketua tim yang bisa mengelola anggota.", "danger")
            return redirect(url_for("monev_bos.staff_my_team"))
            
        action = request.form.get("action")
        staff_id = int(request.form.get("staff_id"))
        
        if action == "add_member":
            queries.add_team_member(team["id"], staff_id)
            _record_monev_admin_action(
                "ADD_MEMBER",
                "MONEV_TEAM_MEMBER",
                target_id=staff_id,
                metadata={"team_id": team["id"]},
            )
            flash("Anggota berhasil ditambahkan.", "success")
        elif action == "remove_member":
            queries.remove_team_member(team["id"], staff_id)
            _record_monev_admin_action(
                "REMOVE_MEMBER",
                "MONEV_TEAM_MEMBER",
                target_id=staff_id,
                metadata={"team_id": team["id"]},
            )
            flash("Anggota berhasil dihapus.", "success")
            
        return redirect(url_for("monev_bos.staff_my_team"))

    members = queries.get_team_members(team["id"])
    all_staff = queries.get_staff_users() if team["is_leader"] else []
    
    # Filter out existing members
    member_ids = [m["id"] for m in members]
    available_staff = [s for s in all_staff if s["id"] not in member_ids]
    
    return render_template("monev_bos/staff/my_team.html", team=team, members=members, available_staff=available_staff)

@monev_bos_bp.route("/staff/audit/<int:report_id>", methods=["GET", "POST"])
@monev_bos_bp.route("/staff/verifikasi/<int:report_id>", methods=["GET", "POST"])
@role_required("staff", "admin")
def staff_audit_report(report_id):
    user = current_user()
    report = queries.get_report_by_id(report_id)
    
    if not report:
        flash("Laporan tidak ditemukan.", "danger")
        return redirect(url_for("monev_bos.staff_dashboard"))
        
    if request.method == "POST":
        action = request.form.get("action")
        if action == "update_report_status":
            status = request.form.get("status")
            if status not in REPORT_AUDIT_STATUSES:
                flash("Status laporan tidak valid.", "warning")
                return redirect(url_for("monev_bos.staff_audit_report", report_id=report_id))
            with queries.get_cursor(commit=True) as cur:
                cur.execute(
                    "UPDATE monev_bos_reports SET status = %s, updated_at = NOW() WHERE id = %s",
                    (status, report_id)
                )
            queries.add_audit_log(report_id, None, user["id"], "UPDATE_STATUS", f"Mengubah status laporan menjadi {status}")
            _record_monev_admin_action(
                "UPDATE_STATUS",
                "MONEV_REPORT",
                target_id=report_id,
                target_name=report.get("school_name"),
                metadata={"status": status},
            )
            flash(f"Status laporan diubah menjadi {status}", "success")
            return redirect(url_for("monev_bos.staff_audit_report", report_id=report_id))
    # Reset any stale in_review activity statuses back to pending when loading page
    with queries.get_cursor(commit=True) as cur:
        cur.execute("UPDATE monev_bos_activities SET status = 'pending' WHERE report_id = %s AND status = 'in_review'", (report_id,))

    fund_source = request.args.get("fund_source", "BOS").upper()
    if fund_source not in ["BOS", "BOP"]:
        fund_source = "BOS"

    checklists = queries.list_checklists(include_inactive=False)
    all_activities = queries.list_activities(report_id)
    activities = [act for act in all_activities if (act.get("fund_source") or "BOS").upper() == fund_source]
    checklist_yes = 0
    checklist_total = 0
    activities_with_school_photo = 0
    for act in activities:
        act["vendor_required_missing"] = not _activity_has_vendor(act)
        act["vendor_verification_blocked"] = _activity_vendor_is_unverified(act)
        docs = queries.get_activity_docs(act["id"])
        act["docs"] = {doc["doc_type"]: doc for doc in docs}
        act["school_photos"] = [doc for doc in docs if doc["doc_type"] == "field_photo"]
        act["valid_school_photos"] = [
            doc for doc in act["school_photos"] if doc.get("is_audit_valid", True)
        ]
        act["invalid_school_photos"] = [
            doc for doc in act["school_photos"] if not doc.get("is_audit_valid", True)
        ]
        act["staff_photos"] = [doc for doc in docs if doc["doc_type"] == "live_photo"]
        for photo in act["staff_photos"]:
            photo["can_delete"] = True
        act["live_photos"] = [doc for doc in docs if doc["doc_type"] in ["live_photo", "field_photo"]]
        act["has_school_live_photo"] = bool(act["valid_school_photos"])
        if act["has_school_live_photo"]:
            activities_with_school_photo += 1
        # get checklist results for this activity
        act["checklist_results"] = queries.get_activity_checklist_results(act["id"])
        act["checklists"] = queries.get_checklists_for_activity(act.get("expense_type_id"))
        act["checklist_total"] = len(act["checklists"])
        act["checklist_yes"] = sum(
            1
            for checklist in act["checklists"]
            if (act["checklist_results"].get(checklist["id"]) or {}).get("status") == "yes"
        )
        act["checklist_no"] = act["checklist_total"] - act["checklist_yes"]
        act["checklist_percent"] = (
            act["checklist_yes"] / act["checklist_total"] * 100
            if act["checklist_total"] else 0
        )
        checklist_yes += act["checklist_yes"]
        checklist_total += act["checklist_total"]

    audit_metrics = {
        "checklist_yes": checklist_yes,
        "checklist_no": checklist_total - checklist_yes,
        "checklist_total": checklist_total,
        "checklist_percent": (checklist_yes / checklist_total * 100) if checklist_total else 0,
        "activities_with_school_photo": activities_with_school_photo,
        "activities_without_school_photo": len(activities) - activities_with_school_photo,
        "activities_total": len(activities),
        "school_photo_percent": (
            activities_with_school_photo / len(activities) * 100 if activities else 0
        ),
    }

    activity_groups_by_name = {}
    for act in activities:
        name = (act.get("activity_name") or "Tanpa nama kegiatan").strip()
        group = activity_groups_by_name.setdefault(name, {
            "name": name,
            "total": 0,
            "pending": 0,
            "in_review": 0,
            "valid": 0,
            "invalid": 0,
            "ready_review": 0,
            "ready_submit": 0,
        })
        group["total"] += 1
        status = act.get("status") or "pending"
        if status in group:
            group[status] += 1

    activity_groups = list(activity_groups_by_name.values())
        
    audit_logs = queries.get_audit_logs(report_id)
    
    total_receipt = float(report.get("bosp_receipt_amount") or 0) if fund_source == "BOS" else float(report.get("bop_receipt_amount") or 0)
    total_realized = sum(float(act.get("realized_amount") or 0) for act in activities)
    remaining_balance = total_receipt - total_realized
    percent_spent = (total_realized / total_receipt * 100) if total_receipt > 0 else 0
    
    return render_template(
        "monev_bos/staff/audit_report.html",
        report=report,
        fund_source=fund_source,
        activities=activities,
        activity_groups=activity_groups,
        audit_logs=audit_logs,
        checklists=checklists,
        audit_metrics=audit_metrics,
        total_receipt=total_receipt,
        total_realized=total_realized,
        remaining_balance=remaining_balance,
        percent_spent=percent_spent,
    )

@monev_bos_bp.route("/staff/audit/activity/<int:activity_id>", methods=["POST"])
@monev_bos_bp.route("/staff/verifikasi/kegiatan/<int:activity_id>", methods=["POST"])
@role_required("staff", "admin")
def staff_audit_activity(activity_id):
    user = current_user()
    act = queries.get_activity_by_id(activity_id)
    
    if not act:
        if request.headers.get("X-Requested-With") == "XMLHttpRequest" or request.headers.get("Accept") == "application/json":
            return jsonify({"success": False, "message": "Kegiatan tidak ditemukan."}), 404
        flash("Kegiatan tidak ditemukan.", "danger")
        return redirect(url_for("monev_bos.staff_dashboard"))

    # The activity is authoritative; never trust a submitted report ID for audit mutations.
    report_id = int(act.get("report_id") or 1)
    
    action = request.form.get("action")
    
    try:
        if action == "check_vendor_status":
            has_vendor = _activity_has_vendor(act)
            vendor_verified = has_vendor and not _activity_vendor_is_unverified(act)
            unverified_names = (
                act.get("unverified_vendor_names")
                or act.get("vendor_display_name")
                or act.get("vendor_name")
                or "Vendor / narasumber terkait"
            )
            return jsonify({
                "success": True,
                "activity_id": activity_id,
                "vendor_verified": vendor_verified,
                "unverified_names": "" if vendor_verified else unverified_names,
                "message": (
                    "Vendor / narasumber sudah terverifikasi. Tombol Sesuai (Valid) kini aktif."
                    if vendor_verified
                    else (
                        "Vendor / narasumber wajib diisi sebelum kegiatan dapat dinyatakan Sesuai (Valid)."
                        if not has_vendor
                        else f"{unverified_names} masih belum terverifikasi."
                    )
                ),
            })

        if action in {"annul_school_photo", "restore_school_photo"}:
            doc_id_raw = request.form.get("doc_id")
            try:
                doc_id = int(doc_id_raw)
            except (TypeError, ValueError):
                return jsonify({"success": False, "message": "Foto sekolah tidak valid."}), 400

            is_valid = action == "restore_school_photo"
            reason = (request.form.get("reason") or "").strip()
            if not is_valid and not reason:
                return jsonify({"success": False, "message": "Alasan anulir wajib diisi."}), 400

            try:
                updated_photo = queries.set_field_photo_audit_validity(
                    activity_id,
                    doc_id,
                    is_valid,
                    int(user["id"]),
                    reason if not is_valid else "Disahkan kembali oleh verifikator",
                )
            except ValueError as exc:
                return jsonify({"success": False, "message": str(exc)}), 400
            if not updated_photo:
                return jsonify({"success": False, "message": "Foto sekolah tidak ditemukan."}), 404

            audit_action = "RESTORE_SCHOOL_PHOTO" if is_valid else "ANNUL_SCHOOL_PHOTO"
            audit_detail = (
                "Mengesahkan kembali foto kegiatan/barang sekolah"
                if is_valid
                else f"Menganulir foto kegiatan/barang sekolah: {reason}"
            )
            queries.add_audit_log(report_id, activity_id, user["id"], audit_action, audit_detail)
            _record_monev_admin_action(
                audit_action,
                "MONEV_ACTIVITY_PHOTO",
                target_id=doc_id,
                target_name=act.get("activity_name"),
                metadata={"report_id": report_id, "activity_id": activity_id, "reason": reason},
            )
            return jsonify({
                "success": True,
                "is_valid": is_valid,
                "message": "Foto kembali disahkan." if is_valid else "Foto berhasil dianulir dan tidak dihitung.",
            })

        if action == "delete_staff_photo":
            doc_id = request.form.get("doc_id", type=int)
            if not doc_id:
                return jsonify({"success": False, "message": "Foto staff tidak valid."}), 400

            deleted_photo = queries.delete_staff_live_photo(
                activity_id,
                doc_id,
                None,
            )
            if not deleted_photo:
                return jsonify({
                    "success": False,
                    "message": "Foto staff tidak ditemukan.",
                }), 404

            photo_path = deleted_photo.get("file_path") or ""
            if photo_path.startswith("static/uploads/monev_bos/"):
                try:
                    os.remove(_absolute_dashboard_file_path(photo_path))
                except FileNotFoundError:
                    pass
                except OSError:
                    current_app.logger.exception("Failed to remove staff live photo %s", photo_path)

            queries.add_audit_log(
                report_id,
                activity_id,
                user["id"],
                "DELETE_PHOTO",
                "Menghapus foto live lapangan yang diambil staff",
            )
            return jsonify({"success": True, "message": "Foto staff berhasil dihapus."})

        if action == "autosave_details":
            notes = request.form.get("audit_notes") or ""
            queries.update_activity_audit_notes(activity_id, notes)

            # Save exactly the active checklist fields rendered and submitted by
            # the form. Some legacy activities have no/mismatched expense type,
            # so filtering again here can silently discard a visible choice.
            checklists = queries.list_checklists(include_inactive=False)
            saved_items = 0
            for cl in checklists:
                cl_status = request.form.get(f"checklist_{cl['id']}")
                cl_notes = request.form.get(f"checklist_notes_{cl['id']}") or ""
                if cl_status in {"yes", "no", "na"}:
                    queries.save_checklist_result(
                        activity_id, cl["id"], cl_status, cl_notes, user["id"]
                    )
                    saved_items += 1

            return jsonify({
                "success": True,
                "activity_id": activity_id,
                "saved_items": saved_items,
                "message": "Perubahan verifikasi tersimpan otomatis.",
            })

        if action == "validate":
            status = request.form.get("status")
            if not status or status not in ["valid", "invalid", "in_review", "pending"]:
                if request.headers.get("X-Requested-With") == "XMLHttpRequest" or request.headers.get("Accept") == "application/json":
                    return jsonify({"success": False, "message": "Pilih status verifikasi yang sesuai."}), 400
                flash("Pilih status verifikasi yang valid (Sesuai atau Perlu Revisi).", "warning")
                return redirect(url_for("monev_bos.staff_audit_report", report_id=report_id))

            if status == "valid" and _activity_vendor_is_unverified(act):
                if not _activity_has_vendor(act):
                    msg = "Kegiatan tidak dapat dinyatakan Sesuai karena vendor / narasumber wajib diisi dan harus sudah terverifikasi."
                else:
                    vendor_name_disp = act.get("unverified_vendor_names") or act.get("vendor_display_name") or act.get("vendor_name") or "terkait"
                    msg = f"Kegiatan tidak dapat dinyatakan Sesuai karena vendor / narasumber '{vendor_name_disp}' belum terverifikasi. Silakan verifikasi terlebih dahulu."
                if request.headers.get("X-Requested-With") == "XMLHttpRequest" or request.headers.get("Accept") == "application/json":
                    return jsonify({"success": False, "message": msg}), 400
                flash(msg, "warning")
                return redirect(url_for("monev_bos.staff_audit_report", report_id=report_id))

            notes = (request.form.get("audit_notes") or "").strip()
            if status == "invalid" and not notes:
                message = "Catatan revisi wajib diisi sebelum kegiatan ditandai Perlu Revisi."
                if request.headers.get("X-Requested-With") == "XMLHttpRequest" or request.headers.get("Accept") == "application/json":
                    return jsonify({"success": False, "message": message}), 400
                flash(message, "warning")
                return redirect(url_for("monev_bos.staff_audit_report", report_id=report_id))

            queries.update_activity_audit(activity_id, status, notes)
            
            # Save active checklist fields submitted by the rendered form.
            checklists = queries.list_checklists(include_inactive=False)
            for cl in checklists:
                cl_status = request.form.get(f"checklist_{cl['id']}")
                cl_notes = request.form.get(f"checklist_notes_{cl['id']}") or ""
                if cl_status and cl_status in ["yes", "no", "na"]:
                    queries.save_checklist_result(activity_id, cl['id'], cl_status, cl_notes, user["id"])
                    
            status_label = "Sesuai" if status == "valid" else ("Tidak Sesuai (Revisi)" if status == "invalid" else ("Proses Verifikasi" if status == "in_review" else "Pending"))
            queries.add_audit_log(report_id, activity_id, user["id"], "VALIDATE", f"Memverifikasi kegiatan dengan status '{status_label}'" + (f": {notes}" if notes else ""))
            _record_monev_admin_action(
                "VALIDATE",
                "MONEV_ACTIVITY",
                target_id=activity_id,
                target_name=act.get("activity_name"),
                metadata={"report_id": report_id, "status": status, "has_notes": bool(notes)},
            )
            
            if request.headers.get("X-Requested-With") == "XMLHttpRequest" or request.headers.get("Accept") == "application/json":
                return jsonify({
                    "success": True,
                    "status": status,
                    "status_label": status_label,
                    "activity_id": activity_id,
                    "auditor_name": user.get("full_name") or user.get("email") or "Staff",
                    "message": f"Hasil verifikasi kegiatan berhasil disimpan ({status_label})."
                })

            flash("Hasil verifikasi kegiatan berhasil disimpan.", "success")
            return redirect(url_for("monev_bos.staff_audit_report", report_id=report_id))

        elif action == "start_audit":
            if act and act.get("status") in ["pending", "invalid"]:
                queries.update_activity_audit(activity_id, "in_review", act.get("audit_notes") or "")
                _record_monev_admin_action(
                    "START_AUDIT",
                    "MONEV_ACTIVITY",
                    target_id=activity_id,
                    target_name=act.get("activity_name"),
                    metadata={"report_id": report_id, "previous_status": act.get("status")},
                )
                staff_name = user.get("full_name") or user.get("email") or "Staff"
                return jsonify({"success": True, "status": "in_review", "original_status": act.get("status"), "message": f"Status kegiatan diubah ke Proses Verifikasi oleh {staff_name}"})
            return jsonify({"success": True, "status": act.get("status") if act else "pending"})

        elif action == "cancel_audit":
            target_status = request.form.get("original_status") or "pending"
            if target_status not in ["pending", "invalid"]:
                target_status = "pending"
            if act and act.get("status") == "in_review":
                queries.update_activity_audit(activity_id, target_status, act.get("audit_notes") or "")
                _record_monev_admin_action(
                    "CANCEL_AUDIT",
                    "MONEV_ACTIVITY",
                    target_id=activity_id,
                    target_name=act.get("activity_name"),
                    metadata={"report_id": report_id, "restored_status": target_status},
                )
                return jsonify({"success": True, "status": target_status, "message": f"Status dikembalikan ke {target_status}"})
            return jsonify({"success": True, "status": act.get("status") if act else target_status})
            
        elif action == "upload_photo":
            image_data = request.form.get("live_photo_data")
            if image_data:
                upload_root = os.path.join(monev_bos_bp.root_path, "..", "static", "uploads")
                relative_dir = f"monev_bos/{report_id}/{activity_id}/live_photo"
                db_path, error = _save_camera_photo(image_data, upload_root, relative_dir)
                if error:
                    if request.headers.get("X-Requested-With") == "XMLHttpRequest" or request.headers.get("Accept") == "application/json":
                        return jsonify({"success": False, "message": error}), 400
                    flash(error, "danger")
                    return redirect(url_for("monev_bos.staff_audit_report", report_id=report_id))

                absolute_path = _absolute_dashboard_file_path(db_path)
                try:
                    photo_id = queries.add_activity_doc(
                        activity_id,
                        "live_photo",
                        db_path,
                        os.path.getsize(absolute_path),
                        user["id"],
                    )
                except Exception:
                    try:
                        os.remove(absolute_path)
                    except OSError:
                        pass
                    raise
                queries.add_audit_log(report_id, activity_id, user["id"], "UPLOAD_PHOTO", "Mengambil foto live lapangan")
                _record_monev_admin_action(
                    "UPLOAD_PHOTO",
                    "MONEV_ACTIVITY",
                    target_id=activity_id,
                    target_name=act.get("activity_name"),
                    metadata={"report_id": report_id},
                )
                if request.headers.get("X-Requested-With") == "XMLHttpRequest" or request.headers.get("Accept") == "application/json":
                    return jsonify({
                        "success": True,
                        "activity_id": activity_id,
                        "photo_id": photo_id,
                        "photo_url": f"/{db_path}",
                        "message": "Foto live staff berhasil disimpan.",
                    })
                flash("Foto live berhasil disimpan.", "success")

    except Exception as e:
        import logging
        logging.error(f"Error processing staff verification activity {activity_id}: {e}", exc_info=True)
        if request.headers.get("X-Requested-With") == "XMLHttpRequest" or request.headers.get("Accept") == "application/json":
            return jsonify({"success": False, "message": f"Terjadi kesalahan pada server saat memproses verifikasi: {str(e)}"}), 500
        flash("Terjadi kesalahan pada server saat memproses verifikasi kegiatan.", "danger")

    return redirect(url_for("monev_bos.staff_audit_report", report_id=report_id))


# --- VENDOR MANAGEMENT ROUTES ---

@monev_bos_bp.route("/sekolah/vendors", methods=["GET", "POST"])
@role_required("sekolah")
def sekolah_vendors():
    user = current_user()
    school_id = user["id"]

    def render_vendor_page(duplicate_warning=None):
        search_query = request.args.get("q", "").strip()
        status_filter = request.args.get("status", "")
        vendors = queries.list_school_vendors(
            school_id,
            status_filter if status_filter in ["pending", "verified", "rejected"] else None,
            search_query=search_query,
        )
        return render_template(
            "monev_bos/sekolah/vendors.html",
            vendors=vendors,
            status_filter=status_filter,
            search_query=search_query,
            master_banks=queries.get_master_banks(),
            duplicate_warning=duplicate_warning,
        )

    if request.method == "POST":
        action = request.form.get("action", "")
        if action in ("create_vendor", "update_vendor"):
            name = (request.form.get("name") or "").strip()
            npwp = (request.form.get("npwp") or "").strip()
            phone = (request.form.get("phone") or "").strip()
            address = (request.form.get("address") or "").strip()
            owner_name = (request.form.get("owner_name") or "").strip()
            bank_name = (request.form.get("bank_name") or "").strip()
            bank_account_type = (request.form.get("bank_account_type") or "rekening").strip().lower()
            if bank_account_type not in ("rekening", "va"):
                bank_account_type = "rekening"
            bank_account = "" if bank_account_type == "va" else (request.form.get("bank_account") or "").strip()
            vendor_type = request.form.get("vendor_type", "vendor")
            if vendor_type not in ("vendor", "narsum"):
                vendor_type = "vendor"

            if not name:
                flash("Nama toko / vendor wajib diisi.", "warning")
            else:
                data = {
                    "name": name,
                    "npwp": npwp,
                    "phone": phone,
                    "address": address,
                    "owner_name": owner_name,
                    "bank_name": bank_name,
                    "bank_account_type": bank_account_type,
                    "bank_account": bank_account,
                    "vendor_type": vendor_type,
                }
                type_label = "Narasumber/Instruktur" if vendor_type == "narsum" else "Vendor"
                vendor_id = request.form.get("vendor_id", type=int) if action == "update_vendor" else None
                duplicate_matches = queries.find_vendor_duplicate_matches_for_data(
                    data,
                    exclude_vendor_id=vendor_id,
                )
                duplicate_confirmation = " ".join(
                    (request.form.get("duplicate_confirmation") or "").casefold().split()
                )
                if duplicate_matches and duplicate_confirmation != "vendor berbeda":
                    return render_vendor_page({
                        "action": action,
                        "vendor_id": vendor_id,
                        "data": data,
                        "matches": duplicate_matches,
                    })
                if action == "create_vendor":
                    saved = queries.create_vendor(school_id, data)
                else:
                    saved = vendor_id and queries.update_pending_vendor(vendor_id, school_id, data)
                if saved and action == "create_vendor":
                    flash(f"{type_label} '{name}' berhasil didaftarkan dan menunggu verifikasi admin/staff.", "success")
                elif saved:
                    flash(f"Data {type_label.lower()} '{name}' berhasil diperbarui.", "success")
                else:
                    flash(f"Gagal menyimpan {type_label.lower()}. Data mungkin sudah diverifikasi atau bukan milik sekolah Anda.", "danger")
            return redirect(url_for("monev_bos.sekolah_vendors"))

        elif action == "delete_vendor":
            vendor_id = request.form.get("vendor_id", type=int)
            if vendor_id and queries.delete_vendor(vendor_id, school_id):
                flash("Pendaftaran vendor berhasil dihapus.", "success")
            else:
                flash("Gagal menghapus vendor.", "danger")
            return redirect(url_for("monev_bos.sekolah_vendors"))

    return render_vendor_page()


@monev_bos_bp.route("/admin/vendors", methods=["GET", "POST"])
@role_required("admin", "staff")
def admin_vendors():
    user = current_user()

    def filtered_redirect():
        return redirect(url_for(
            "monev_bos.admin_vendors",
            status=request.args.get("status", ""),
            q=request.args.get("q", ""),
            vendor_type=request.args.get("vendor_type", "vendor"),
            school_id=request.args.get("school_id", ""),
            scan=request.args.get("scan", ""),
        ))

    if request.method == "POST":
        action = request.form.get("action", "")
        vendor_id = request.form.get("vendor_id", type=int)

        if action == "verify_vendor" and vendor_id:
            verified_duplicate_matches = [
                match for match in queries.find_vendor_duplicate_matches(vendor_id)
                if match.get("status") == "verified"
            ]
            if verified_duplicate_matches:
                flash(
                    "Verifikasi diblokir karena data vendor/narasumber yang sama sudah terverifikasi. Gunakan data yang sudah ada atau tolak pengajuan duplikat ini.",
                    "danger",
                )
                return filtered_redirect()

            checklist_keys = ("identity", "npwp", "phone", "address", "owner", "bank")
            verification_checklist = {
                key: request.form.get(f"check_{key}") == "sesuai"
                for key in checklist_keys
            }
            missing_checks = [key for key, checked in verification_checklist.items() if not checked]
            if missing_checks:
                flash(
                    "Semua kolom checklist pemeriksaan harus ditandai Sesuai sebelum vendor/narasumber dapat diverifikasi.",
                    "warning",
                )
                return filtered_redirect()

            review_notes = (request.form.get("review_notes") or "").strip() or None
            if queries.update_vendor_status(
                vendor_id,
                "verified",
                user["id"],
                verification_notes=None,
                verification_checklist=verification_checklist,
                review_notes=review_notes,
            ):
                v_obj = queries.get_vendor_by_id(vendor_id)
                v_name = (
                    (v_obj.get("owner_name") or v_obj.get("name"))
                    if v_obj and v_obj.get("vendor_type") == "narsum"
                    else (v_obj.get("name") if v_obj else "Vendor")
                )
                type_label = "Narasumber" if v_obj and v_obj.get("vendor_type") == "narsum" else "Vendor"
                verification_metadata = {
                    "verification_checklist": verification_checklist,
                    "has_review_notes": bool(review_notes),
                }
                _record_monev_admin_action(
                    "VERIFY_APPROVE",
                    "MONEV_VENDOR",
                    target_id=vendor_id,
                    target_name=v_name,
                    metadata={**verification_metadata, "review_notes": review_notes},
                    allow_staff=True,
                )
                flash(f"{type_label} '{v_name}' berhasil diverifikasi dan disetujui.", "success")
            else:
                flash("Gagal memverifikasi data vendor/narasumber.", "danger")
            return filtered_redirect()

        elif action == "reject_vendor" and vendor_id:
            reason = (request.form.get("rejection_reason") or "").strip()
            if not reason:
                flash("Alasan penolakan vendor wajib diisi.", "warning")
            else:
                if queries.update_vendor_status(vendor_id, "rejected", user["id"], rejection_reason=reason):
                    v_obj = queries.get_vendor_by_id(vendor_id)
                    _record_monev_admin_action(
                        "VERIFY_REJECT",
                        "MONEV_VENDOR",
                        target_id=vendor_id,
                        target_name=v_obj.get("name") if v_obj else None,
                        metadata={"has_rejection_reason": True, "rejection_reason": reason},
                        allow_staff=True,
                    )
                    type_label = "Narasumber" if v_obj and v_obj.get("vendor_type") == "narsum" else "Vendor"
                    flash(f"{type_label} berhasil ditolak.", "info")
                else:
                    flash("Gagal menolak data vendor/narasumber.", "danger")
            return filtered_redirect()

        elif action == "update_master_banks":
            if user.get("role") != "admin":
                flash("Hanya Admin yang berhak mengubah daftar pilihan bank.", "danger")
                return filtered_redirect()

            bank_lines = request.form.get("bank_list", "").splitlines()
            bank_list = [b.strip() for b in bank_lines if b.strip()]
            if queries.save_master_banks(bank_list, user["id"]):
                _record_monev_admin_action(
                    "UPDATE",
                    "MONEV_MASTER_BANKS",
                    target_name="Daftar Bank Vendor",
                    metadata={"bank_count": len(bank_list)},
                )
                flash("Daftar master pilihan bank berhasil diperbarui.", "success")
            else:
                flash("Gagal memperbarui daftar bank.", "danger")
            return filtered_redirect()

    search_query = request.args.get("q", "").strip()
    scan_mode = request.args.get("scan", "")
    if scan_mode not in {"verified_duplicates", "verified_incomplete"}:
        scan_mode = ""
    duplicate_scan = scan_mode == "verified_duplicates"
    incomplete_scan = scan_mode == "verified_incomplete"
    status_filter = "verified" if scan_mode else request.args.get("status", "")
    vendor_type_filter = request.args.get("vendor_type", "vendor")
    if vendor_type_filter not in ["vendor", "narsum"]:
        vendor_type_filter = "vendor"
    school_id_filter = request.args.get("school_id", type=int)
    vendors = queries.list_all_vendors_for_admin(
        status_filter if status_filter in ["pending", "verified", "rejected"] else None,
        search_query=search_query,
        vendor_type_filter=vendor_type_filter,
        school_id_filter=school_id_filter,
    )
    queries.attach_vendor_duplicate_matches(vendors)
    queries.attach_vendor_action_history(vendors)
    if duplicate_scan:
        vendors = queries.filter_verified_duplicate_vendors(vendors)
    elif incomplete_scan:
        vendors = queries.filter_verified_incomplete_vendors(vendors)
    else:
        queries.attach_vendor_missing_fields(vendors)
    vendor_schools = queries.list_vendor_schools_for_admin()
    master_banks = queries.get_master_banks()
    return render_template(
        "monev_bos/admin/admin_vendors.html",
        vendors=vendors,
        status_filter=status_filter,
        search_query=search_query,
        vendor_type_filter=vendor_type_filter,
        school_id_filter=school_id_filter,
        vendor_schools=vendor_schools,
        master_banks=master_banks,
        scan_mode=scan_mode,
        duplicate_scan=duplicate_scan,
        incomplete_scan=incomplete_scan,
    )
